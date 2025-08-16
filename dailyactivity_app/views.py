from django.shortcuts import render, redirect, get_object_or_404
from dailyactivity_app.models import Shift, Location, Category, Status, Machinemechanical, Machineelectrical, Machineutility, MechanicalData, ElectricalData, UtilityData, PICMechanical, PICElectrical, PICUtility, TabelMain, MechanicalData2, PICMechanical2, UtilityData2, PICUtility2, ItData, PICIt, LaporanData, LaporanDataPIC, PICLaporan, PICLembur, LaporanMechanicalData, PICLaporanMechanical, PICLemburMechanical, ScheduleMechanicalData, LaporanMechanicalDataPIC, LemburMechanicalDataPIC, Project, ProjectIssue, LemburDataPIC, LaporanUtility, LaporanPekerjaan, DetailPekerjaan
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from dailyactivity_app.forms import MechanicalDataForm, ElectricalDataForm, UtilityDataForm, PICMechanicalForm, PICElectricalForm, PICUtilityForm, MechanicalData2Form, PICMechanical2Form, UtilityData2Form, PICUtility2Form, ItDataForm, PICItForm, LaporanDataForm, PICLaporanForm, PICLemburForm, LaporanMechanicalDataForm, PICLaporanMechanicalForm, PICLemburMechanicalForm, ScheduleMechanicalDataForm, IssueForm
from django.db.models import DateField, Count
from django.db.models.functions import ExtractMonth, ExtractDay, Trunc
from django.utils.dateparse import parse_date
import pandas as pd
from collections import defaultdict
from django.http import JsonResponse, HttpResponseNotFound, HttpResponse
from django.db import connections
from django.db.models.functions import TruncDay
from django.db.models.functions import TruncDate
import logging
# import datetime
from django.db.models.functions import TruncMonth
from collections import defaultdict
import calendar
import openpyxl
from django.db import transaction
from openpyxl.styles import Alignment, Font, Border, Side
from django.db.models import Q, F
import csv
import xlwt
from django.db.models.functions import Extract
from datetime import datetime, timedelta
from django.utils import timezone
from hashlib import md5


@login_required
def get_laporan_data_ajax(request, tanggal):
    """
    API endpoint untuk mengambil data laporan terbaru via AJAX
    """
    try:
        tanggal_parsed = parse_date(tanggal)
        if not tanggal_parsed:
            return JsonResponse({'error': 'Invalid date format'}, status=400)
        
        # Ambil data laporan (sama seperti di data_laporan view)
        laporan_data = LaporanData.objects.filter(
            tanggal=tanggal_parsed
        ).order_by('shift', 'user', 'created_at')
        
        # Group laporan seperti di view utama
        grouped_laporan = {}
        
        for laporan in laporan_data:
            time_key = laporan.created_at.replace(second=0, microsecond=0)
            minute = (time_key.minute // 5) * 5
            time_key = time_key.replace(minute=minute)
            
            key = f"{laporan.shift.id}_{laporan.user.id}_{time_key.strftime('%Y%m%d_%H%M')}"
            
            if key not in grouped_laporan:
                grouped_laporan[key] = {
                    'meta': {
                        'tanggal': laporan.tanggal.strftime('%Y-%m-%d'),
                        'shift': {
                            'id': laporan.shift.id,
                            'name': laporan.shift.name
                        },
                        'user': {
                            'id': laporan.user.id,
                            'username': laporan.user.username,
                            'full_name': laporan.user.get_full_name() or laporan.user.username
                        },
                        'catatan': laporan.catatan or '',
                        'image_url': laporan.image.url if laporan.image else None,
                        'pic': [{'id': pic.id, 'name': pic.name} for pic in laporan.pic.all()],
                        'piclembur': [{'id': pic.id, 'name': pic.name} for pic in laporan.piclembur.all()],
                        'created_at': laporan.created_at.strftime('%d %B %Y %H:%i'),
                    },
                    'pekerjaan_list': []
                }
            
            grouped_laporan[key]['pekerjaan_list'].append({
                'id': laporan.id,
                'masalah': laporan.masalah,
                'jenis_pekerjaan': laporan.jenis_pekerjaan or 'Tidak ditentukan',
                'lama_pekerjaan': laporan.lama_pekerjaan or 'Tidak ditentukan',
                'pic_pekerjaan': laporan.pic_pekerjaan or 'Tidak ditentukan',
            })
        
        return JsonResponse({
            'grouped_laporan': grouped_laporan,
            'total_groups': len(grouped_laporan),
            'current_time': timezone.now().timestamp()
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'Error fetching data: {str(e)}'
        }, status=500)

@login_required
def check_laporan_updates(request, tanggal):
    """
    API endpoint untuk mengecek apakah ada perubahan data pada tanggal tertentu
    """
    try:
        tanggal_parsed = parse_date(tanggal)
        if not tanggal_parsed:
            return JsonResponse({'error': 'Invalid date format'}, status=400)
        
        # Ambil parameter last_check_time dari request
        last_check_str = request.GET.get('last_check')
        last_check_time = None
        
        if last_check_str:
            try:
                # Parse timestamp dari client
                last_check_time = timezone.datetime.fromtimestamp(
                    float(last_check_str), tz=timezone.get_current_timezone()
                )
            except (ValueError, TypeError):
                pass
        
        # Ambil data laporan terbaru
        laporan_data = LaporanData.objects.filter(
            tanggal=tanggal_parsed
        ).order_by('shift', 'user', 'created_at')
        
        # Hitung hash dari data untuk mendeteksi perubahan
        data_string = ""
        total_records = 0
        latest_update = None
        
        for laporan in laporan_data:
            total_records += 1
            # Gunakan updated_at jika tersedia, otherwise gunakan created_at
            update_time = getattr(laporan, 'updated_at', None) or laporan.created_at
            
            data_string += f"{laporan.id}_{laporan.masalah}_{laporan.jenis_pekerjaan}_"
            data_string += f"{laporan.lama_pekerjaan}_{laporan.pic_pekerjaan}_"
            data_string += f"{laporan.catatan}_{update_time.timestamp()}_"
            
            # Track latest update time
            if not latest_update or update_time > latest_update:
                latest_update = update_time
        
        # Buat hash dari data
        current_hash = md5(data_string.encode()).hexdigest()
        
        # Cek apakah ada perubahan sejak last check
        has_changes = False
        if last_check_time and latest_update:
            has_changes = latest_update > last_check_time
        elif not last_check_time:
            has_changes = True  # First time check
        
        response_data = {
            'has_changes': has_changes,
            'total_records': total_records,
            'data_hash': current_hash,
            'last_update': latest_update.timestamp() if latest_update else None,
            'current_time': timezone.now().timestamp(),
            'message': 'Data berhasil dicek'
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'error': f'Error checking updates: {str(e)}',
            'has_changes': False
        }, status=500)

# logger = logging.getLogger(__name__)
@login_required
def dashboard_index(request):
    # Ambil semua data dari model MechanicalData
    mechanical_data = MechanicalData.objects.all()
    electrical_data = ElectricalData.objects.all()
    utility_data = UtilityData.objects.all()  

    # Mengelompokkan data Mechanical berdasarkan lokasi dan mesin
    grouped_mechanical_data = defaultdict(lambda: {"total_masalah": 0, "dates": []})
    for item in mechanical_data:
        key = (item.location.id, item.machine.id)
        grouped_mechanical_data[key]["total_masalah"] += 1
        if item.tanggal not in grouped_mechanical_data[key]["dates"]:
            grouped_mechanical_data[key]["dates"].append(item.tanggal)

    # Mempersiapkan data untuk grafik Mechanical
    labels_mechanical = []
    total_masalah_mechanical = []

    for (location_id, machine_id), values in grouped_mechanical_data.items():
        location_name = Location.objects.get(id=location_id).name
        machine_name = Machinemechanical.objects.get(id=machine_id).name
        labels_mechanical.append(f"{location_name} - {machine_name} | Tanggal: {', '.join(map(str, values['dates']))}")
        total_masalah_mechanical.append(values["total_masalah"])

    # Mengelompokkan data Electrical berdasarkan lokasi dan mesin
    grouped_electrical_data = defaultdict(lambda: {"total_masalah": 0, "dates": []})
    for item in electrical_data:
        key = (item.location.id, item.machine.id)
        grouped_electrical_data[key]["total_masalah"] += 1
        if item.tanggal not in grouped_electrical_data[key]["dates"]:
            grouped_electrical_data[key]["dates"].append(item.tanggal)
    # Mempersiapkan data untuk grafik Electrical
    labels_electrical = []
    total_masalah_electrical = []

    for (location_id, machine_id), values in grouped_electrical_data.items():
        location_name = Location.objects.get(id=location_id).name
        machine_name = Machineelectrical.objects.get(id=machine_id).name
        labels_electrical.append(f"{location_name} - {machine_name} | Tanggal: {', '.join(map(str, values['dates']))}")
        total_masalah_electrical.append(values["total_masalah"])
    # Mengelompokkan data Utility berdasarkan lokasi dan mesin
    grouped_utility_data = defaultdict(lambda: {"total_masalah": 0, "dates": []})
    for item in utility_data:
        key = (item.location.id, item.machine.id)
        grouped_utility_data[key]["total_masalah"] += 1
        if item.tanggal not in grouped_utility_data[key]["dates"]:
            grouped_utility_data[key]["dates"].append(item.tanggal)
    # Mempersiapkan data untuk grafik Utility
    labels_utility = []
    total_masalah_utility = []

    for (location_id, machine_id), values in grouped_utility_data.items():
        location_name = Location.objects.get(id=location_id).name
        machine_name = Machineutility.objects.get(id=machine_id).name
        labels_utility.append(f"{location_name} - {machine_name} | Tanggal: {', '.join(map(str, values['dates']))}")
        total_masalah_utility.append(values["total_masalah"])
    # Menyusun konteks untuk template
    context = {
        'labels_mechanical': labels_mechanical,
        'total_masalah_mechanical': total_masalah_mechanical,
        'labels_electrical': labels_electrical,
        'total_masalah_electrical': total_masalah_electrical,
        'labels_utility': labels_utility,
        'total_masalah_utility': total_masalah_utility,
    }
    
    return render(request, 'dailyactivity_app/dashboard_index.html', context)

@login_required
def upload_machinemechanical_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        try:
            # Baca file Excel
            df = pd.read_excel(excel_file)
            # Pastikan kolom yang diperlukan ada di file Excel
            required_columns = ['name']
            if not all(column in df.columns for column in required_columns):
                messages.error(request, 'Excel file does not have the required columns.')
                return redirect('dailyactivity_app:machinemechanical_index')
            # Loop melalui setiap baris data di Excel dan simpan ke database
            for _, row in df.iterrows():
                Machinemechanical.objects.create(
                    name=row['name']
                )
            messages.success(request, 'Data uploaded successfully!')
        except Exception as e:
            messages.error(request, f'An error occurred: {e}')

    return redirect('dailyactivity_app:machinemechanical_index')

@login_required
def upload_location_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        try:
            # Baca file Excel
            df = pd.read_excel(excel_file)
            # Pastikan kolom yang diperlukan ada di file Excel
            required_columns = ['name']
            if not all(column in df.columns for column in required_columns):
                messages.error(request, 'Excel file does not have the required columns.')
                return redirect('dailyactivity_app:location_index')
            # Loop melalui setiap baris data di Excel dan simpan ke database
            for _, row in df.iterrows():
                Machinemechanical.objects.create(
                    name=row['name']
                )
            messages.success(request, 'Data uploaded successfully!')
        except Exception as e:
            messages.error(request, f'An error occurred: {e}')
    return redirect('dailyactivity_app:location_index')

@login_required
def lookup_data_table(request):
    nomor_wo = request.GET.get('nomor_wo')
    data = []

    if nomor_wo:
        with connections['DB_Maintenance'].cursor() as cursor:
            cursor.execute("""
                SELECT [tgl_his]
                           ,[line]
      ,[akar_masalah]
      ,[mesin]
      ,[nomer]
      ,[seksi]
      ,[pekerjaan]
      ,[number_wo]
      ,[deskripsi_perbaikan]
      ,[penyebab]
      ,[tindakan_perbaikan]
      ,[tindakan_pencegahan]
      ,[status_pekerjaan]
                FROM [DB_Maintenance].[dbo].[view_main]
                WHERE number_wo LIKE %s
            """, [f"%{nomor_wo}%"])

            rows = cursor.fetchall()
            for row in rows:
                data.append({
                    'tgl_his': row[0],
                    'line': row[1],
                    'akar_masalah': row[2],
                    'mesin': row[3],
                    'nomer': row[4],
                    'seksi': row[5],
                    'pekerjaan': row[6],
                    'number_wo': row[7],
                    'deskripsi_perbaikan': row[8],
                    'penyebab': row[9],
                    'tindakan_perbaikan': row[10],
                    'tindakan_pencegahan': row[11],
                    'status_pekerjaan': row[12],
                })

    return JsonResponse({'data': data})

# Update untuk mechanical_index di views.py

@login_required 
def mechanical_index(request):
    shifts = Shift.objects.all()
    locations = Location.objects.all()
    machines = Machinemechanical.objects.all()
    categories = Category.objects.all()
    status = Status.objects.all()
    pic_mechanical = PICMechanical2.objects.all()

    # Ambil hanya 30 nomor WO paling terbaru
    nomor_wo_list = []
    with connections['DB_Maintenance'].cursor() as cursor:
        cursor.execute("""
            SELECT TOP 30 number_wo
            FROM dbo.view_main
            WHERE id_section = 4
            AND YEAR(tgl_his) BETWEEN 2024 AND 2025
            ORDER BY history_id DESC
        """)
        nomor_wo_list = [row[0] for row in cursor.fetchall()]

    # Default values
    deskripsi_perbaikan = None
    tgl_his = None
    penyebab = None
    line = None
    mesin = None
    nomer = None
    pekerjaan = None
    status_pekerjaan = None
    tindakan_perbaikan = None

    # Tangani Permintaan AJAX untuk nomor WO
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.method == 'GET':
        nomor_wo_selected = request.GET.get('nomor_wo')
        
        # Prepare response data
        response_data = {
            'deskripsi_perbaikan': None,
            'tgl_his': None,
            'penyebab': None,
            'line': None,
            'mesin': None,
            'nomer': None,
            'pekerjaan': None,
            'status_pekerjaan': None,
            'tindakan_perbaikan': None,
            'location_id': None,
            'machine_id': None
        }
        
        with connections['DB_Maintenance'].cursor() as cursor:
            cursor.execute("""
                SELECT deskripsi_perbaikan, tgl_his, penyebab, line, mesin, nomer, 
                       pekerjaan, status_pekerjaan, tindakan_perbaikan
                FROM dbo.view_main
                WHERE number_wo = %s
            """, [nomor_wo_selected])
            row = cursor.fetchone()
            
            if row:
                (deskripsi_perbaikan, tgl_his, penyebab, line, mesin, 
                 nomer, pekerjaan, status_pekerjaan, tindakan_perbaikan) = row
                
                response_data.update({
                    'deskripsi_perbaikan': deskripsi_perbaikan,
                    'tgl_his': tgl_his,
                    'penyebab': penyebab,
                    'line': line,
                    'mesin': mesin,
                    'nomer': nomer,
                    'pekerjaan': pekerjaan,
                    'status_pekerjaan': status_pekerjaan,
                    'tindakan_perbaikan': tindakan_perbaikan
                })
                
                # TAMBAHAN: Cari location_id berdasarkan nama line
                if line:
                    try:
                        location_obj = Location.objects.filter(name__icontains=line).first()
                        if location_obj:
                            response_data['location_id'] = location_obj.id
                    except:
                        pass
                
                # TAMBAHAN: Cari machine_id berdasarkan nama mesin dan nomor
                if mesin:
                    try:
                        machine_query = Machinemechanical.objects.filter(name__icontains=mesin)
                        if nomer:
                            machine_query = machine_query.filter(nomor__icontains=nomer)
                        
                        machine_obj = machine_query.first()
                        if machine_obj:
                            response_data['machine_id'] = machine_obj.id
                    except:
                        pass

        return JsonResponse(response_data)

    # Context untuk template
    context = {
        'shifts': shifts,
        'locations': locations,
        'machines': machines,
        'categories': categories,
        'status': status,
        'pic_mechanical': pic_mechanical,
        'nomor_wo_list': nomor_wo_list,
        'deskripsi_perbaikan': deskripsi_perbaikan,
        'tgl_his': tgl_his,
        'penyebab': penyebab,
        'line': line,
        'mesin': mesin,
        'nomer': nomer,
        'pekerjaan': pekerjaan,
        'status_pekerjaan': status_pekerjaan,
        'tindakan_perbaikan': tindakan_perbaikan,
    }
    return render(request, 'dailyactivity_app/mechanical_index.html', context)


@login_required  
def mechanical_submit(request):
    if request.method == 'POST':
        try:
            # Function to safely convert empty string to None, then to int
            def safe_int_or_none(value):
                if not value or value == '' or value == 'None':
                    return None
                try:
                    return int(value)
                except (ValueError, TypeError):
                    return None
            
            # Function to safely get string value
            def safe_string(value):
                return value.strip() if value and value.strip() else None

            # Ambil dan clean semua data dari formulir
            tanggal = request.POST.get('tanggal')
            jam = request.POST.get('jam')
            tgl_his = request.POST.get('tgl_his')
            
            # Convert ALL ID fields safely
            shift_id = safe_int_or_none(request.POST.get('shift'))
            location_id = safe_int_or_none(request.POST.get('location'))
            machine_id = safe_int_or_none(request.POST.get('machine'))
            category_id = safe_int_or_none(request.POST.get('category'))
            status_id = safe_int_or_none(request.POST.get('status'))
            
            # Text fields
            masalah = request.POST.get('masalah', '').strip()
            penyebab = request.POST.get('penyebab', '').strip()
            tindakan_perbaikan = request.POST.get('tindakan_perbaikan', '').strip()
            nomor_wo = safe_string(request.POST.get('nomor_wo'))
            waktu_pengerjaan = safe_string(request.POST.get('waktu_pengerjaan'))
            machine_number = safe_string(request.POST.get('machine_number'))
            
            # Manual machine inputs
            manual_machine_name = safe_string(request.POST.get('manual_machine_name'))
            manual_machine_number = safe_string(request.POST.get('manual_machine_number'))
            
            # File
            image = request.FILES.get('image')
            
            # PIC IDs - filter empty values
            pic_ids_raw = request.POST.getlist('pic')
            pic_ids = [safe_int_or_none(pic_id) for pic_id in pic_ids_raw if safe_int_or_none(pic_id) is not None]

            print(f"DEBUG - shift_id: {shift_id}, category_id: {category_id}, status_id: {status_id}")
            print(f"DEBUG - location_id: {location_id}, machine_id: {machine_id}")
            print(f"DEBUG - manual_machine_name: {manual_machine_name}")

            # Validasi REQUIRED fields dengan proper error messages
            if not shift_id:
                messages.error(request, 'Shift harus dipilih.')
                return redirect('dailyactivity_app:mechanical_index')
                
            if not category_id:
                messages.error(request, 'Jenis Pekerjaan harus dipilih.')
                return redirect('dailyactivity_app:mechanical_index')
                
            if not status_id:
                messages.error(request, 'Status harus dipilih.')
                return redirect('dailyactivity_app:mechanical_index')

            if not masalah:
                messages.error(request, 'Masalah harus diisi.')
                return redirect('dailyactivity_app:mechanical_index')

            if not penyebab:
                messages.error(request, 'Penyebab harus diisi.')
                return redirect('dailyactivity_app:mechanical_index')

            if not tindakan_perbaikan:
                messages.error(request, 'Tindakan Perbaikan harus diisi.')
                return redirect('dailyactivity_app:mechanical_index')

            # Get required instances
            try:
                shift_instance = Shift.objects.get(id=shift_id)
                category_instance = Category.objects.get(id=category_id)
                status_instance = Status.objects.get(id=status_id)
            except (Shift.DoesNotExist, Category.DoesNotExist, Status.DoesNotExist) as e:
                messages.error(request, f'Data referensi tidak ditemukan: {str(e)}')
                return redirect('dailyactivity_app:mechanical_index')

            # HANDLE LOCATION - ALWAYS CREATE ONE
            location_instance = None
            if location_id:
                try:
                    location_instance = Location.objects.get(id=location_id)
                except Location.DoesNotExist:
                    messages.error(request, 'Location tidak ditemukan.')
                    return redirect('dailyactivity_app:mechanical_index')
            
            # Create default location if none selected
            if not location_instance:
                location_instance, created = Location.objects.get_or_create(
                    name="Unknown Location",
                    defaults={'name': "Unknown Location"}
                )
                print(f"DEBUG - Created/found default location: {location_instance.id}")

            # HANDLE MACHINE - ALWAYS CREATE ONE
            machine_instance = None
            
            # Option 1: Machine selected from dropdown
            if machine_id:
                try:
                    machine_instance = Machinemechanical.objects.get(id=machine_id)
                    print(f"DEBUG - Found machine from dropdown: {machine_instance.id}")
                    
                    # Update machine number if provided and machine doesn't have one
                    if machine_number and not machine_instance.nomor:
                        machine_instance.nomor = machine_number
                        machine_instance.save()
                        
                except Machinemechanical.DoesNotExist:
                    print(f"DEBUG - Machine ID {machine_id} not found")
                    machine_instance = None
            
            # Option 2: Manual machine input
            if not machine_instance and manual_machine_name:
                try:
                    print(f"DEBUG - Trying to create/find manual machine: {manual_machine_name}")
                    
                    # Check if machine already exists
                    machine_instance = Machinemechanical.objects.filter(
                        name__iexact=manual_machine_name
                    ).first()
                    
                    if not machine_instance:
                        # Create new machine
                        machine_instance = Machinemechanical.objects.create(
                            name=manual_machine_name,
                            location=location_instance,
                            nomor=manual_machine_number
                        )
                        print(f"DEBUG - Created new machine: {machine_instance.id}")
                    else:
                        # Update existing machine number if needed
                        if manual_machine_number and not machine_instance.nomor:
                            machine_instance.nomor = manual_machine_number
                            machine_instance.save()
                        print(f"DEBUG - Found existing machine: {machine_instance.id}")
                        
                except Exception as e:
                    print(f"DEBUG - Error creating manual machine: {e}")
                    machine_instance = None
            
            # Option 3: Create default machine if still none
            if not machine_instance:
                try:
                    machine_instance, created = Machinemechanical.objects.get_or_create(
                        name="Unknown Machine",
                        location=location_instance,
                        defaults={
                            'name': "Unknown Machine",
                            'location': location_instance,
                            'nomor': None
                        }
                    )
                    print(f"DEBUG - Created/found default machine: {machine_instance.id}, created: {created}")
                except Exception as e:
                    print(f"DEBUG - Error creating default machine: {e}")
                    # Last resort with timestamp to avoid conflicts
                    import time
                    timestamp = str(int(time.time()))
                    machine_instance = Machinemechanical.objects.create(
                        name=f"Unknown Machine {timestamp}",
                        location=location_instance,
                        nomor=None
                    )
                    print(f"DEBUG - Created timestamped machine: {machine_instance.id}")

            # Pastikan semua instance yang dibutuhkan ada
            if not all([shift_instance, location_instance, machine_instance, category_instance, status_instance]):
                missing = []
                if not shift_instance: missing.append("shift")
                if not location_instance: missing.append("location")
                if not machine_instance: missing.append("machine")
                if not category_instance: missing.append("category") 
                if not status_instance: missing.append("status")
                
                messages.error(request, f'Instance tidak lengkap: {", ".join(missing)}')
                return redirect('dailyactivity_app:mechanical_index')

            # Process jam field
            jam_value = None
            if tgl_his:
                jam_value = tgl_his
            elif jam:
                jam_value = jam

            print(f"DEBUG - About to save with: shift={shift_instance.id}, location={location_instance.id}, machine={machine_instance.id}")

            # SAVE DATA TO DATABASE
            mechanical_data = MechanicalData.objects.create(
                tanggal=tanggal,
                jam=jam_value,
                shift=shift_instance,
                location=location_instance,
                machine=machine_instance,
                category=category_instance,
                status=status_instance,
                user=request.user,  # Use request.user instead of user_id
                masalah=masalah,
                penyebab=penyebab,
                tindakan=tindakan_perbaikan,
                tindakan_perbaikan=tindakan_perbaikan,
                image=image,
                nomor_wo=nomor_wo,
                waktu_pengerjaan=waktu_pengerjaan
            )

            print(f"DEBUG - Mechanical data created with ID: {mechanical_data.id}")

            # Add PICs
            for pic_id in pic_ids:
                try:
                    pic_instance = PICMechanical2.objects.get(id=pic_id)
                    mechanical_data.pic.add(pic_instance)
                    print(f"DEBUG - Added PIC: {pic_id}")
                except PICMechanical2.DoesNotExist:
                    print(f"DEBUG - PIC not found: {pic_id}")
                    continue

            messages.success(request, 'Data berhasil disimpan!')
            return redirect('dailyactivity_app:mechanical_index')

        except Exception as e:
            print(f"ERROR - Exception in mechanical_submit: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
            return redirect('dailyactivity_app:mechanical_index')

    return redirect('dailyactivity_app:mechanical_index')


# @login_required
# def edit_mechanical_data(request, id):
#     # Ambil data yang akan diedit berdasarkan id
#     mechanical_data = get_object_or_404(MechanicalData2, id=id)
#     shifts = Shift.objects.all()
#     status = Status.objects.all()
#     pic_mechanical = PICMechanical2.objects.all()
    
#     if request.method == 'POST':
#         # Memproses form data yang telah diisi ulang
#         form = MechanicalData2Form(request.POST, request.FILES, instance=mechanical_data)
#         if form.is_valid():
#             updated_data = form.save(commit=False)
#             updated_data.user = request.user
#             updated_data.nomor_wo = form.cleaned_data.get('nomor_wo')
#             updated_data.waktu_pengerjaan = form.cleaned_data.get('waktu_pengerjaan')
#             updated_data.line = form.cleaned_data.get('line')
#             updated_data.mesin = form.cleaned_data.get('mesin')
#             updated_data.nomer = form.cleaned_data.get('nomer')
#             updated_data.pekerjaan = form.cleaned_data.get('pekerjaan')
#             updated_data.save()
#             # Update PIC terkait
#             pic_ids = form.cleaned_data.get('pic')
#             updated_data.pic.set(pic_ids)
#             messages.success(request, 'Data berhasil diperbarui!')
#             return redirect('dailyactivity_app:data_mechanical', tanggal=updated_data.tanggal.strftime('%Y-%m-%d'))
#         else:
#             messages.error(request, 'Terjadi kesalahan saat memperbarui data. Periksa kembali isian Anda.')

#     else:
#         form = MechanicalData2Form(instance=mechanical_data)

#     context = {
#         'form': form,
#         'shifts': shifts,
#         'status': status,
#         'pic_mechanical': pic_mechanical,
#         'data': mechanical_data,
#         'tanggal': mechanical_data.tanggal,
#         'jam': mechanical_data.jam,
#         'nomor_wo': mechanical_data.nomor_wo,
#         'waktu_pengerjaan': mechanical_data.waktu_pengerjaan,
#         'line': mechanical_data.line,
#         'mesin': mechanical_data.mesin,
#         'nomer': mechanical_data.nomer,
#         'pekerjaan': mechanical_data.pekerjaan,
#         'pic': mechanical_data.pic.all(),
#     }
#     return render(request, 'dailyactivity_app/edit_mechanical_data.html', context)

@login_required
def get_machines_by_location_mechanical(request, location_id):
    # Mengambil data mesin berdasarkan location_id
    machines = Machinemechanical.objects.filter(location_id=location_id).order_by('name', 'nomor') 
    
    # Mengambil hanya beberapa field yang relevan
    machine_list = list(machines.values('id', 'name', 'location_id', 'nomor'))
    
    # Mengembalikan response dalam format JSON
    return JsonResponse(machine_list, safe=False)
@login_required
def get_machine_number_mechanical(request, machine_id):
    try:
        # Mengambil mesin berdasarkan ID
        machine = Machinemechanical.objects.get(id=machine_id)
        # Mengembalikan nomor mesin
        return JsonResponse({'nomor': machine.nomor})
    except Machinemechanical.DoesNotExist:
        return JsonResponse({'error': 'Machine not found'}, status=404)


# Update juga fungsi untuk tanggal dan data mechanical agar menggunakan MechanicalData
# @login_required
# def tanggal_mechanical(request):
#     # Mengambil tanggal-tanggal unik dan mengurutkan berdasarkan tanggal secara descending
#     dates = MechanicalData.objects.annotate(
#         date=TruncDate('tanggal', output_field=DateField())  # Truncate to date only
#     ).values('date').distinct().order_by('-date')  # Order by date descending

#     context = {
#         'dates': dates,
#     }
#     return render(request, 'dailyactivity_app/tanggal_mechanical.html', context)

# @login_required
# def tanggal_mechanical(request):
#     # Ambil parameter bulan dan tahun dari URL jika ada
#     selected_month = request.GET.get('month')
#     selected_year = request.GET.get('year')
    
#     print(f"DEBUG - Raw parameters: month='{selected_month}', year='{selected_year}'")
    
#     if selected_month and selected_year:
#         try:
#             # Fungsi untuk membersihkan dan konversi nilai
#             def clean_and_convert(value):
#                 if value is None:
#                     return None
                
#                 # Convert to string terlebih dahulu
#                 str_value = str(value).strip()
                
#                 # Jika kosong, return None
#                 if not str_value:
#                     return None
                
#                 # Jika ada decimal (seperti 2.025), ambil bagian integer saja
#                 if '.' in str_value:
#                     # Split dan ambil bagian sebelum decimal
#                     integer_part = str_value.split('.')[0]
#                     return int(integer_part) if integer_part.isdigit() else None
                
#                 # Jika sudah integer/string digit biasa
#                 if str_value.isdigit():
#                     return int(str_value)
                
#                 # Coba konversi float dulu baru ke int (untuk handle kasus lain)
#                 try:
#                     return int(float(str_value))
#                 except (ValueError, TypeError):
#                     return None
            
#             # Konversi parameter
#             selected_month = clean_and_convert(selected_month)
#             selected_year = clean_and_convert(selected_year)
            
#             print(f"DEBUG - Cleaned: month={selected_month}, year={selected_year}")
            
#             # Validasi hasil konversi
#             if selected_month is None or selected_year is None:
#                 messages.error(request, 'Parameter bulan atau tahun tidak dapat dikonversi')
#                 return redirect('dailyactivity_app:tanggal_mechanical')
            
#             # Validasi range
#             if not (1 <= selected_month <= 12):
#                 messages.error(request, f'Bulan harus antara 1-12, diterima: {selected_month}')
#                 return redirect('dailyactivity_app:tanggal_mechanical')
                
#             # Validasi tahun dengan range yang realistis
#             current_year = datetime.now().year
#             if not (2020 <= selected_year <= current_year + 2):
#                 messages.error(request, f'Tahun harus antara 2020-{current_year + 2}, diterima: {selected_year}')
#                 return redirect('dailyactivity_app:tanggal_mechanical')
            
#             # Query data dengan error handling
#             try:
#                 dates = MechanicalData.objects.filter(
#                     tanggal__month=selected_month,
#                     tanggal__year=selected_year
#                 ).annotate(
#                     date=TruncDate('tanggal', output_field=DateField())
#                 ).values('date').distinct().order_by('-date')
                
#                 dates_count = dates.count()
#                 print(f"DEBUG - Query successful, found {dates_count} dates")
                
#             except Exception as query_error:
#                 print(f"DEBUG - Query error: {query_error}")
#                 messages.error(request, f'Error saat mengambil data: {query_error}')
#                 return redirect('dailyactivity_app:tanggal_mechanical')
            
#             # Nama bulan
#             month_names = {
#                 1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
#                 5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
#                 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
#             }
            
#             context = {
#                 'dates': dates,
#                 'selected_month': selected_month,
#                 'selected_year': selected_year,
#                 'selected_month_name': month_names.get(selected_month, f'Bulan {selected_month}'),
#                 'show_dates': True,
#             }
            
#         except Exception as e:
#             print(f"DEBUG - General exception: {e}")
#             messages.error(request, f'Terjadi kesalahan: {str(e)}')
#             return redirect('dailyactivity_app:tanggal_mechanical')
#     else:
#         print("DEBUG - Showing month/year list")
        
#         try:
#             # Cek apakah ada data sama sekali
#             total_count = MechanicalData.objects.count()
#             print(f"DEBUG - Total MechanicalData records: {total_count}")
            
#             if total_count == 0:
#                 messages.info(request, 'Belum ada data mechanical dalam sistem')
#                 context = {
#                     'month_year_data': [],
#                     'show_dates': False,
#                 }
#                 return render(request, 'dailyactivity_app/tanggal_mechanical.html', context)
            
#             # Ambil sample data untuk debugging
#             sample_data = MechanicalData.objects.first()
#             if sample_data and sample_data.tanggal:
#                 print(f"DEBUG - Sample date: {sample_data.tanggal}, type: {type(sample_data.tanggal)}")
            
#             # Ambil semua data dengan tanggal yang valid
#             all_mechanical_data = MechanicalData.objects.filter(
#                 tanggal__isnull=False
#             ).values('tanggal')
            
#             print(f"DEBUG - Records with valid dates: {all_mechanical_data.count()}")
            
#             # Dictionary untuk menyimpan count
#             month_year_dict = {}
            
#             for data in all_mechanical_data:
#                 tanggal = data['tanggal']
#                 if tanggal:
#                     try:
#                         month = tanggal.month
#                         year = tanggal.year
#                         key = f"{year}-{month:02d}"
                        
#                         if key in month_year_dict:
#                             month_year_dict[key]['count'] += 1
#                         else:
#                             month_year_dict[key] = {
#                                 'month': month,
#                                 'year': year,
#                                 'count': 1
#                             }
#                     except Exception as date_error:
#                         print(f"DEBUG - Error processing date {tanggal}: {date_error}")
#                         continue
            
#             print(f"DEBUG - Month/Year dict keys: {list(month_year_dict.keys())}")
            
#             # Konversi ke list
#             month_names = {
#                 1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
#                 5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
#                 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
#             }
            
#             data_list = []
#             for key, data in month_year_dict.items():
#                 month = data['month']
#                 year = data['year']
#                 count = data['count']
                
#                 data_list.append({
#                     'month': month,
#                     'year': year,
#                     'count': count,
#                     'month_name': month_names.get(month, f'Bulan {month}')
#                 })
            
#             # Sorting
#             current_date = datetime.now()
#             current_month = current_date.month
#             current_year = current_date.year
            
#             def get_sort_priority(item):
#                 year = item['year']
#                 month = item['month']
                
#                 if year == current_year:
#                     if month <= current_month:
#                         return (0, current_month - month)
#                     else:
#                         return (1, month)
#                 elif year < current_year:
#                     return (2, -year, -month)
#                 else:
#                     return (3, year, month)
            
#             data_list.sort(key=get_sort_priority)
            
#             print(f"DEBUG - Final data list count: {len(data_list)}")
            
#             context = {
#                 'month_year_data': data_list,
#                 'show_dates': False,
#             }
            
#         except Exception as list_error:
#             print(f"DEBUG - Error creating month/year list: {list_error}")
#             messages.error(request, f'Error saat mengambil daftar bulan/tahun: {list_error}')
#             context = {
#                 'month_year_data': [],
#                 'show_dates': False,
#             }
    
#     return render(request, 'dailyactivity_app/tanggal_mechanical.html', context)
    
@login_required
def tanggal_mechanical(request):
    """
    View untuk menampilkan tanggal-tanggal yang ada data mechanical & laporan mechanical
    Menggabungkan data dari MechanicalData dan LaporanMechanicalData
    """
    # Ambil parameter bulan dan tahun dari URL jika ada
    selected_month = request.GET.get('month')
    selected_year = request.GET.get('year')
    
    print(f"DEBUG - Raw parameters: month='{selected_month}', year='{selected_year}'")
    
    if selected_month and selected_year:
        try:
            # Fungsi untuk membersihkan dan konversi nilai
            def clean_and_convert(value):
                if value is None:
                    return None
                
                str_value = str(value).strip()
                
                if not str_value:
                    return None
                
                if '.' in str_value:
                    integer_part = str_value.split('.')[0]
                    return int(integer_part) if integer_part.isdigit() else None
                
                if str_value.isdigit():
                    return int(str_value)
                
                try:
                    return int(float(str_value))
                except (ValueError, TypeError):
                    return None
            
            # Konversi parameter
            selected_month = clean_and_convert(selected_month)
            selected_year = clean_and_convert(selected_year)
            
            print(f"DEBUG - Cleaned: month={selected_month}, year={selected_year}")
            
            # Validasi hasil konversi
            if selected_month is None or selected_year is None:
                messages.error(request, 'Parameter bulan atau tahun tidak dapat dikonversi')
                return redirect('dailyactivity_app:tanggal_mechanical')
            
            # Validasi range
            if not (1 <= selected_month <= 12):
                messages.error(request, f'Bulan harus antara 1-12, diterima: {selected_month}')
                return redirect('dailyactivity_app:tanggal_mechanical')
                
            current_year = datetime.now().year
            if not (2020 <= selected_year <= current_year + 2):
                messages.error(request, f'Tahun harus antara 2020-{current_year + 2}, diterima: {selected_year}')
                return redirect('dailyactivity_app:tanggal_mechanical')
            
            # Query data dari MechanicalData
            try:
                mechanical_dates = MechanicalData.objects.filter(
                    tanggal__month=selected_month,
                    tanggal__year=selected_year
                ).annotate(
                    date=TruncDate('tanggal', output_field=DateField())
                ).values('date').distinct()
                
                # Query data dari LaporanMechanicalData
                laporan_dates = LaporanMechanicalData.objects.filter(
                    tanggal__month=selected_month,
                    tanggal__year=selected_year
                ).annotate(
                    date=TruncDate('tanggal', output_field=DateField())
                ).values('date').distinct()
                
                print(f"DEBUG - MechanicalData dates: {mechanical_dates.count()}")
                print(f"DEBUG - LaporanMechanicalData dates: {laporan_dates.count()}")
                
                # Gabungkan tanggal dan hitung jumlah data per tanggal
                combined_dates = {}
                
                # Proses MechanicalData
                for date_obj in mechanical_dates:
                    date_key = date_obj['date']
                    if date_key not in combined_dates:
                        combined_dates[date_key] = {
                            'date': date_key,
                            'mechanical_count': 0,
                            'laporan_count': 0,
                            'total_count': 0
                        }
                    
                    # Hitung jumlah data mechanical untuk tanggal ini
                    mechanical_count = MechanicalData.objects.filter(
                        tanggal=date_key
                    ).count()
                    combined_dates[date_key]['mechanical_count'] = mechanical_count
                
                # Proses LaporanMechanicalData
                for date_obj in laporan_dates:
                    date_key = date_obj['date']
                    if date_key not in combined_dates:
                        combined_dates[date_key] = {
                            'date': date_key,
                            'mechanical_count': 0,
                            'laporan_count': 0,
                            'total_count': 0
                        }
                    
                    # Hitung jumlah data laporan untuk tanggal ini
                    laporan_count = LaporanMechanicalData.objects.filter(
                        tanggal=date_key
                    ).count()
                    combined_dates[date_key]['laporan_count'] = laporan_count
                
                # Hitung total count dan convert ke list
                dates_list = []
                for date_key, data in combined_dates.items():
                    data['total_count'] = data['mechanical_count'] + data['laporan_count']
                    dates_list.append(data)
                
                # Sort berdasarkan tanggal descending
                dates_list.sort(key=lambda x: x['date'], reverse=True)
                
                total_dates = len(dates_list)
                print(f"DEBUG - Combined dates: {total_dates}")
                
            except Exception as query_error:
                print(f"DEBUG - Query error: {query_error}")
                messages.error(request, f'Error saat mengambil data: {query_error}')
                return redirect('dailyactivity_app:tanggal_mechanical')
            
            # Nama bulan
            month_names = {
                1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
                5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
                9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
            }
            
            context = {
                'dates': dates_list,
                'selected_month': selected_month,
                'selected_year': selected_year,
                'selected_month_name': month_names.get(selected_month, f'Bulan {selected_month}'),
                'show_dates': True,
                'total_mechanical': sum(item['mechanical_count'] for item in dates_list),
                'total_laporan': sum(item['laporan_count'] for item in dates_list),
                'total_combined': sum(item['total_count'] for item in dates_list),
            }
            
        except Exception as e:
            print(f"DEBUG - General exception: {e}")
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
            return redirect('dailyactivity_app:tanggal_mechanical')
    else:
        print("DEBUG - Showing month/year list")
        
        try:
            # Cek data dari kedua model
            mechanical_count = MechanicalData.objects.count()
            laporan_count = LaporanMechanicalData.objects.count()
            total_count = mechanical_count + laporan_count
            
            print(f"DEBUG - MechanicalData records: {mechanical_count}")
            print(f"DEBUG - LaporanMechanicalData records: {laporan_count}")
            print(f"DEBUG - Total records: {total_count}")
            
            if total_count == 0:
                messages.info(request, 'Belum ada data mechanical atau laporan dalam sistem')
                context = {
                    'month_year_data': [],
                    'show_dates': False,
                }
                return render(request, 'dailyactivity_app/tanggal_mechanical.html', context)
            
            # Dictionary untuk menyimpan count gabungan
            month_year_dict = {}
            
            # Proses data MechanicalData
            mechanical_data = MechanicalData.objects.filter(
                tanggal__isnull=False
            ).values('tanggal')
            
            for data in mechanical_data:
                tanggal = data['tanggal']
                if tanggal:
                    try:
                        month = tanggal.month
                        year = tanggal.year
                        key = f"{year}-{month:02d}"
                        
                        if key in month_year_dict:
                            month_year_dict[key]['mechanical_count'] += 1
                        else:
                            month_year_dict[key] = {
                                'month': month,
                                'year': year,
                                'mechanical_count': 1,
                                'laporan_count': 0,
                                'total_count': 1
                            }
                    except Exception as date_error:
                        print(f"DEBUG - Error processing mechanical date {tanggal}: {date_error}")
                        continue
            
            # Proses data LaporanMechanicalData
            laporan_data = LaporanMechanicalData.objects.filter(
                tanggal__isnull=False
            ).values('tanggal')
            
            for data in laporan_data:
                tanggal = data['tanggal']
                if tanggal:
                    try:
                        month = tanggal.month
                        year = tanggal.year
                        key = f"{year}-{month:02d}"
                        
                        if key in month_year_dict:
                            month_year_dict[key]['laporan_count'] += 1
                            month_year_dict[key]['total_count'] += 1
                        else:
                            month_year_dict[key] = {
                                'month': month,
                                'year': year,
                                'mechanical_count': 0,
                                'laporan_count': 1,
                                'total_count': 1
                            }
                    except Exception as date_error:
                        print(f"DEBUG - Error processing laporan date {tanggal}: {date_error}")
                        continue
            
            print(f"DEBUG - Month/Year dict keys: {list(month_year_dict.keys())}")
            
            # Konversi ke list
            month_names = {
                1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
                5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
                9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
            }
            
            data_list = []
            for key, data in month_year_dict.items():
                month = data['month']
                year = data['year']
                mechanical_count = data['mechanical_count']
                laporan_count = data['laporan_count']
                total_count = data['total_count']
                
                data_list.append({
                    'month': month,
                    'year': year,
                    'mechanical_count': mechanical_count,
                    'laporan_count': laporan_count,
                    'total_count': total_count,
                    'month_name': month_names.get(month, f'Bulan {month}')
                })
            
            # Sorting
            current_date = datetime.now()
            current_month = current_date.month
            current_year = current_date.year
            
            def get_sort_priority(item):
                year = item['year']
                month = item['month']
                
                if year == current_year:
                    if month <= current_month:
                        return (0, current_month - month)
                    else:
                        return (1, month)
                elif year < current_year:
                    return (2, -year, -month)
                else:
                    return (3, year, month)
            
            data_list.sort(key=get_sort_priority)
            
            print(f"DEBUG - Final data list count: {len(data_list)}")
            
            context = {
                'month_year_data': data_list,
                'show_dates': False,
                'total_mechanical_all': mechanical_count,
                'total_laporan_all': laporan_count,
                'total_combined_all': total_count,
            }
            
        except Exception as list_error:
            print(f"DEBUG - Error creating month/year list: {list_error}")
            messages.error(request, f'Error saat mengambil daftar bulan/tahun: {list_error}')
            context = {
                'month_year_data': [],
                'show_dates': False,
            }
    
    return render(request, 'dailyactivity_app/tanggal_mechanical.html', context)
# @login_required
# def data_mechanical(request, tanggal):
#     # Parsing tanggal dari URL
#     tanggal_parsed = parse_date(tanggal)
    
#     # Menyaring data berdasarkan tanggal yang dipilih dengan select_related untuk optimasi
#     mechanical_data = MechanicalData.objects.filter(
#         tanggal=tanggal_parsed
#     ).select_related(
#         'shift', 'location', 'machine', 'category', 'status', 'user'
#     ).prefetch_related(
#         'pic'  # Untuk many-to-many relationship
#     ).order_by('-jam', '-id')  # Order by jam descending, then ID descending
    
#     context = {
#         'mechanical_data': mechanical_data,
#         'selected_date': tanggal_parsed,
#     }
#     return render(request, 'dailyactivity_app/data_mechanical.html', context)

@login_required
def data_mechanical(request, tanggal):
    """
    View buat nampilir data mechanical dan laporan mechanical dalam satu halaman
    Kombinasi data dari MechanicalData dan LaporanMechanicalData
    """
    from datetime import datetime
    
    # Parsing tanggal dari URL
    tanggal_parsed = parse_date(tanggal)
    
    # Ambil data MechanicalData (data mechanical biasa)
    mechanical_data = MechanicalData.objects.filter(
        tanggal=tanggal_parsed
    ).select_related(
        'shift', 'location', 'machine', 'category', 'status', 'user'
    ).prefetch_related(
        'pic'
    ).order_by('-jam', '-id')
    
    # Ambil data LaporanMechanicalData (data laporan mechanical)
    laporan_mechanical_data = LaporanMechanicalData.objects.filter(
        tanggal=tanggal_parsed
    ).select_related(
        'shift', 'user'
    ).prefetch_related(
        'pic',  # untuk PIC laporan
        'piclembur',  # untuk PIC lembur
        'laporan_pekerjaan'  # untuk detail pekerjaan
    ).order_by('-id')
    
    # Gabungin data jadi unified structure buat template
    combined_data = []
    
    # Proses data MechanicalData
    for data in mechanical_data:
        combined_data.append({
            'id': data.id,
            'type': 'mechanical',  # buat identifier
            'tanggal': data.tanggal,
            'jam': data.jam,
            'shift': data.shift,
            'nomor_wo': data.nomor_wo,
            'waktu_pengerjaan': data.waktu_pengerjaan,
            'line': data.location.name if data.location else None,  # ambil dari location.name
            'mesin': data.machine.name if data.machine else None,  # ambil dari machine.name
            'nomer': data.machine.nomor if data.machine and data.machine.nomor else None,  # ambil dari machine.nomor
            'masalah': data.masalah,
            'penyebab': data.penyebab,
            'pekerjaan': None,  # MechanicalData ga punya field pekerjaan
            'status_pekerjaan': None,  # MechanicalData ga punya field status_pekerjaan
            'tindakan_perbaikan': data.tindakan_perbaikan if hasattr(data, 'tindakan_perbaikan') and data.tindakan_perbaikan else data.tindakan,
            'tindakan_pencegahan': data.tindakan_pencegahan if hasattr(data, 'tindakan_pencegahan') and data.tindakan_pencegahan else None,
            'tindakan': data.tindakan,  # field asli dari MechanicalData
            'status': data.status,
            'image': data.image,
            'pic': data.pic.all(),
            'pic_lembur': None,  # mechanical data ga punya pic lembur
            'user': data.user,
            'catatan': None,  # mechanical data ga punya catatan
            'lama_pekerjaan': None,
            'pic_masalah': None,
            'location': data.location,  # tambah field location
            'machine': data.machine,  # tambah field machine
            'category': data.category,  # tambah field category
        })
    
    # Proses data LaporanMechanicalData
    for laporan in laporan_mechanical_data:
        # Ambil detail pekerjaan kalo ada
        detail_pekerjaan = laporan.laporan_pekerjaan.all()
        pekerjaan_list = []
        jenis_pekerjaan_list = []
        
        if detail_pekerjaan.exists():
            for detail in detail_pekerjaan:
                pekerjaan_list.append(detail.masalah)
                jenis_pekerjaan_list.append(detail.jenis_pekerjaan)
        
        # Gabungin masalah utama dengan detail pekerjaan
        masalah_lengkap = laporan.masalah
        if pekerjaan_list:
            masalah_lengkap += " | Detail: " + " • ".join(pekerjaan_list)
        
        combined_data.append({
            'id': laporan.id,
            'type': 'laporan',  # buat identifier
            'tanggal': laporan.tanggal,
            'jam': None,  # laporan ga punya jam spesifik
            'shift': laporan.shift,
            'nomor_wo': None,  # laporan ga punya nomor WO
            'waktu_pengerjaan': laporan.lama_pekerjaan,
            'line': None,
            'mesin': None,
            'nomer': None,
            'masalah': masalah_lengkap,  # masalah + detail pekerjaan
            'penyebab': None,
            'pekerjaan': " • ".join(pekerjaan_list) if pekerjaan_list else None,
            'status_pekerjaan': None,
            'tindakan_perbaikan': None,
            'tindakan_pencegahan': None,
            'tindakan': None,  # laporan ga punya field tindakan
            'status': None,  # laporan ga punya status
            'image': laporan.image,
            'pic': laporan.pic.all(),
            'pic_lembur': laporan.piclembur.all(),
            'user': laporan.user,
            'catatan': laporan.catatan,
            'lama_pekerjaan': laporan.lama_pekerjaan,
            'pic_masalah': laporan.pic_masalah,
            'jenis_pekerjaan': " • ".join(jenis_pekerjaan_list) if jenis_pekerjaan_list else None,
            'location': None,  # laporan ga punya location
            'machine': None,  # laporan ga punya machine
            'category': None,  # laporan ga punya category
        })
    
    # Sort combined data berdasarkan tanggal dan jam (simple sorting)
    def safe_sort_key(item):
        """Helper function untuk sorting yang aman"""
        tanggal = item['tanggal']
        jam = item['jam']
        
        # Convert semua ke string buat sorting, atau set default value
        if jam is None:
            jam_sort = "00:00:00"  # Default time string
        elif isinstance(jam, datetime):
            jam_sort = jam.strftime("%H:%M:%S")
        elif hasattr(jam, 'strftime'):
            jam_sort = jam.strftime("%H:%M:%S")  
        else:
            jam_sort = str(jam) if jam else "00:00:00"
        
        return (tanggal, jam_sort, -item['id'])
    
    try:
        combined_data.sort(key=safe_sort_key, reverse=True)
    except Exception as sort_error:
        # Fallback: sort cuma berdasarkan tanggal dan ID aja
        print(f"Sorting error: {sort_error}, using fallback sort")
        combined_data.sort(key=lambda x: (x['tanggal'], -x['id']), reverse=True)
    
    context = {
        'combined_data': combined_data,
        'mechanical_data': mechanical_data,  # tetep kirim yang original buat compatibility
        'laporan_mechanical_data': laporan_mechanical_data,
        'selected_date': tanggal_parsed,
        'total_mechanical': mechanical_data.count(),
        'total_laporan': laporan_mechanical_data.count(),
        'total_combined': len(combined_data),
        'data_info': {
            'mechanical_model': 'MechanicalData',  # info model yang dipake
            'laporan_model': 'LaporanMechanicalData',
            'has_mechanical': mechanical_data.exists(),
            'has_laporan': laporan_mechanical_data.exists(),
        }
    }
    
    return render(request, 'dailyactivity_app/data_mechanical.html', context)

@login_required
def edit_mechanical_data(request, id):
    """
    Edit data mechanical untuk model MechanicalData dengan design modern
    """
    # Ambil data yang akan diedit berdasarkan id - gunakan MechanicalData
    mechanical_data = get_object_or_404(MechanicalData, id=id)
    
    # Ambil semua data referensi
    shifts = Shift.objects.all()
    locations = Location.objects.all()
    machines = Machinemechanical.objects.all()
    categories = Category.objects.all()
    status = Status.objects.all()
    pic_mechanical = PICMechanical2.objects.all()  # Gunakan PICMechanical untuk MechanicalData

    if request.method == 'POST':
        try:
            # Function to safely convert empty string to None, then to int
            def safe_int_or_none(value):
                if not value or value == '' or value == 'None':
                    return None
                try:
                    return int(value)
                except (ValueError, TypeError):
                    return None
            
            # Function to safely get string value
            def safe_string(value):
                return value.strip() if value and value.strip() else None

            # Ambil dan clean semua data dari formulir
            tanggal_str = request.POST.get('tanggal')
            jam = request.POST.get('jam')

            # Convert string ke date
            tanggal = None
            if tanggal_str:
                try:
                    tanggal = datetime.strptime(tanggal_str, "%Y-%m-%d").date()
                except ValueError:
                    messages.error(request, 'Format tanggal tidak valid. Gunakan YYYY-MM-DD.')
                    return redirect('dailyactivity_app:edit_mechanical_data', id=id)
            
            # Convert ALL ID fields safely
            shift_id = safe_int_or_none(request.POST.get('shift'))
            location_id = safe_int_or_none(request.POST.get('location'))
            machine_id = safe_int_or_none(request.POST.get('machine'))
            category_id = safe_int_or_none(request.POST.get('category'))
            status_id = safe_int_or_none(request.POST.get('status'))
            
            # Text fields
            masalah = request.POST.get('masalah', '').strip()
            penyebab = request.POST.get('penyebab', '').strip()
            tindakan_perbaikan = request.POST.get('tindakan_perbaikan', '').strip()
            nomor_wo = safe_string(request.POST.get('nomor_wo'))
            waktu_pengerjaan = safe_string(request.POST.get('waktu_pengerjaan'))
            
            # File
            image = request.FILES.get('image')
            
            # PIC IDs - filter empty values
            pic_ids_raw = request.POST.getlist('pic')
            pic_ids = [safe_int_or_none(pic_id) for pic_id in pic_ids_raw if safe_int_or_none(pic_id) is not None]

            print(f"DEBUG EDIT - shift_id: {shift_id}, category_id: {category_id}, status_id: {status_id}")
            print(f"DEBUG EDIT - location_id: {location_id}, machine_id: {machine_id}")

            # Validasi REQUIRED fields
            if not shift_id:
                messages.error(request, 'Shift harus dipilih.')
                return redirect('dailyactivity_app:edit_mechanical_data', id=id)
                
            if not category_id:
                messages.error(request, 'Jenis Pekerjaan harus dipilih.')
                return redirect('dailyactivity_app:edit_mechanical_data', id=id)
                
            if not status_id:
                messages.error(request, 'Status harus dipilih.')
                return redirect('dailyactivity_app:edit_mechanical_data', id=id)

            if not masalah:
                messages.error(request, 'Masalah harus diisi.')
                return redirect('dailyactivity_app:edit_mechanical_data', id=id)

            if not penyebab:
                messages.error(request, 'Penyebab harus diisi.')
                return redirect('dailyactivity_app:edit_mechanical_data', id=id)

            if not tindakan_perbaikan:
                messages.error(request, 'Tindakan Perbaikan harus diisi.')
                return redirect('dailyactivity_app:edit_mechanical_data', id=id)

            # Get required instances
            try:
                shift_instance = Shift.objects.get(id=shift_id)
                category_instance = Category.objects.get(id=category_id)
                status_instance = Status.objects.get(id=status_id)
            except (Shift.DoesNotExist, Category.DoesNotExist, Status.DoesNotExist) as e:
                messages.error(request, f'Data referensi tidak ditemukan: {str(e)}')
                return redirect('dailyactivity_app:edit_mechanical_data', id=id)

            # Handle Location - bisa None
            location_instance = None
            if location_id:
                try:
                    location_instance = Location.objects.get(id=location_id)
                except Location.DoesNotExist:
                    messages.error(request, 'Location tidak ditemukan.')
                    return redirect('dailyactivity_app:edit_mechanical_data', id=id)
            
            # Kalau location kosong, pakai yang lama atau buat default
            if not location_instance:
                location_instance = mechanical_data.location  # Keep existing location
                if not location_instance:
                    # Create default location if needed
                    location_instance, created = Location.objects.get_or_create(
                        name="Unknown Location",
                        defaults={'name': "Unknown Location"}
                    )

            # Handle Machine - bisa None
            machine_instance = None
            if machine_id:
                try:
                    machine_instance = Machinemechanical.objects.get(id=machine_id)
                except Machinemechanical.DoesNotExist:
                    messages.error(request, 'Machine tidak ditemukan.')
                    return redirect('dailyactivity_app:edit_mechanical_data', id=id)
            
            # Kalau machine kosong, pakai yang lama atau buat default  
            if not machine_instance:
                machine_instance = mechanical_data.machine  # Keep existing machine
                if not machine_instance:
                    # Create default machine if needed
                    machine_instance, created = Machinemechanical.objects.get_or_create(
                        name="Unknown Machine",
                        location=location_instance,
                        defaults={
                            'name': "Unknown Machine",
                            'location': location_instance,
                            'nomor': None
                        }
                    )

             # Process jam field
            jam_value = jam if jam else None

            print(f"DEBUG EDIT - About to update with: shift={shift_instance.id}, location={location_instance.id}, machine={machine_instance.id}")

            # UPDATE DATA MECHANICAL
            mechanical_data.tanggal = tanggal
            mechanical_data.jam = jam_value
            mechanical_data.shift = shift_instance
            mechanical_data.location = location_instance
            mechanical_data.machine = machine_instance
            mechanical_data.category = category_instance
            mechanical_data.status = status_instance
            mechanical_data.masalah = masalah
            mechanical_data.penyebab = penyebab
            mechanical_data.tindakan = tindakan_perbaikan  # Backward compatibility
            mechanical_data.tindakan_perbaikan = tindakan_perbaikan
            mechanical_data.nomor_wo = nomor_wo
            mechanical_data.waktu_pengerjaan = waktu_pengerjaan
            
            # Update image only if new file uploaded
            if image:
                mechanical_data.image = image
                
            mechanical_data.save()

            print(f"DEBUG EDIT - Mechanical data updated with ID: {mechanical_data.id}")

            # Update PICs - clear existing and add new ones
            mechanical_data.pic.clear()
            for pic_id in pic_ids:
                try:
                    pic_instance = PICMechanical2.objects.get(id=pic_id)
                    mechanical_data.pic.add(pic_instance)
                    print(f"DEBUG EDIT - Added PIC: {pic_id}")
                except PICMechanical2.DoesNotExist:
                    print(f"DEBUG EDIT - PIC not found: {pic_id}")
                    continue

            messages.success(request, 'Data berhasil diperbarui!')
            return redirect('dailyactivity_app:data_mechanical', tanggal=mechanical_data.tanggal.strftime('%Y-%m-%d'))

        except Exception as e:
            print(f"ERROR EDIT - Exception in edit_mechanical_data: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Terjadi kesalahan saat memperbarui data: {str(e)}')
            return redirect('dailyactivity_app:edit_mechanical_data', id=id)

    # Context untuk GET request
    context = {
        'shifts': shifts,
        'locations': locations,
        'machines': machines,
        'categories': categories,
        'status': status,
        'pic_mechanical': pic_mechanical,
        'data': mechanical_data,
    }
    
    return render(request, 'dailyactivity_app/edit_mechanical_data.html', context)

# @login_required
# def delete_mechanical_data(request, id):
#     # Mengambil data berdasarkan ID - gunakan MechanicalData
#     mechanical_data = get_object_or_404(MechanicalData, id=id)
#     if request.method == 'POST':
#         # Hapus data dari database
#         mechanical_data.delete()
#         messages.success(request, 'Data berhasil dihapus!')
#         return redirect('dailyactivity_app:data_mechanical', tanggal=mechanical_data.tanggal.strftime('%Y-%m-%d'))   
#     # Jika bukan POST, redirect ke halaman data mechanical
#     return redirect('dailyactivity_app:data_mechanical', tanggal=mechanical_data.tanggal.strftime('%Y-%m-%d'))
@login_required
def delete_mechanical_data(request, id):
    """
    Delete data mechanical untuk model MechanicalData
    """
    # Mengambil data berdasarkan ID - gunakan MechanicalData
    mechanical_data = get_object_or_404(MechanicalData, id=id)
    
    if request.method == 'POST':
        try:
            # Simpan tanggal sebelum dihapus untuk redirect
            tanggal_redirect = mechanical_data.tanggal.strftime('%Y-%m-%d')
            
            # Log informasi sebelum hapus
            print(f"DEBUG DELETE - Deleting MechanicalData ID: {id}")
            print(f"DEBUG DELETE - Data: {mechanical_data.tanggal} - {mechanical_data.masalah[:50]}...")
            
            # Hapus data dari database
            mechanical_data.delete()
            
            messages.success(request, 'Data mechanical berhasil dihapus!')
            print(f"DEBUG DELETE - Successfully deleted MechanicalData ID: {id}")
            
            return redirect('dailyactivity_app:data_mechanical', tanggal=tanggal_redirect)
            
        except Exception as e:
            print(f"ERROR DELETE - Exception in delete_mechanical_data: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Terjadi kesalahan saat menghapus data: {str(e)}')
            
            # Redirect ke data mechanical dengan tanggal yang sama
            tanggal_redirect = mechanical_data.tanggal.strftime('%Y-%m-%d')
            return redirect('dailyactivity_app:data_mechanical', tanggal=tanggal_redirect)
    
    # Jika bukan POST request, redirect ke data mechanical
    tanggal_redirect = mechanical_data.tanggal.strftime('%Y-%m-%d')
    return redirect('dailyactivity_app:data_mechanical', tanggal=tanggal_redirect)
   
# @login_required
# def electrical_index(request):
#     shifts = Shift.objects.all()
#     locations = Location.objects.all()
#     machines = Machineelectrical.objects.all()
#     categories = Category.objects.all()
#     status = Status.objects.all()
#     pic_electrical = PICElectrical.objects.all()

#     # Ambil 20 nomor_wo dari tabel_main di DB_Maintenance dengan filter berdasarkan section_id = 5
#     nomor_wo_list = []
#     with connections['DB_Maintenance'].cursor() as cursor:
#         cursor.execute("""
#             SELECT TOP 20 number_wo
#             FROM dbo.tabel_main
#             WHERE id_section = 5
#             ORDER BY history_id DESC
#         """)
#         nomor_wo_list = [row[0] for row in cursor.fetchall()]  # Simpan hasil query sebelum keluar dari blok `with`

#     deskripsi_perbaikan = None  # Default value untuk deskripsi_perbaikan
#     tgl_his = None  # Default value untuk tgl_his  
#     if request.method == 'POST':
#         form = ElectricalDataForm(request.POST, request.FILES)
#         if form.is_valid():

#             # Ambil nilai dari form
#             machine_number = form.cleaned_data.get('machine_number')
#             machine_instance = form.cleaned_data.get('machine')

#             # Cek apakah mesin sudah memiliki nomor
#             if machine_number:
#                 # Jika tidak ada nomor mesin pada instance machine, inputkan nomor mesin baru
#                 if not machine_instance.nomor:
#                     machine_instance.nomor = machine_number
#                     machine_instance.save()  # Update nomor mesin
#                 else:
#                     # Jika nomor sudah ada, gunakan nomor yang ada
#                     pass
#             electrical_data = form.save(commit=False)
#             electrical_data.user = request.user
#             electrical_data.machine = machine_instance  # Set machine_instance sebagai mesin yang dipilih atau baru
#             electrical_data.save()
#             pic_ids = form.cleaned_data.get('pic')
#             electrical_data.pic.set(pic_ids)
#             electrical_data.nomor_wo = form.cleaned_data.get('nomor_wo')
#             electrical_data.waktu_pengerjaan = form.cleaned_data.get('waktu_pengerjaan')

#             electrical_data.save()
#             return redirect('success_page')
#     else:
#         form = ElectricalDataForm()

#     # Cek apakah request adalah AJAX dengan header 'HTTP_X_REQUESTED_WITH'
#     if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.method == 'GET':
#         nomor_wo_selected = request.GET.get('nomor_wo')
#         with connections['DB_Maintenance'].cursor() as cursor:
#             cursor.execute("""
#                 SELECT deskripsi_perbaikan, tgl_his
#                 FROM dbo.tabel_main
#                 WHERE number_wo = %s
#             """, [nomor_wo_selected])
#             row = cursor.fetchone()
#             if row:
#                 deskripsi_perbaikan = row[0]
#                 tgl_his = row[1]
#         return JsonResponse({
#             'deskripsi_perbaikan': deskripsi_perbaikan,
#             'tgl_his': tgl_his
#         })

#     context = {
#         'shifts': shifts,
#         'locations': locations,
#         'machines': machines,
#         'categories': categories,
#         'status': status,
#         'pic_electrical': pic_electrical,
#         'nomor_wo_list': nomor_wo_list,  # Kirim nomor_wo_list ke template
#         'form': form,
#         'deskripsi_perbaikan': deskripsi_perbaikan,  # Menambahkan deskripsi_perbaikan
#         'tgl_his': tgl_his,  # Menambahkan tgl_his
#     }
#     return render(request, 'dailyactivity_app/electrical_index.html', context)

@login_required 
def electrical_index(request):
    shifts = Shift.objects.all()
    locations = Location.objects.all()
    machines = Machineelectrical.objects.all()
    categories = Category.objects.all()
    status = Status.objects.all()
    pic_electrical = PICElectrical.objects.all()

    # Ambil hanya 30 nomor WO paling terbaru
    nomor_wo_list = []
    with connections['DB_Maintenance'].cursor() as cursor:
        cursor.execute("""
            SELECT TOP 30 number_wo
            FROM dbo.view_main
            WHERE id_section = 5
            AND YEAR(tgl_his) BETWEEN 2024 AND 2025
            ORDER BY history_id DESC
        """)
        nomor_wo_list = [row[0] for row in cursor.fetchall()]

    # Default values
    deskripsi_perbaikan = None
    tgl_his = None
    penyebab = None
    line = None
    mesin = None
    nomer = None
    pekerjaan = None
    status_pekerjaan = None
    tindakan_perbaikan = None

    # Tangani Permintaan AJAX untuk nomor WO
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.method == 'GET':
        nomor_wo_selected = request.GET.get('nomor_wo')
        
        # Prepare response data
        response_data = {
            'deskripsi_perbaikan': None,
            'tgl_his': None,
            'penyebab': None,
            'line': None,
            'mesin': None,
            'nomer': None,
            'pekerjaan': None,
            'status_pekerjaan': None,
            'tindakan_perbaikan': None,
            'location_id': None,
            'machine_id': None
        }
        
        with connections['DB_Maintenance'].cursor() as cursor:
            cursor.execute("""
                SELECT deskripsi_perbaikan, tgl_his, penyebab, line, mesin, nomer, 
                       pekerjaan, status_pekerjaan, tindakan_perbaikan
                FROM dbo.view_main
                WHERE number_wo = %s
            """, [nomor_wo_selected])
            row = cursor.fetchone()
            
            if row:
                (deskripsi_perbaikan, tgl_his, penyebab, line, mesin, 
                 nomer, pekerjaan, status_pekerjaan, tindakan_perbaikan) = row
                
                response_data.update({
                    'deskripsi_perbaikan': deskripsi_perbaikan,
                    'tgl_his': tgl_his,
                    'penyebab': penyebab,
                    'line': line,
                    'mesin': mesin,
                    'nomer': nomer,
                    'pekerjaan': pekerjaan,
                    'status_pekerjaan': status_pekerjaan,
                    'tindakan_perbaikan': tindakan_perbaikan
                })
                
                # TAMBAHAN: Cari location_id berdasarkan nama line
                if line:
                    try:
                        location_obj = Location.objects.filter(name__icontains=line).first()
                        if location_obj:
                            response_data['location_id'] = location_obj.id
                    except:
                        pass
                
                # TAMBAHAN: Cari machine_id berdasarkan nama mesin dan nomor
                if mesin:
                    try:
                        machine_query = Machineelectrical.objects.filter(name__icontains=mesin)
                        if nomer:
                            machine_query = machine_query.filter(nomor__icontains=nomer)
                        
                        machine_obj = machine_query.first()
                        if machine_obj:
                            response_data['machine_id'] = machine_obj.id
                    except:
                        pass

        return JsonResponse(response_data)

    # Context untuk template
    context = {
        'shifts': shifts,
        'locations': locations,
        'machines': machines,
        'categories': categories,
        'status': status,
        'pic_electrical': pic_electrical,
        'nomor_wo_list': nomor_wo_list,
        'deskripsi_perbaikan': deskripsi_perbaikan,
        'tgl_his': tgl_his,
        'penyebab': penyebab,
        'line': line,
        'mesin': mesin,
        'nomer': nomer,
        'pekerjaan': pekerjaan,
        'status_pekerjaan': status_pekerjaan,
        'tindakan_perbaikan': tindakan_perbaikan,
    }
    return render(request, 'dailyactivity_app/electrical_index.html', context)

#     return redirect('dailyactivity_app:electrical_index')
def get_machines_by_location_electrical(request, location_id):
    # Mengambil data mesin berdasarkan location_id dan mengurutkan berdasarkan `name` dan `nomor`
    machines = Machineelectrical.objects.filter(location_id=location_id).order_by('name', 'nomor') 
    # Mengambil hanya beberapa field yang relevan
    machine_list = list(machines.values('id', 'name', 'location_id', 'nomor')) 
    # Mengembalikan response dalam format JSON
    return JsonResponse(machine_list, safe=False)
def get_machine_number_electrical(request, machine_id):
    try:
        # Mengambil mesin berdasarkan ID
        machine = Machineelectrical.objects.get(id=machine_id)
        # Mengembalikan nomor mesin
        return JsonResponse({'nomor': machine.nomor})
    except Machineelectrical.DoesNotExist:
        return JsonResponse({'error': 'Machine not found'}, status=404)

# @login_required  # Pastikan pengguna harus login untuk mengakses fungsi ini
# def electrical_submit(request):
#     if request.method == 'POST':
#         # Ambil data dari formulir
#         tanggal = request.POST.get('tanggal')
#         jam = request.POST.get('jam')
#         tgl_his = request.POST.get('tgl_his')
#         shift_id = request.POST.get('shift')
#         location_id = request.POST.get('location')
#         machine_id = request.POST.get('machine')
#         category_id = request.POST.get('category')
#         status_id = request.POST.get('status')
#         masalah = request.POST.get('masalah')
#         penyebab = request.POST.get('penyebab')
#         tindakan = request.POST.get('tindakan')
#         image = request.FILES.get('image')
#         pic_ids = request.POST.getlist('pic')  # Ambil daftar PIC yang dipilih
#         nomor_wo = request.POST.get('nomor_wo')  # Ambil nomor WO
#         waktu_pengerjaan = request.POST.get('waktu_pengerjaan')  # Ambil waktu pengerjaan

#         # Ambil instance Shift
#         try:
#             shift_instance = Shift.objects.get(id=shift_id)
#         except Shift.DoesNotExist:
#             messages.error(request, 'Shift tidak ditemukan.')
#             return redirect('dailyactivity_app:electrical_index')

#         # Ambil instance Location
#         try:
#             location_instance = Location.objects.get(id=location_id)
#         except Location.DoesNotExist:
#             messages.error(request, 'Location tidak ditemukan.')
#             return redirect('dailyactivity_app:electrical_index')

#         # Ambil instance Machinemechanical
#         try:
#             machine_instance = Machineelectrical.objects.get(id=machine_id)
#         except Machineelectrical.DoesNotExist:
#             messages.error(request, 'Machine tidak ditemukan.')
#             return redirect('dailyactivity_app:electrical_index')

#         # Ambil instance Category
#         try:
#             category_instance = Category.objects.get(id=category_id)
#         except Category.DoesNotExist:
#             messages.error(request, 'Category tidak ditemukan.')
#             return redirect('dailyactivity_app:electrical_index')

#         # Ambil instance Status
#         try:
#             status_instance = Status.objects.get(id=status_id)
#         except Status.DoesNotExist:
#             messages.error(request, 'Status tidak ditemukan.')
#             return redirect('dailyactivity_app:electrical_index')

#         # Ambil user_id dari pengguna yang sedang login
#         user_id = request.user.id

#         jam_value = tgl_his if tgl_his else jam

#         # Simpan data ke database
#         electrical_data = ElectricalData.objects.create(
#             tanggal=tanggal,
#             jam=jam_value,
#             shift=shift_instance,
#             location=location_instance,
#             machine=machine_instance,
#             category=category_instance,
#             status=status_instance,
#             user_id=user_id,
#             masalah=masalah,
#             penyebab=penyebab,
#             tindakan=tindakan,
#             image=image,
#             nomor_wo=nomor_wo,  # Simpan nomor WO
#             waktu_pengerjaan=waktu_pengerjaan  # Simpan waktu pengerjaan3
#         )

#         # Menyimpan PIC yang dipilih
#         for pic_id in pic_ids:
#             try:
#                 pic_instance = PICElectrical.objects.get(id=pic_id)
#                 electrical_data.pic.add(pic_instance)
#             except PICElectrical.DoesNotExist:
#                 continue

#         # Simpan pesan sukses
#         messages.success(request, 'Data berhasil disimpan!')

#         # Redirect ke mechanical_index
#         return redirect('dailyactivity_app:electrical_index')

#     return redirect('dailyactivity_app:electrical_index')

@login_required  
def electrical_submit(request):
    if request.method == 'POST':
        try:
            # Function to safely convert empty string to None, then to int
            def safe_int_or_none(value):
                if not value or value == '' or value == 'None':
                    return None
                try:
                    return int(value)
                except (ValueError, TypeError):
                    return None
            
            # Function to safely get string value
            def safe_string(value):
                return value.strip() if value and value.strip() else None

            # Ambil dan clean semua data dari formulir
            tanggal = request.POST.get('tanggal')
            jam = request.POST.get('jam')
            tgl_his = request.POST.get('tgl_his')
            
            # Convert ALL ID fields safely
            shift_id = safe_int_or_none(request.POST.get('shift'))
            location_id = safe_int_or_none(request.POST.get('location'))
            machine_id = safe_int_or_none(request.POST.get('machine'))
            category_id = safe_int_or_none(request.POST.get('category'))
            status_id = safe_int_or_none(request.POST.get('status'))
            
            # Text fields
            masalah = request.POST.get('masalah', '').strip()
            penyebab = request.POST.get('penyebab', '').strip()
            tindakan_perbaikan = request.POST.get('tindakan_perbaikan', '').strip()
            nomor_wo = safe_string(request.POST.get('nomor_wo'))
            waktu_pengerjaan = safe_string(request.POST.get('waktu_pengerjaan'))
            machine_number = safe_string(request.POST.get('machine_number'))
            
            # Manual machine inputs
            manual_machine_name = safe_string(request.POST.get('manual_machine_name'))
            manual_machine_number = safe_string(request.POST.get('manual_machine_number'))
            
            # File
            image = request.FILES.get('image')
            
            # PIC IDs - filter empty values
            pic_ids_raw = request.POST.getlist('pic')
            pic_ids = [safe_int_or_none(pic_id) for pic_id in pic_ids_raw if safe_int_or_none(pic_id) is not None]

            print(f"DEBUG - shift_id: {shift_id}, category_id: {category_id}, status_id: {status_id}")
            print(f"DEBUG - location_id: {location_id}, machine_id: {machine_id}")
            print(f"DEBUG - manual_machine_name: {manual_machine_name}")

            # Validasi REQUIRED fields dengan proper error messages
            if not shift_id:
                messages.error(request, 'Shift harus dipilih.')
                return redirect('dailyactivity_app:electrical_index')
                
            if not category_id:
                messages.error(request, 'Jenis Pekerjaan harus dipilih.')
                return redirect('dailyactivity_app:electrical_index')
                
            if not status_id:
                messages.error(request, 'Status harus dipilih.')
                return redirect('dailyactivity_app:electrical_index')

            if not masalah:
                messages.error(request, 'Masalah harus diisi.')
                return redirect('dailyactivity_app:electrical_index')

            if not penyebab:
                messages.error(request, 'Penyebab harus diisi.')
                return redirect('dailyactivity_app:electrical_index')

            if not tindakan_perbaikan:
                messages.error(request, 'Tindakan Perbaikan harus diisi.')
                return redirect('dailyactivity_app:electrical_index')

            # Get required instances
            try:
                shift_instance = Shift.objects.get(id=shift_id)
                category_instance = Category.objects.get(id=category_id)
                status_instance = Status.objects.get(id=status_id)
            except (Shift.DoesNotExist, Category.DoesNotExist, Status.DoesNotExist) as e:
                messages.error(request, f'Data referensi tidak ditemukan: {str(e)}')
                return redirect('dailyactivity_app:electrical_index')

            # HANDLE LOCATION - ALWAYS CREATE ONE
            location_instance = None
            if location_id:
                try:
                    location_instance = Location.objects.get(id=location_id)
                except Location.DoesNotExist:
                    messages.error(request, 'Location tidak ditemukan.')
                    return redirect('dailyactivity_app:electrical_index')
            
            # Create default location if none selected
            if not location_instance:
                location_instance, created = Location.objects.get_or_create(
                    name="Unknown Location",
                    defaults={'name': "Unknown Location"}
                )
                print(f"DEBUG - Created/found default location: {location_instance.id}")

            # HANDLE MACHINE - ALWAYS CREATE ONE
            machine_instance = None
            
            # Option 1: Machine selected from dropdown
            if machine_id:
                try:
                    machine_instance = Machineelectrical.objects.get(id=machine_id)
                    print(f"DEBUG - Found machine from dropdown: {machine_instance.id}")
                    
                    # Update machine number if provided and machine doesn't have one
                    if machine_number and not machine_instance.nomor:
                        machine_instance.nomor = machine_number
                        machine_instance.save()
                        
                except Machineelectrical.DoesNotExist:
                    print(f"DEBUG - Machine ID {machine_id} not found")
                    machine_instance = None
            
            # Option 2: Manual machine input
            if not machine_instance and manual_machine_name:
                try:
                    print(f"DEBUG - Trying to create/find manual machine: {manual_machine_name}")
                    
                    # Check if machine already exists
                    machine_instance = Machineelectrical.objects.filter(
                        name__iexact=manual_machine_name
                    ).first()
                    
                    if not machine_instance:
                        # Create new machine
                        machine_instance = Machineelectrical.objects.create(
                            name=manual_machine_name,
                            location=location_instance,
                            nomor=manual_machine_number
                        )
                        print(f"DEBUG - Created new machine: {machine_instance.id}")
                    else:
                        # Update existing machine number if needed
                        if manual_machine_number and not machine_instance.nomor:
                            machine_instance.nomor = manual_machine_number
                            machine_instance.save()
                        print(f"DEBUG - Found existing machine: {machine_instance.id}")
                        
                except Exception as e:
                    print(f"DEBUG - Error creating manual machine: {e}")
                    machine_instance = None
            
            # Option 3: Create default machine if still none
            if not machine_instance:
                try:
                    machine_instance, created = Machineelectrical.objects.get_or_create(
                        name="Unknown Machine",
                        location=location_instance,
                        defaults={
                            'name': "Unknown Machine",
                            'location': location_instance,
                            'nomor': None
                        }
                    )
                    print(f"DEBUG - Created/found default machine: {machine_instance.id}, created: {created}")
                except Exception as e:
                    print(f"DEBUG - Error creating default machine: {e}")
                    # Last resort with timestamp to avoid conflicts
                    import time
                    timestamp = str(int(time.time()))
                    machine_instance = Machineelectrical.objects.create(
                        name=f"Unknown Machine {timestamp}",
                        location=location_instance,
                        nomor=None
                    )
                    print(f"DEBUG - Created timestamped machine: {machine_instance.id}")

            # Pastikan semua instance yang dibutuhkan ada
            if not all([shift_instance, location_instance, machine_instance, category_instance, status_instance]):
                missing = []
                if not shift_instance: missing.append("shift")
                if not location_instance: missing.append("location")
                if not machine_instance: missing.append("machine")
                if not category_instance: missing.append("category") 
                if not status_instance: missing.append("status")
                
                messages.error(request, f'Instance tidak lengkap: {", ".join(missing)}')
                return redirect('dailyactivity_app:electrical_index')

            # Process jam field
            jam_value = None
            if tgl_his:
                jam_value = tgl_his
            elif jam:
                jam_value = jam

            print(f"DEBUG - About to save with: shift={shift_instance.id}, location={location_instance.id}, machine={machine_instance.id}")

            # SAVE DATA TO DATABASE
            electrical_data = ElectricalData.objects.create(
                tanggal=tanggal,
                jam=jam_value,
                shift=shift_instance,
                location=location_instance,
                machine=machine_instance,
                category=category_instance,
                status=status_instance,
                user=request.user,  # Use request.user instead of user_id
                masalah=masalah,
                penyebab=penyebab,
                tindakan=tindakan_perbaikan,
                # tindakan_perbaikan=tindakan_perbaikan,
                image=image,
                nomor_wo=nomor_wo,
                waktu_pengerjaan=waktu_pengerjaan
            )

            print(f"DEBUG - electrical_data data created with ID: {electrical_data.id}")

            # Add PICs
            for pic_id in pic_ids:
                try:
                    pic_instance = PICElectrical.objects.get(id=pic_id)
                    electrical_data.pic.add(pic_instance)
                    print(f"DEBUG - Added PIC: {pic_id}")
                except PICElectrical.DoesNotExist:
                    print(f"DEBUG - PIC not found: {pic_id}")
                    continue

            messages.success(request, 'Data berhasil disimpan!')
            return redirect('dailyactivity_app:electrical_index')

        except Exception as e:
            print(f"ERROR - Exception in electrical_submit: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
            return redirect('dailyactivity_app:electrical_index')

    return redirect('dailyactivity_app:electrical_index')

@login_required
def data_electrical(request, tanggal):
    # Parsing tanggal dari URL
    tanggal_parsed = parse_date(tanggal)
    # Menyaring data berdasarkan tanggal yang dipilih
    electrical_data = ElectricalData.objects.filter(tanggal=tanggal_parsed)
    context = {
        'electrical_data': electrical_data,
        'selected_date': tanggal_parsed,
    }
    return render(request, 'dailyactivity_app/data_electrical.html', context)

# @login_required
# def edit_electrical_data(request, id):
#     # Ambil data yang akan diedit berdasarkan id
#     electrical_data = get_object_or_404(ElectricalData, id=id)
#     shifts = Shift.objects.all()
#     locations = Location.objects.all()
#     machines = Machineelectrical.objects.all()
#     categories = Category.objects.all()
#     status = Status.objects.all()
#     pic_electrical = PICElectrical.objects.all()

#     if request.method == 'POST':
#         # Memproses form data yang telah diisi ulang
#         form = ElectricalDataForm(request.POST, request.FILES, instance=electrical_data)
#         if form.is_valid():
#             updated_data = form.save(commit=False)
#             updated_data.user = request.user
#             updated_data.nomor_wo = form.cleaned_data.get('nomor_wo')
#             updated_data.waktu_pengerjaan = form.cleaned_data.get('waktu_pengerjaan')
#             updated_data.save()

#             # Update PIC terkait
#             pic_ids = form.cleaned_data.get('pic')
#             updated_data.pic.set(pic_ids)

#             messages.success(request, 'Data berhasil diperbarui!')
#             return redirect('dailyactivity_app:data_electrical', tanggal=updated_data.tanggal.strftime('%Y-%m-%d'))
#         else:
#             messages.error(request, 'Terjadi kesalahan saat memperbarui data. Periksa kembali isian Anda.')

#     else:
#         # Memuat form dengan data yang ada untuk ditampilkan di template
#         form = ElectricalDataForm(instance=electrical_data)

#     context = {
#         'form': form,
#         'shifts': shifts,
#         'locations': locations,
#         'machines': machines,
#         'categories': categories,
#         'status': status,
#         'pic_electrical': pic_electrical,
#         'data': electrical_data,
#         'tanggal': electrical_data.tanggal,
#         'jam': electrical_data.jam,
#         'nomor_wo': electrical_data.nomor_wo,
#         'waktu_pengerjaan': electrical_data.waktu_pengerjaan,
#         'pic': electrical_data.pic.all(),
#     }
#     return render(request, 'dailyactivity_app/edit_electrical_data.html', context)

@login_required
def edit_electrical_data(request, id):
    """
    Edit data electrical untuk model ElectricalData
    """
    electrical_data = get_object_or_404(ElectricalData, id=id)
    shifts = Shift.objects.all()
    locations = Location.objects.all()
    machines = Machineelectrical.objects.all()
    categories = Category.objects.all()
    status = Status.objects.all()
    pic_electrical = PICElectrical.objects.all()

    if request.method == 'POST':
        try:
            # Function to safely convert empty string to None, then to int
            def safe_int_or_none(value):
                if not value or value == '' or value == 'None':
                    return None
                try:
                    return int(value)
                except (ValueError, TypeError):
                    return None
            
            # Function to safely get string value
            def safe_string(value):
                return value.strip() if value and value.strip() else None

            # Ambil dan clean semua data dari formulir
            tanggal_str = request.POST.get('tanggal')
            jam = request.POST.get('jam')

            # Convert string ke date
            tanggal = None
            if tanggal_str:
                try:
                    tanggal = datetime.strptime(tanggal_str, "%Y-%m-%d").date()
                except ValueError:
                    messages.error(request, 'Format tanggal tidak valid. Gunakan YYYY-MM-DD.')
                    return redirect('dailyactivity_app:edit_electrical_data', id=id)
            
            # Convert ALL ID fields safely
            shift_id = safe_int_or_none(request.POST.get('shift'))
            location_id = safe_int_or_none(request.POST.get('location'))
            machine_id = safe_int_or_none(request.POST.get('machine'))
            category_id = safe_int_or_none(request.POST.get('category'))
            status_id = safe_int_or_none(request.POST.get('status'))
            
            # Text fields
            masalah = request.POST.get('masalah', '').strip()
            penyebab = request.POST.get('penyebab', '').strip()
            tindakan_perbaikan = request.POST.get('tindakan_perbaikan', '').strip()
            nomor_wo = safe_string(request.POST.get('nomor_wo'))
            waktu_pengerjaan = safe_string(request.POST.get('waktu_pengerjaan'))
            
            # File
            image = request.FILES.get('image')
            
            # PIC IDs - filter empty values
            pic_ids_raw = request.POST.getlist('pic')
            pic_ids = [safe_int_or_none(pic_id) for pic_id in pic_ids_raw if safe_int_or_none(pic_id) is not None]

            # Validasi REQUIRED fields
            if not shift_id:
                messages.error(request, 'Shift harus dipilih.')
                return redirect('dailyactivity_app:edit_electrical_data', id=id)
                
            if not category_id:
                messages.error(request, 'Jenis Pekerjaan harus dipilih.')
                return redirect('dailyactivity_app:edit_electrical_data', id=id)
                
            if not status_id:
                messages.error(request, 'Status harus dipilih.')
                return redirect('dailyactivity_app:edit_electrical_data', id=id)

            if not masalah:
                messages.error(request, 'Masalah harus diisi.')
                return redirect('dailyactivity_app:edit_electrical_data', id=id)

            if not penyebab:
                messages.error(request, 'Penyebab harus diisi.')
                return redirect('dailyactivity_app:edit_electrical_data', id=id)

            if not tindakan_perbaikan:
                messages.error(request, 'Tindakan Perbaikan harus diisi.')
                return redirect('dailyactivity_app:edit_electrical_data', id=id)

            # Get required instances
            try:
                shift_instance = Shift.objects.get(id=shift_id)
                category_instance = Category.objects.get(id=category_id)
                status_instance = Status.objects.get(id=status_id)
            except (Shift.DoesNotExist, Category.DoesNotExist, Status.DoesNotExist) as e:
                messages.error(request, f'Data referensi tidak ditemukan: {str(e)}')
                return redirect('dailyactivity_app:edit_electrical_data', id=id)

            # Handle Location - bisa None
            location_instance = None
            if location_id:
                try:
                    location_instance = Location.objects.get(id=location_id)
                except Location.DoesNotExist:
                    messages.error(request, 'Location tidak ditemukan.')
                    return redirect('dailyactivity_app:edit_electrical_data', id=id)
            
            if not location_instance:
                location_instance = electrical_data.location  # Keep existing location
                if not location_instance:
                    location_instance, created = Location.objects.get_or_create(
                        name="Unknown Location",
                        defaults={'name': "Unknown Location"}
                    )

            # Handle Machine - bisa None
            machine_instance = None
            if machine_id:
                try:
                    machine_instance = Machineelectrical.objects.get(id=machine_id)
                except Machineelectrical.DoesNotExist:
                    messages.error(request, 'Machine tidak ditemukan.')
                    return redirect('dailyactivity_app:edit_electrical_data', id=id)
            
            if not machine_instance:
                machine_instance = electrical_data.machine  # Keep existing machine
                if not machine_instance:
                    machine_instance, created = Machineelectrical.objects.get_or_create(
                        name="Unknown Machine",
                        location=location_instance,
                        defaults={
                            'name': "Unknown Machine",
                            'location': location_instance,
                            'nomor': None
                        }
                    )

            # Process jam field
            jam_value = jam if jam else None

            # UPDATE DATA ELECTRICAL
            electrical_data.tanggal = tanggal
            electrical_data.jam = jam_value
            electrical_data.shift = shift_instance
            electrical_data.location = location_instance
            electrical_data.machine = machine_instance
            electrical_data.category = category_instance
            electrical_data.status = status_instance
            electrical_data.masalah = masalah
            electrical_data.penyebab = penyebab
            electrical_data.tindakan = tindakan_perbaikan  # Backward compatibility
            # electrical_data.tindakan_perbaikan = tindakan_perbaikan
            electrical_data.nomor_wo = nomor_wo
            electrical_data.waktu_pengerjaan = waktu_pengerjaan
            
            if image:
                electrical_data.image = image
                
            electrical_data.save()

            # Update PICs - clear existing and add new ones
            electrical_data.pic.clear()
            for pic_id in pic_ids:
                try:
                    pic_instance = PICElectrical.objects.get(id=pic_id)
                    electrical_data.pic.add(pic_instance)
                except PICElectrical.DoesNotExist:
                    continue

            messages.success(request, 'Data berhasil diperbarui!')
            return redirect('dailyactivity_app:data_electrical', tanggal=electrical_data.tanggal.strftime('%Y-%m-%d'))

        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f'Terjadi kesalahan saat memperbarui data: {str(e)}')
            return redirect('dailyactivity_app:edit_electrical_data', id=id)

    # Context untuk GET request
    context = {
        'shifts': shifts,
        'locations': locations,
        'machines': machines,
        'categories': categories,
        'status': status,
        'pic_electrical': pic_electrical,
        'data': electrical_data,
    }
    
    return render(request, 'dailyactivity_app/edit_electrical_data.html', context)



@login_required
def delete_electrical_data(request, id):
    # Mengambil data berdasarkan ID
    electrical_data = get_object_or_404(ElectricalData, id=id)

    if request.method == 'POST':
        # Hapus data dari database
        electrical_data.delete()
        messages.success(request, 'Data berhasil dihapus!')
        return redirect('dailyactivity_app:data_electrical', tanggal=electrical_data.tanggal.strftime('%Y-%m-%d'))
    
    # Jika bukan POST, redirect ke halaman data electrical
    return redirect('dailyactivity_app:data_electrical', tanggal=electrical_data.tanggal.strftime('%Y-%m-%d'))


@login_required
def tanggal_electrical(request):
    # Ambil parameter bulan dan tahun dari URL jika ada
    selected_month = request.GET.get('month')
    selected_year = request.GET.get('year')
    
    print(f"DEBUG - Raw parameters: month='{selected_month}', year='{selected_year}'")
    
    if selected_month and selected_year:
        try:
            # Fungsi untuk membersihkan dan konversi nilai
            def clean_and_convert(value):
                if value is None:
                    return None
                
                # Convert to string terlebih dahulu
                str_value = str(value).strip()
                
                # Jika kosong, return None
                if not str_value:
                    return None
                
                # Jika ada decimal (seperti 2.025), ambil bagian integer saja
                if '.' in str_value:
                    # Split dan ambil bagian sebelum decimal
                    integer_part = str_value.split('.')[0]
                    return int(integer_part) if integer_part.isdigit() else None
                
                # Jika sudah integer/string digit biasa
                if str_value.isdigit():
                    return int(str_value)
                
                # Coba konversi float dulu baru ke int (untuk handle kasus lain)
                try:
                    return int(float(str_value))
                except (ValueError, TypeError):
                    return None
            
            # Konversi parameter
            selected_month = clean_and_convert(selected_month)
            selected_year = clean_and_convert(selected_year)
            
            print(f"DEBUG - Cleaned: month={selected_month}, year={selected_year}")
            
            # Validasi hasil konversi
            if selected_month is None or selected_year is None:
                messages.error(request, 'Parameter bulan atau tahun tidak dapat dikonversi')
                return redirect('dailyactivity_app:tanggal_electrical')
            
            # Validasi range
            if not (1 <= selected_month <= 12):
                messages.error(request, f'Bulan harus antara 1-12, diterima: {selected_month}')
                return redirect('dailyactivity_app:tanggal_electrical')
                
            # Validasi tahun dengan range yang realistis
            current_year = datetime.now().year
            if not (2020 <= selected_year <= current_year + 2):
                messages.error(request, f'Tahun harus antara 2020-{current_year + 2}, diterima: {selected_year}')
                return redirect('dailyactivity_app:tanggal_electrical')
            
            # Query data dengan error handling
            try:
                dates = ElectricalData.objects.filter(
                    tanggal__month=selected_month,
                    tanggal__year=selected_year
                ).annotate(
                    date=TruncDate('tanggal', output_field=DateField())
                ).values('date').distinct().order_by('-date')
                
                dates_count = dates.count()
                print(f"DEBUG - Query successful, found {dates_count} dates")
                
            except Exception as query_error:
                print(f"DEBUG - Query error: {query_error}")
                messages.error(request, f'Error saat mengambil data: {query_error}')
                return redirect('dailyactivity_app:tanggal_electrical')
            
            # Nama bulan
            month_names = {
                1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
                5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
                9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
            }
            
            context = {
                'dates': dates,
                'selected_month': selected_month,
                'selected_year': selected_year,
                'selected_month_name': month_names.get(selected_month, f'Bulan {selected_month}'),
                'show_dates': True,
            }
            
        except Exception as e:
            print(f"DEBUG - General exception: {e}")
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
            return redirect('dailyactivity_app:tanggal_electrical')
    else:
        print("DEBUG - Showing month/year list")
        
        try:
            # Cek apakah ada data sama sekali
            total_count = ElectricalData.objects.count()
            print(f"DEBUG - Total ElectricalData records: {total_count}")
            
            if total_count == 0:
                messages.info(request, 'Belum ada data electrical dalam sistem')
                context = {
                    'month_year_data': [],
                    'show_dates': False,
                }
                return render(request, 'dailyactivity_app/tanggal_electrical.html', context)
            
            # Ambil sample data untuk debugging
            sample_data = ElectricalData.objects.first()
            if sample_data and sample_data.tanggal:
                print(f"DEBUG - Sample date: {sample_data.tanggal}, type: {type(sample_data.tanggal)}")
            
            # Ambil semua data dengan tanggal yang valid
            all_electrical_data = ElectricalData.objects.filter(
                tanggal__isnull=False
            ).values('tanggal')
            
            print(f"DEBUG - Records with valid dates: {all_electrical_data.count()}")
            
            # Dictionary untuk menyimpan count
            month_year_dict = {}
            
            for data in all_electrical_data:
                tanggal = data['tanggal']
                if tanggal:
                    try:
                        month = tanggal.month
                        year = tanggal.year
                        key = f"{year}-{month:02d}"
                        
                        if key in month_year_dict:
                            month_year_dict[key]['count'] += 1
                        else:
                            month_year_dict[key] = {
                                'month': month,
                                'year': year,
                                'count': 1
                            }
                    except Exception as date_error:
                        print(f"DEBUG - Error processing date {tanggal}: {date_error}")
                        continue
            
            print(f"DEBUG - Month/Year dict keys: {list(month_year_dict.keys())}")
            
            # Konversi ke list
            month_names = {
                1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
                5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
                9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
            }
            
            data_list = []
            for key, data in month_year_dict.items():
                month = data['month']
                year = data['year']
                count = data['count']
                
                data_list.append({
                    'month': month,
                    'year': year,
                    'count': count,
                    'month_name': month_names.get(month, f'Bulan {month}')
                })
            
            # Sorting
            current_date = datetime.now()
            current_month = current_date.month
            current_year = current_date.year
            
            def get_sort_priority(item):
                year = item['year']
                month = item['month']
                
                if year == current_year:
                    if month <= current_month:
                        return (0, current_month - month)
                    else:
                        return (1, month)
                elif year < current_year:
                    return (2, -year, -month)
                else:
                    return (3, year, month)
            
            data_list.sort(key=get_sort_priority)
            
            print(f"DEBUG - Final data list count: {len(data_list)}")
            
            context = {
                'month_year_data': data_list,
                'show_dates': False,
            }
            
        except Exception as list_error:
            print(f"DEBUG - Error creating month/year list: {list_error}")
            messages.error(request, f'Error saat mengambil daftar bulan/tahun: {list_error}')
            context = {
                'month_year_data': [],
                'show_dates': False,
            }
    
    return render(request, 'dailyactivity_app/tanggal_electrical.html', context)

def upload_machineutility_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']

        try:
            # Baca file Excel
            df = pd.read_excel(excel_file)

            # Pastikan kolom yang diperlukan ada di file Excel
            required_columns = ['name']
            if not all(column in df.columns for column in required_columns):
                messages.error(request, 'Excel file does not have the required columns.')
                return redirect('dailyactivity_app:machineutility_index')

            # Loop melalui setiap baris data di Excel dan simpan ke database
            for _, row in df.iterrows():
                Machineutility.objects.create(
                    name=row['name']
                )

            messages.success(request, 'Data uploaded successfully!')
        except Exception as e:
            messages.error(request, f'An error occurred: {e}')

    return redirect('dailyactivity_app:machineutility_index')



# @login_required 
# def utility_index(request):
#     shifts = Shift.objects.all()
#     status = Status.objects.all()
#     pic_utility = PICUtility2.objects.all()

#     # Ambil daftar nomor WO
#     nomor_wo_list = []
#     with connections['DB_Maintenance'].cursor() as cursor:
#         cursor.execute("""
#             SELECT number_wo, status_pekerjaan
#             FROM dbo.view_main
#             WHERE id_section = 6
#         AND YEAR(tgl_his) BETWEEN 2024 AND 2024
#         ORDER BY history_id DESC
#         """)
#         nomor_wo_list = [(row[0], row[1]) for row in cursor.fetchall()]  # Simpan hanya kolom number_wo dan status_pekerjaan

#     # Default untuk deskripsi_perbaikan, tgl_his, dan penyebab
#     deskripsi_perbaikan = None
#     tgl_his = None
#     penyebab = None
#     line = None
#     mesin = None
#     nomer = None
#     pekerjaan = None
#     status_pekerjaan = None
#     tindakan_perbaikan = None
#     tindakan_pencegahan = None

#     # Tangani Form Submission
#     if request.method == 'POST':
#         form = UtilityData2(request.POST, request.FILES)
#         if form.is_valid():
#             # Proses data form
#             machine_number = form.cleaned_data.get('machine_number')
#             machine_instance = form.cleaned_data.get('machine')

#             # Cek apakah mesin sudah memiliki nomor
#             if machine_number and machine_instance and not machine_instance.nomor:
#                 machine_instance.nomor = machine_number
#                 machine_instance.save()

#             # Simpan data MechanicalData
#             utility_data = form.save(commit=False)
#             utility_data.user = request.user
#             utility_data.machine = machine_instance
#             utility_data.save()

#             # Hubungkan PIC dengan MechanicalData
#             pic_ids = form.cleaned_data.get('pic')
#             utility_data.pic.set(pic_ids)

#             # Simpan nomor WO dan waktu pengerjaan
#             utility_data.nomor_wo = form.cleaned_data.get('nomor_wo')
#             utility_data.waktu_pengerjaan = form.cleaned_data.get('waktu_pengerjaan')
#             utility_data.save()
            
#             return redirect('success_page')
#     else:
#         form = UtilityData2Form()

#     # Tangani Permintaan AJAX
#     if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.method == 'GET':
#         nomor_wo_selected = request.GET.get('nomor_wo')
#         with connections['DB_Maintenance'].cursor() as cursor:
#             cursor.execute("""
#                 SELECT deskripsi_perbaikan, tgl_his, penyebab, line, mesin, nomer, pekerjaan, status_pekerjaan, tindakan_perbaikan, tindakan_pencegahan
#                 FROM dbo.view_main
#                 WHERE number_wo = %s
#             """, [nomor_wo_selected])
#             row = cursor.fetchone()
#             if row:
#                 deskripsi_perbaikan, tgl_his, penyebab, line, mesin, nomer, pekerjaan, status_pekerjaan, tindakan_perbaikan, tindakan_pencegahan = row

#         return JsonResponse({
#             'deskripsi_perbaikan': deskripsi_perbaikan,
#             'tgl_his': tgl_his,
#             'penyebab': penyebab,
#             'line': line,
#             'mesin': mesin,
#             'nomer': nomer,
#             'pekerjaan': pekerjaan,
#             'status_pekerjaan': status_pekerjaan,
#             'tindakan_perbaikan': tindakan_perbaikan,
#             'tindakan_pencegahan': tindakan_pencegahan
#         })

#     # Context untuk template
#     context = {
#         'shifts': shifts,
#         'status': status,
#         'pic_utility': pic_utility,
#         'nomor_wo_list': nomor_wo_list,
#         'deskripsi_perbaikan': deskripsi_perbaikan,
#         'tgl_his': tgl_his,
#         'penyebab': penyebab,  # Perbaikan typo dari 'penyabab' ke 'penyebab'
#         'line': line,
#         'mesin': mesin,
#         'nomer': nomer,
#         'pekerjaan': pekerjaan,
#         'status_pekerjaan': status_pekerjaan,
#         'tindakan_perbaikan': tindakan_perbaikan,
#         'tindakan_pencegahan': tindakan_pencegahan,
#         'form': form,
#     }
#     return render(request, 'dailyactivity_app/utility_index.html', context)

@login_required 
def utility_index(request):
    shifts = Shift.objects.all()
    locations = Location.objects.all()
    machines = Machineutility.objects.all()
    categories = Category.objects.all()
    status = Status.objects.all()
    pic_utility = PICUtility.objects.all()

    # Ambil hanya 30 nomor WO paling terbaru
    nomor_wo_list = []
    with connections['DB_Maintenance'].cursor() as cursor:
        cursor.execute("""
            SELECT TOP 30 number_wo
            FROM dbo.view_main
            WHERE id_section = 6
            AND YEAR(tgl_his) BETWEEN 2024 AND 2025
            ORDER BY history_id DESC
        """)
        nomor_wo_list = [row[0] for row in cursor.fetchall()]

    # Default values
    deskripsi_perbaikan = None
    tgl_his = None
    penyebab = None
    line = None
    mesin = None
    nomer = None
    pekerjaan = None
    status_pekerjaan = None
    tindakan_perbaikan = None

    # Tangani Permintaan AJAX untuk nomor WO
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.method == 'GET':
        nomor_wo_selected = request.GET.get('nomor_wo')
        
        # Prepare response data
        response_data = {
            'deskripsi_perbaikan': None,
            'tgl_his': None,
            'penyebab': None,
            'line': None,
            'mesin': None,
            'nomer': None,
            'pekerjaan': None,
            'status_pekerjaan': None,
            'tindakan_perbaikan': None,
            'location_id': None,
            'machine_id': None
        }
        
        with connections['DB_Maintenance'].cursor() as cursor:
            cursor.execute("""
                SELECT deskripsi_perbaikan, tgl_his, penyebab, line, mesin, nomer, 
                       pekerjaan, status_pekerjaan, tindakan_perbaikan
                FROM dbo.view_main
                WHERE number_wo = %s
            """, [nomor_wo_selected])
            row = cursor.fetchone()
            
            if row:
                (deskripsi_perbaikan, tgl_his, penyebab, line, mesin, 
                 nomer, pekerjaan, status_pekerjaan, tindakan_perbaikan) = row
                
                response_data.update({
                    'deskripsi_perbaikan': deskripsi_perbaikan,
                    'tgl_his': tgl_his,
                    'penyebab': penyebab,
                    'line': line,
                    'mesin': mesin,
                    'nomer': nomer,
                    'pekerjaan': pekerjaan,
                    'status_pekerjaan': status_pekerjaan,
                    'tindakan_perbaikan': tindakan_perbaikan
                })
                
                # TAMBAHAN: Cari location_id berdasarkan nama line
                if line:
                    try:
                        location_obj = Location.objects.filter(name__icontains=line).first()
                        if location_obj:
                            response_data['location_id'] = location_obj.id
                    except:
                        pass
                
                # TAMBAHAN: Cari machine_id berdasarkan nama mesin dan nomor
                if mesin:
                    try:
                        machine_query = Machineutility.objects.filter(name__icontains=mesin)
                        if nomer:
                            machine_query = machine_query.filter(nomor__icontains=nomer)
                        
                        machine_obj = machine_query.first()
                        if machine_obj:
                            response_data['machine_id'] = machine_obj.id
                    except:
                        pass

        return JsonResponse(response_data)

    # Context untuk template
    context = {
        'shifts': shifts,
        'locations': locations,
        'machines': machines,
        'categories': categories,
        'status': status,
        'pic_utility': pic_utility,
        'nomor_wo_list': nomor_wo_list,
        'deskripsi_perbaikan': deskripsi_perbaikan,
        'tgl_his': tgl_his,
        'penyebab': penyebab,
        'line': line,
        'mesin': mesin,
        'nomer': nomer,
        'pekerjaan': pekerjaan,
        'status_pekerjaan': status_pekerjaan,
        'tindakan_perbaikan': tindakan_perbaikan,
    }
    return render(request, 'dailyactivity_app/utility_index.html', context)

# @login_required  # Pastikan pengguna harus login untuk mengakses fungsi ini
# def utility_submit(request):
#     if request.method == 'POST':
#         # Ambil data dari formulir
#         tanggal = request.POST.get('tanggal')
#         jam = request.POST.get('jam')
#         tgl_his = request.POST.get('tgl_his')
#         shift_id = request.POST.get('shift')
#         # machine_id = request.POST.get('machine')
#         status_id = request.POST.get('status')
#         masalah = request.POST.get('masalah')
#         penyebab = request.POST.get('penyebab')  # Default None jika tidak diisi
#         line = request.POST.get('line')   # Default None jika tidak diisi
#         mesin = request.POST.get('mesin')  # Default None jika tidak diisi
#         nomer = request.POST.get('nomer') # Default None jika tidak diisi
#         pekerjaan = request.POST.get('pekerjaan')   # Default None jika tidak diisi
#         status_pekerjaan = request.POST.get('status_pekerjaan')   # Default None jika tidak diisi
#         tindakan_perbaikan = request.POST.get('tindakan_perbaikan')   # Default None jika tidak diisi
#         tindakan_pencegahan = request.POST.get('tindakan_pencegahan')   # Default None jika tidak diisi
#         # tindakan = request.POST.get('tindakan')
#         image = request.FILES.get('image')
#         pic_ids = request.POST.getlist('pic')  # Ambil daftar PIC yang dipilih
#         nomor_wo = request.POST.get('nomor_wo')  # Ambil nomor WO
#         waktu_pengerjaan = request.POST.get('waktu_pengerjaan')  # Ambil waktu pengerjaan

#         # Ambil instance Shift
#         try:
#             shift_instance = Shift.objects.get(id=shift_id)
#         except Shift.DoesNotExist:
#             messages.error(request, 'Shift tidak ditemukan.')
#             return redirect('dailyactivity_app:utility_index')
#         try:
#             status_instance = Status.objects.get(id=status_id)
#         except Status.DoesNotExist:
#             messages.error(request, 'Status tidak ditemukan.')
#             return redirect('dailyactivity_app:utility_index')

#         # Ambil user_id dari pengguna yang sedang login
#         user_id = request.user.id

#         jam_value = tgl_his if tgl_his else jam

#         # Simpan data ke database
#         utility_data = UtilityData2.objects.create(
#             tanggal=tanggal,
#             jam=jam_value,
#             shift=shift_instance,
#             # machine=machine_instance,
#             status=status_instance,
#             user_id=user_id,
#             masalah=masalah,
#             penyebab=penyebab,
#             line=line,
#             mesin=mesin,
#             nomer=nomer,
#             pekerjaan=pekerjaan,
#             status_pekerjaan=status_pekerjaan,
#             # tindakan=tindakan,
#             tindakan_perbaikan=tindakan_perbaikan,
#             tindakan_pencegahan=tindakan_pencegahan,
#             image=image,
#             nomor_wo=nomor_wo,
#             waktu_pengerjaan=waktu_pengerjaan
#         )
#         # Menyimpan PIC yang dipilih
#         for pic_id in pic_ids:
#             try:
#                 pic_instance = PICUtility2.objects.get(id=pic_id)
#                 utility_data.pic.add(pic_instance)
#             except PICUtility2.DoesNotExist:
#                 continue

#         # Simpan pesan sukses
#         messages.success(request, 'Data berhasil disimpan!')

#         # Redirect ke mechanical_index
#         return redirect('dailyactivity_app:utility_index')

#     return redirect('dailyactivity_app:utility_index')

@login_required  
def utility_submit(request):
    if request.method == 'POST':
        try:
            # Function to safely convert empty string to None, then to int
            def safe_int_or_none(value):
                if not value or value == '' or value == 'None':
                    return None
                try:
                    return int(value)
                except (ValueError, TypeError):
                    return None
            
            # Function to safely get string value
            def safe_string(value):
                return value.strip() if value and value.strip() else None

            # Ambil dan clean semua data dari formulir
            tanggal = request.POST.get('tanggal')
            jam = request.POST.get('jam')
            tgl_his = request.POST.get('tgl_his')
            
            # Convert ALL ID fields safely
            shift_id = safe_int_or_none(request.POST.get('shift'))
            location_id = safe_int_or_none(request.POST.get('location'))
            machine_id = safe_int_or_none(request.POST.get('machine'))
            category_id = safe_int_or_none(request.POST.get('category'))
            status_id = safe_int_or_none(request.POST.get('status'))
            
            # Text fields
            masalah = request.POST.get('masalah', '').strip()
            penyebab = request.POST.get('penyebab', '').strip()
            tindakan_perbaikan = request.POST.get('tindakan_perbaikan', '').strip()
            nomor_wo = safe_string(request.POST.get('nomor_wo'))
            waktu_pengerjaan = safe_string(request.POST.get('waktu_pengerjaan'))
            machine_number = safe_string(request.POST.get('machine_number'))
            
            # Manual machine inputs
            manual_machine_name = safe_string(request.POST.get('manual_machine_name'))
            manual_machine_number = safe_string(request.POST.get('manual_machine_number'))
            
            # File
            image = request.FILES.get('image')
            
            # PIC IDs - filter empty values
            pic_ids_raw = request.POST.getlist('pic')
            pic_ids = [safe_int_or_none(pic_id) for pic_id in pic_ids_raw if safe_int_or_none(pic_id) is not None]

            print(f"DEBUG - shift_id: {shift_id}, category_id: {category_id}, status_id: {status_id}")
            print(f"DEBUG - location_id: {location_id}, machine_id: {machine_id}")
            print(f"DEBUG - manual_machine_name: {manual_machine_name}")

            # Validasi REQUIRED fields dengan proper error messages
            if not shift_id:
                messages.error(request, 'Shift harus dipilih.')
                return redirect('dailyactivity_app:utility_index')
                
            if not category_id:
                messages.error(request, 'Jenis Pekerjaan harus dipilih.')
                return redirect('dailyactivity_app:utility_index')
                
            if not status_id:
                messages.error(request, 'Status harus dipilih.')
                return redirect('dailyactivity_app:utility_index')

            if not masalah:
                messages.error(request, 'Masalah harus diisi.')
                return redirect('dailyactivity_app:utility_index')

            if not penyebab:
                messages.error(request, 'Penyebab harus diisi.')
                return redirect('dailyactivity_app:utility_index')

            if not tindakan_perbaikan:
                messages.error(request, 'Tindakan Perbaikan harus diisi.')
                return redirect('dailyactivity_app:utility_index')

            # Get required instances
            try:
                shift_instance = Shift.objects.get(id=shift_id)
                category_instance = Category.objects.get(id=category_id)
                status_instance = Status.objects.get(id=status_id)
            except (Shift.DoesNotExist, Category.DoesNotExist, Status.DoesNotExist) as e:
                messages.error(request, f'Data referensi tidak ditemukan: {str(e)}')
                return redirect('dailyactivity_app:utility_index')

            # HANDLE LOCATION - ALWAYS CREATE ONE
            location_instance = None
            if location_id:
                try:
                    location_instance = Location.objects.get(id=location_id)
                except Location.DoesNotExist:
                    messages.error(request, 'Location tidak ditemukan.')
                    return redirect('dailyactivity_app:utility_index')
            
            # Create default location if none selected
            if not location_instance:
                location_instance, created = Location.objects.get_or_create(
                    name="Unknown Location",
                    defaults={'name': "Unknown Location"}
                )
                print(f"DEBUG - Created/found default location: {location_instance.id}")

            # HANDLE MACHINE - ALWAYS CREATE ONE
            machine_instance = None
            
            # Option 1: Machine selected from dropdown
            if machine_id:
                try:
                    machine_instance = Machineutility.objects.get(id=machine_id)
                    print(f"DEBUG - Found machine from dropdown: {machine_instance.id}")
                    
                    # Update machine number if provided and machine doesn't have one
                    if machine_number and not machine_instance.nomor:
                        machine_instance.nomor = machine_number
                        machine_instance.save()
                        
                except Machineutility.DoesNotExist:
                    print(f"DEBUG - Machine ID {machine_id} not found")
                    machine_instance = None
            
            # Option 2: Manual machine input
            if not machine_instance and manual_machine_name:
                try:
                    print(f"DEBUG - Trying to create/find manual machine: {manual_machine_name}")
                    
                    # Check if machine already exists
                    machine_instance = Machineutility.objects.filter(
                        name__iexact=manual_machine_name
                    ).first()
                    
                    if not machine_instance:
                        # Create new machine
                        machine_instance = Machineutility.objects.create(
                            name=manual_machine_name,
                            location=location_instance,
                            nomor=manual_machine_number
                        )
                        print(f"DEBUG - Created new machine: {machine_instance.id}")
                    else:
                        # Update existing machine number if needed
                        if manual_machine_number and not machine_instance.nomor:
                            machine_instance.nomor = manual_machine_number
                            machine_instance.save()
                        print(f"DEBUG - Found existing machine: {machine_instance.id}")
                        
                except Exception as e:
                    print(f"DEBUG - Error creating manual machine: {e}")
                    machine_instance = None
            
            # Option 3: Create default machine if still none
            if not machine_instance:
                try:
                    machine_instance, created = Machineutility.objects.get_or_create(
                        name="Unknown Machine",
                        location=location_instance,
                        defaults={
                            'name': "Unknown Machine",
                            'location': location_instance,
                            'nomor': None
                        }
                    )
                    print(f"DEBUG - Created/found default machine: {machine_instance.id}, created: {created}")
                except Exception as e:
                    print(f"DEBUG - Error creating default machine: {e}")
                    # Last resort with timestamp to avoid conflicts
                    import time
                    timestamp = str(int(time.time()))
                    machine_instance = Machineutility.objects.create(
                        name=f"Unknown Machine {timestamp}",
                        location=location_instance,
                        nomor=None
                    )
                    print(f"DEBUG - Created timestamped machine: {machine_instance.id}")

            # Pastikan semua instance yang dibutuhkan ada
            if not all([shift_instance, location_instance, machine_instance, category_instance, status_instance]):
                missing = []
                if not shift_instance: missing.append("shift")
                if not location_instance: missing.append("location")
                if not machine_instance: missing.append("machine")
                if not category_instance: missing.append("category") 
                if not status_instance: missing.append("status")
                
                messages.error(request, f'Instance tidak lengkap: {", ".join(missing)}')
                return redirect('dailyactivity_app:utility_index')

            from datetime import datetime

            # Di dalam function utility_submit, ganti bagian processing jam jadi:
            jam_value = None
            if tgl_his:
                try:
                    # Parse datetime string dan ambil time portion
                    if 'T' in str(tgl_his):
                        time_str = str(tgl_his).split('T')[1]  # Ambil bagian setelah 'T'
                        if '.' in time_str:
                            time_str = time_str.split('.')[0]  # Remove microseconds
                        jam_value = time_str
                    else:
                        jam_value = tgl_his
                except Exception as e:
                    print(f"ERROR parsing tgl_his: {e}")
                    jam_value = None
            elif jam:
                jam_value = jam
            print(f"DEBUG - About to save with: shift={shift_instance.id}, location={location_instance.id}, machine={machine_instance.id}")

            # SAVE DATA TO DATABASE
            utility_data = UtilityData.objects.create(
                tanggal=tanggal,
                jam=jam_value,
                shift=shift_instance,
                location=location_instance,
                machine=machine_instance,
                category=category_instance,
                status=status_instance,
                user=request.user,  # Use request.user instead of user_id
                masalah=masalah,
                penyebab=penyebab,
                tindakan=tindakan_perbaikan,
                image=image,
                nomor_wo=nomor_wo,
                waktu_pengerjaan=waktu_pengerjaan
            )

            print(f"DEBUG - Utility data created with ID: {utility_data.id}")

            # Add PICs
            for pic_id in pic_ids:
                try:
                    pic_instance = PICUtility.objects.get(id=pic_id)
                    utility_data.pic.add(pic_instance)
                    print(f"DEBUG - Added PIC: {pic_id}")
                except PICUtility.DoesNotExist:
                    print(f"DEBUG - PIC not found: {pic_id}")
                    continue

            messages.success(request, 'Data berhasil disimpan!')
            return redirect('dailyactivity_app:utility_index')

        except Exception as e:
            print(f"ERROR - Exception in utility_submit: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
            return redirect('dailyactivity_app:utility_index')

    return redirect('dailyactivity_app:utility_index')

# @login_required
# def edit_utility_data(request, id):
#     # Ambil data yang akan diedit berdasarkan id
#     utility_data = get_object_or_404(UtilityData2, id=id)
#     shifts = Shift.objects.all()
#     # locations = Location.objects.all()
#     # machines = Machinemechanical.objects.all()
#     # categories = Category.objects.all()
#     status = Status.objects.all()
#     pic_utility = PICUtility2.objects.all()

#     if request.method == 'POST':
#         # Memproses form data yang telah diisi ulang
#         form = UtilityData2Form(request.POST, request.FILES, instance=utility_data)
#         if form.is_valid():
#             updated_data = form.save(commit=False)
#             updated_data.user = request.user
#             updated_data.nomor_wo = form.cleaned_data.get('nomor_wo')
#             updated_data.waktu_pengerjaan = form.cleaned_data.get('waktu_pengerjaan')
#             updated_data.line = form.cleaned_data.get('line')
#             updated_data.mesin = form.cleaned_data.get('mesin')
#             updated_data.nomer = form.cleaned_data.get('nomer')
#             updated_data.pekerjaan = form.cleaned_data.get('pekerjaan')

#             updated_data.save()

#             # Update PIC terkait
#             pic_ids = form.cleaned_data.get('pic')
#             updated_data.pic.set(pic_ids)

#             messages.success(request, 'Data berhasil diperbarui!')
#             return redirect('dailyactivity_app:data_utility', tanggal=updated_data.tanggal.strftime('%Y-%m-%d'))
#         else:
#             messages.error(request, 'Terjadi kesalahan saat memperbarui data. Periksa kembali isian Anda.')

#     else:
#         # Memuat form dengan data yang ada untuk ditampilkan di template
#         form = UtilityData2Form(instance=utility_data)

#     context = {
#         'form': form,
#         'shifts': shifts,
#         # 'locations': locations,
#         # 'machines': machines,
#         # 'categories': categories,
#         'status': status,
#         'pic_utility': pic_utility,
#         'data': utility_data,
#         'tanggal': utility_data.tanggal,
#         'jam': utility_data.jam,
#         'nomor_wo': utility_data.nomor_wo,
#         'waktu_pengerjaan': utility_data.waktu_pengerjaan,
#         'line': utility_data.line,
#         'mesin': utility_data.mesin,
#         'nomer': utility_data.nomer,
#         'pekerjaan': utility_data.pekerjaan,
#         'pic': utility_data.pic.all(),
#     }
#     return render(request, 'dailyactivity_app/edit_utility_data.html', context)

@login_required
def edit_utility_data(request, id):
    """
    Edit data mechanical untuk model MechanicalData dengan design modern
    """
    # Ambil data yang akan diedit berdasarkan id - gunakan MechanicalData
    utility_data = get_object_or_404(UtilityData, id=id)
    
    # Ambil semua data referensi
    shifts = Shift.objects.all()
    locations = Location.objects.all()
    machines = Machineutility.objects.all()
    categories = Category.objects.all()
    status = Status.objects.all()
    pic_utility = PICUtility.objects.all()  # Gunakan PICMechanical untuk MechanicalData

    if request.method == 'POST':
        try:
            # Function to safely convert empty string to None, then to int
            def safe_int_or_none(value):
                if not value or value == '' or value == 'None':
                    return None
                try:
                    return int(value)
                except (ValueError, TypeError):
                    return None
            
            # Function to safely get string value
            def safe_string(value):
                return value.strip() if value and value.strip() else None

            # Ambil dan clean semua data dari formulir
            tanggal = request.POST.get('tanggal')
            jam = request.POST.get('jam')
            
            # Convert ALL ID fields safely
            shift_id = safe_int_or_none(request.POST.get('shift'))
            location_id = safe_int_or_none(request.POST.get('location'))
            machine_id = safe_int_or_none(request.POST.get('machine'))
            category_id = safe_int_or_none(request.POST.get('category'))
            status_id = safe_int_or_none(request.POST.get('status'))
            
            # Text fields
            masalah = request.POST.get('masalah', '').strip()
            penyebab = request.POST.get('penyebab', '').strip()
            tindakan_perbaikan = request.POST.get('tindakan_perbaikan', '').strip()
            nomor_wo = safe_string(request.POST.get('nomor_wo'))
            waktu_pengerjaan = safe_string(request.POST.get('waktu_pengerjaan'))
            
            # File
            image = request.FILES.get('image')
            
            # PIC IDs - filter empty values
            pic_ids_raw = request.POST.getlist('pic')
            pic_ids = [safe_int_or_none(pic_id) for pic_id in pic_ids_raw if safe_int_or_none(pic_id) is not None]

            print(f"DEBUG EDIT - shift_id: {shift_id}, category_id: {category_id}, status_id: {status_id}")
            print(f"DEBUG EDIT - location_id: {location_id}, machine_id: {machine_id}")

            # Validasi REQUIRED fields
            if not shift_id:
                messages.error(request, 'Shift harus dipilih.')
                return redirect('dailyactivity_app:edit_utility_data', id=id)
                
            if not category_id:
                messages.error(request, 'Jenis Pekerjaan harus dipilih.')
                return redirect('dailyactivity_app:edit_utility_data', id=id)
                
            if not status_id:
                messages.error(request, 'Status harus dipilih.')
                return redirect('dailyactivity_app:edit_utility_data', id=id)

            if not masalah:
                messages.error(request, 'Masalah harus diisi.')
                return redirect('dailyactivity_app:edit_utility_data', id=id)

            if not penyebab:
                messages.error(request, 'Penyebab harus diisi.')
                return redirect('dailyactivity_app:edit_utility_data', id=id)

            if not tindakan_perbaikan:
                messages.error(request, 'Tindakan Perbaikan harus diisi.')
                return redirect('dailyactivity_app:edit_utility_data', id=id)

            # Get required instances
            try:
                shift_instance = Shift.objects.get(id=shift_id)
                category_instance = Category.objects.get(id=category_id)
                status_instance = Status.objects.get(id=status_id)
            except (Shift.DoesNotExist, Category.DoesNotExist, Status.DoesNotExist) as e:
                messages.error(request, f'Data referensi tidak ditemukan: {str(e)}')
                return redirect('dailyactivity_app:edit_utility_data', id=id)

            # Handle Location - bisa None
            location_instance = None
            if location_id:
                try:
                    location_instance = Location.objects.get(id=location_id)
                except Location.DoesNotExist:
                    messages.error(request, 'Location tidak ditemukan.')
                    return redirect('dailyactivity_app:edit_utility_data', id=id)
            
            # Kalau location kosong, pakai yang lama atau buat default
            if not location_instance:
                location_instance = utility_data.location  # Keep existing location
                if not location_instance:
                    # Create default location if needed
                    location_instance, created = Location.objects.get_or_create(
                        name="Unknown Location",
                        defaults={'name': "Unknown Location"}
                    )

            # Handle Machine - bisa None
            machine_instance = None
            if machine_id:
                try:
                    machine_instance = Machineutility.objects.get(id=machine_id)
                except Machineutility.DoesNotExist:
                    messages.error(request, 'Machine tidak ditemukan.')
                    return redirect('dailyactivity_app:edit_utility_data', id=id)
            
            # Kalau machine kosong, pakai yang lama atau buat default  
            if not machine_instance:
                machine_instance = utility_data.machine  # Keep existing machine
                if not machine_instance:
                    # Create default machine if needed
                    machine_instance, created = Machineutility.objects.get_or_create(
                        name="Unknown Machine",
                        location=location_instance,
                        defaults={
                            'name': "Unknown Machine",
                            'location': location_instance,
                            'nomor': None
                        }
                    )

            # Process jam field
            jam_value = None
            if jam:
                jam_value = jam

            print(f"DEBUG EDIT - About to update with: shift={shift_instance.id}, location={location_instance.id}, machine={machine_instance.id}")

            # UPDATE DATA 
            utility_data.tanggal = tanggal
            utility_data.jam = jam_value
            utility_data.shift = shift_instance
            utility_data.location = location_instance
            utility_data.machine = machine_instance
            utility_data.category = category_instance
            utility_data.status = status_instance
            utility_data.masalah = masalah
            utility_data.penyebab = penyebab
            utility_data.tindakan = tindakan_perbaikan  # Backward compatibilit
            utility_data.nomor_wo = nomor_wo
            utility_data.waktu_pengerjaan = waktu_pengerjaan
            
            # Update image only if new file uploaded
            if image:
                utility_data.image = image
                
            utility_data.save()

            print(f"DEBUG EDIT - Utility data updated with ID: {utility_data.id}")

            # Update PICs - clear existing and add new ones
            utility_data.pic.clear()
            for pic_id in pic_ids:
                try:
                    pic_instance = PICUtility.objects.get(id=pic_id)
                    utility_data.pic.add(pic_instance)
                    print(f"DEBUG EDIT - Added PIC: {pic_id}")
                except PICUtility.DoesNotExist:
                    print(f"DEBUG EDIT - PIC not found: {pic_id}")
                    continue

            messages.success(request, 'Data berhasil diperbarui!')
            return redirect('dailyactivity_app:data_utility', tanggal=utility_data.tanggal.strftime('%Y-%m-%d'))

        except Exception as e:
            print(f"ERROR EDIT - Exception in edit_utility_data: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Terjadi kesalahan saat memperbarui data: {str(e)}')
            return redirect('dailyactivity_app:edit_utility_data', id=id)

    # Context untuk GET request
    context = {
        'shifts': shifts,
        'locations': locations,
        'machines': machines,
        'categories': categories,
        'status': status,
        'pic_utility': pic_utility,
        'data': utility_data,
    }
    
    return render(request, 'dailyactivity_app/edit_utility_data.html', context)

# @login_required
# def delete_utility_data(request, id):
#     # Mengambil data berdasarkan ID
#     utility_data = get_object_or_404(UtilityData2, id=id)

#     if request.method == 'POST':
#         # Hapus data dari database
#         utility_data.delete()
#         messages.success(request, 'Data berhasil dihapus!')
#         return redirect('dailyactivity_app:data_utility', tanggal=utility_data.tanggal.strftime('%Y-%m-%d'))
    
#     # Jika bukan POST, redirect ke halaman data electrical
#     return redirect('dailyactivity_app:data_utility', tanggal=utility_data.tanggal.strftime('%Y-%m-%d'))

@login_required
def delete_utility_data(request, id):
    """
    Delete data utility untuk model MechanicalData
    """
    # Mengambil data berdasarkan ID - gunakan MechanicalData
    utility_data = get_object_or_404(UtilityData, id=id)
    
    if request.method == 'POST':
        try:
            # Simpan tanggal sebelum dihapus untuk redirect
            tanggal_redirect = utility_data.tanggal.strftime('%Y-%m-%d')
            
            # Log informasi sebelum hapus
            print(f"DEBUG DELETE - Deleting utility_data ID: {id}")
            print(f"DEBUG DELETE - Data: {utility_data.tanggal} - {utility_data.masalah[:50]}...")
            
            # Hapus data dari database
            utility_data.delete()
            
            messages.success(request, 'Data utility_data berhasil dihapus!')
            print(f"DEBUG DELETE - Successfully deleted utility_data ID: {id}")
            
            return redirect('dailyactivity_app:data_utility', tanggal=tanggal_redirect)
            
        except Exception as e:
            print(f"ERROR DELETE - Exception in delete_utility_data: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Terjadi kesalahan saat menghapus data: {str(e)}')
            
            # Redirect ke data mechanical dengan tanggal yang sama
            tanggal_redirect = utility_data.tanggal.strftime('%Y-%m-%d')
            return redirect('dailyactivity_app:data_utility', tanggal=tanggal_redirect)
    
    # Jika bukan POST request, redirect ke data mechanical
    tanggal_redirect = utility_data.tanggal.strftime('%Y-%m-%d')
    return redirect('dailyactivity_app:data_utility', tanggal=tanggal_redirect)

def get_machines_by_location_utility(request, location_id):
    # Mengambil data mesin berdasarkan location_id
    machines = Machineutility.objects.filter(location_id=location_id)
    
    # Mengambil hanya beberapa field yang relevan
    machine_list = list(machines.values('id', 'name', 'location_id', 'nomor'))
    
    # Mengembalikan response dalam format JSON
    return JsonResponse(machine_list, safe=False)

def get_machine_number_utility(request, machine_id):
    try: 
        # Mengambil mesin berdasarkan ID
        machine = Machineutility.objects.get(id=machine_id)
        # Mengembalikan nomor mesin
        return JsonResponse({'nomor': machine.nomor})
    except Machineutility.DoesNotExist:
        return JsonResponse({'error': 'Machine not found'}, status=404)


# @login_required
# def tanggal_utility(request):
#     # Mengambil tanggal-tanggal unik dan mengurutkan berdasarkan tanggal secara descending
#     dates = UtilityData2.objects.annotate(
#         date=TruncDate('tanggal', output_field=DateField())  # Truncate to date only
#     ).values('date').distinct().order_by('-date')  # Order by date descending

#     context = {
#         'dates': dates,
#     }
#     return render(request, 'dailyactivity_app/tanggal_utility.html', context)

@login_required
def tanggal_utility(request):
    """
    View untuk menampilkan tanggal-tanggal yang ada data utility & laporan utility
    Menggabungkan data dari utilityData dan LaporanutilityData
    """
    # Ambil parameter bulan dan tahun dari URL jika ada
    selected_month = request.GET.get('month')
    selected_year = request.GET.get('year')
    
    print(f"DEBUG - Raw parameters: month='{selected_month}', year='{selected_year}'")
    
    if selected_month and selected_year:
        try:
            # Fungsi untuk membersihkan dan konversi nilai
            def clean_and_convert(value):
                if value is None:
                    return None
                
                str_value = str(value).strip()
                
                if not str_value:
                    return None
                
                if '.' in str_value:
                    integer_part = str_value.split('.')[0]
                    return int(integer_part) if integer_part.isdigit() else None
                
                if str_value.isdigit():
                    return int(str_value)
                
                try:
                    return int(float(str_value))
                except (ValueError, TypeError):
                    return None
            
            # Konversi parameter
            selected_month = clean_and_convert(selected_month)
            selected_year = clean_and_convert(selected_year)
            
            print(f"DEBUG - Cleaned: month={selected_month}, year={selected_year}")
            
            # Validasi hasil konversi
            if selected_month is None or selected_year is None:
                messages.error(request, 'Parameter bulan atau tahun tidak dapat dikonversi')
                return redirect('dailyactivity_app:tanggal_utility')
            
            # Validasi range
            if not (1 <= selected_month <= 12):
                messages.error(request, f'Bulan harus antara 1-12, diterima: {selected_month}')
                return redirect('dailyactivity_app:tanggal_utility')
                
            current_year = datetime.now().year
            if not (2020 <= selected_year <= current_year + 2):
                messages.error(request, f'Tahun harus antara 2020-{current_year + 2}, diterima: {selected_year}')
                return redirect('dailyactivity_app:tanggal_utility')
            
            # Query data dari utilityData
            try:
                utility_dates = UtilityData.objects.filter(
                    tanggal__month=selected_month,
                    tanggal__year=selected_year
                ).annotate(
                    date=TruncDate('tanggal', output_field=DateField())
                ).values('date').distinct()
                
                # Query data dari LaporanUtility
                laporan_dates = LaporanData.objects.filter(
                    tanggal__month=selected_month,
                    tanggal__year=selected_year
                ).annotate(
                    date=TruncDate('tanggal', output_field=DateField())
                ).values('date').distinct()
                
                print(f"DEBUG - utilityData dates: {utility_dates.count()}")
                print(f"DEBUG - LaporanUtility dates: {laporan_dates.count()}")
                
                # Gabungkan tanggal dan hitung jumlah data per tanggal
                combined_dates = {}
                
                # Proses utilityData
                for date_obj in utility_dates:
                    date_key = date_obj['date']
                    if date_key not in combined_dates:
                        combined_dates[date_key] = {
                            'date': date_key,
                            'utility_count': 0,
                            'laporan_count': 0,
                            'total_count': 0
                        }
                    
                    # Hitung jumlah data utility untuk tanggal ini
                    utility_count = UtilityData.objects.filter(
                        tanggal=date_key
                    ).count()
                    combined_dates[date_key]['utility_count'] = utility_count
                
                # Proses LaporanutilityData
                for date_obj in laporan_dates:
                    date_key = date_obj['date']
                    if date_key not in combined_dates:
                        combined_dates[date_key] = {
                            'date': date_key,
                            'utility_count': 0,
                            'laporan_count': 0,
                            'total_count': 0
                        }
                    
                    # Hitung jumlah data laporan untuk tanggal ini
                    laporan_count = LaporanData.objects.filter(
                        tanggal=date_key
                    ).count()
                    combined_dates[date_key]['laporan_count'] = laporan_count
                
                # Hitung total count dan convert ke list
                dates_list = []
                for date_key, data in combined_dates.items():
                    data['total_count'] = data['utility_count'] + data['laporan_count']
                    dates_list.append(data)
                
                # Sort berdasarkan tanggal descending
                dates_list.sort(key=lambda x: x['date'], reverse=True)
                
                total_dates = len(dates_list)
                print(f"DEBUG - Combined dates: {total_dates}")
                
            except Exception as query_error:
                print(f"DEBUG - Query error: {query_error}")
                messages.error(request, f'Error saat mengambil data: {query_error}')
                return redirect('dailyactivity_app:tanggal_utility')
            
            # Nama bulan
            month_names = {
                1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
                5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
                9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
            }
            
            context = {
                'dates': dates_list,
                'selected_month': selected_month,
                'selected_year': selected_year,
                'selected_month_name': month_names.get(selected_month, f'Bulan {selected_month}'),
                'show_dates': True,
                'total_utility': sum(item['utility_count'] for item in dates_list),
                'total_laporan': sum(item['laporan_count'] for item in dates_list),
                'total_combined': sum(item['total_count'] for item in dates_list),
            }
            
        except Exception as e:
            print(f"DEBUG - General exception: {e}")
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
            return redirect('dailyactivity_app:tanggal_utility')
    else:
        print("DEBUG - Showing month/year list")
        
        try:
            # Cek data dari kedua model
            utility_count = UtilityData.objects.count()
            laporan_count = LaporanData.objects.count()
            total_count = utility_count + laporan_count
            
            print(f"DEBUG - UtilityData records: {utility_count}")
            print(f"DEBUG - LaporanData records: {laporan_count}")
            print(f"DEBUG - Total records: {total_count}")
            
            if total_count == 0:
                messages.info(request, 'Belum ada data utility atau laporan dalam sistem')
                context = {
                    'month_year_data': [],
                    'show_dates': False,
                }
                return render(request, 'dailyactivity_app/tanggal_utility.html', context)
            
            # Dictionary untuk menyimpan count gabungan
            month_year_dict = {}
            
            # Proses data utilityData
            utility_data = UtilityData.objects.filter(
                tanggal__isnull=False
            ).values('tanggal')
            
            for data in utility_data:
                tanggal = data['tanggal']
                if tanggal:
                    try:
                        month = tanggal.month
                        year = tanggal.year
                        key = f"{year}-{month:02d}"
                        
                        if key in month_year_dict:
                            month_year_dict[key]['utility_count'] += 1
                        else:
                            month_year_dict[key] = {
                                'month': month,
                                'year': year,
                                'utility_count': 1,
                                'laporan_count': 0,
                                'total_count': 1
                            }
                    except Exception as date_error:
                        print(f"DEBUG - Error processing utility date {tanggal}: {date_error}")
                        continue
            
            # Proses data LaporanutilityData
            laporan_data = LaporanData.objects.filter(
                tanggal__isnull=False
            ).values('tanggal')
            
            for data in laporan_data:
                tanggal = data['tanggal']
                if tanggal:
                    try:
                        month = tanggal.month
                        year = tanggal.year
                        key = f"{year}-{month:02d}"
                        
                        if key in month_year_dict:
                            month_year_dict[key]['laporan_count'] += 1
                            month_year_dict[key]['total_count'] += 1
                        else:
                            month_year_dict[key] = {
                                'month': month,
                                'year': year,
                                'utility_count': 0,
                                'laporan_count': 1,
                                'total_count': 1
                            }
                    except Exception as date_error:
                        print(f"DEBUG - Error processing laporan date {tanggal}: {date_error}")
                        continue
            
            print(f"DEBUG - Month/Year dict keys: {list(month_year_dict.keys())}")
            
            # Konversi ke list
            month_names = {
                1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
                5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
                9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
            }
            
            data_list = []
            for key, data in month_year_dict.items():
                month = data['month']
                year = data['year']
                utility_count = data['utility_count']
                laporan_count = data['laporan_count']
                total_count = data['total_count']
                
                data_list.append({
                    'month': month,
                    'year': year,
                    'utility_count': utility_count,
                    'laporan_count': laporan_count,
                    'total_count': total_count,
                    'month_name': month_names.get(month, f'Bulan {month}')
                })
            
            # Sorting
            current_date = datetime.now()
            current_month = current_date.month
            current_year = current_date.year
            
            def get_sort_priority(item):
                year = item['year']
                month = item['month']
                
                if year == current_year:
                    if month <= current_month:
                        return (0, current_month - month)
                    else:
                        return (1, month)
                elif year < current_year:
                    return (2, -year, -month)
                else:
                    return (3, year, month)
            
            data_list.sort(key=get_sort_priority)
            
            print(f"DEBUG - Final data list count: {len(data_list)}")
            
            context = {
                'month_year_data': data_list,
                'show_dates': False,
                'total_utility_all': utility_count,
                'total_laporan_all': laporan_count,
                'total_combined_all': total_count,
            }
            
        except Exception as list_error:
            print(f"DEBUG - Error creating month/year list: {list_error}")
            messages.error(request, f'Error saat mengambil daftar bulan/tahun: {list_error}')
            context = {
                'month_year_data': [],
                'show_dates': False,
            }
    
    return render(request, 'dailyactivity_app/tanggal_utility.html', context)



# @login_required
# def data_utility(request, tanggal):
#     # Parsing tanggal dari URL
#     tanggal_parsed = parse_date(tanggal)
#     # Menyaring data berdasarkan tanggal yang dipilih
#     utility_data = UtilityData2.objects.filter(tanggal=tanggal_parsed)
#     context = {
#         'utility_data': utility_data,
#         'selected_date': tanggal_parsed,
#     }
#     return render(request, 'dailyactivity_app/data_utility.html', context)

@login_required
def data_utility(request, tanggal):
    """
    View buat nampilir data utility dan laporan utility dalam satu halaman
    Kombinasi data dari UtilityData (bukan UtilityData2) dan LaporanData
    
    CATATAN STRUKTUR MODEL:
    - UtilityData: Punya field location, machine, category (ForeignKey), tindakan (TextField)
    - LaporanData: Punya field detail_pekerjaan (related from DetailPekerjaan model)
    """
    from datetime import datetime
    
    # Parsing tanggal dari URL
    tanggal_parsed = parse_date(tanggal)
    
    # Ambil data UtilityData (data utility biasa)
    # FIXED: Sekarang pake UtilityData dengan select_related yang sesuai
    utility_data = UtilityData.objects.filter(
        tanggal=tanggal_parsed
    ).select_related(
        'shift', 'status', 'user', 'location', 'machine', 'category'  # Field ForeignKey yang ada di UtilityData
    ).prefetch_related(
        'pic'
    ).order_by('-jam', '-id')
    
    # Ambil data LaporanData (data laporan utility)
    laporan_data = LaporanData.objects.filter(
        tanggal=tanggal_parsed
    ).select_related(
        'shift', 'user'
    ).prefetch_related(
        'pic',  # untuk PIC laporan
        'piclembur',  # untuk PIC lembur
        'detail_pekerjaan'  # Field yang benar adalah 'detail_pekerjaan', bukan 'laporan_pekerjaan'
    ).order_by('-id')
    
    # Gabungin data jadi unified structure buat template
    combined_data = []
    
    # Proses data utility - FIXED untuk UtilityData
    for data in utility_data:
        combined_data.append({
            'id': data.id,
            'type': 'utility',  # buat identifier
            'tanggal': data.tanggal,
            'jam': data.jam,
            'shift': data.shift,
            'nomor_wo': data.nomor_wo,
            'waktu_pengerjaan': data.waktu_pengerjaan,
            
            # MAPPING FIELD SESUAI UtilityData STRUCTURE
            'line': data.location.name if data.location else None,  # ambil dari location.name
            'mesin': data.machine.name if data.machine else None,  # ambil dari machine.name
            'nomer': data.machine.nomor if data.machine and data.machine.nomor else None,  # ambil dari machine.nomor
            
            # FIELD YANG ADA DI UtilityData
            'masalah': data.masalah,
            'penyebab': data.penyebab,
            'tindakan': data.tindakan,  # Field tindakan ADA di UtilityData
            
            # FIELD YANG GAK ADA DI UtilityData - set None atau default value
            'pekerjaan': None,  # Field pekerjaan GAK ADA di UtilityData
            'status_pekerjaan': None,  # Field status_pekerjaan GAK ADA di UtilityData
            'tindakan_perbaikan': data.tindakan,  # Map ke field tindakan yang ada
            'tindakan_pencegahan': None,  # Field tindakan_pencegahan GAK ADA di UtilityData
            
            # FIELD YANG ADA DI UtilityData
            'status': data.status,
            'image': data.image,
            'pic': data.pic.all(),
            'user': data.user,
            
            # FIELD YANG GAK ADA - set default
            'pic_lembur': None,  # utility data ga punya pic lembur
            'catatan': None,  # utility data ga punya catatan
            'lama_pekerjaan': None,
            'pic_masalah': None,
            'jenis_pekerjaan': None,
            
            # FIELD YANG ADA DI UtilityData - pass as objects
            'location': data.location,  # Field ADA di UtilityData
            'machine': data.machine,  # Field ADA di UtilityData
            'category': data.category,  # Field ADA di UtilityData
        })
    
    # Proses data LaporanData - SAME AS BEFORE
    for laporan in laporan_data:
        # Ambil detail pekerjaan kalo ada
        detail_pekerjaan = []
        pekerjaan_list = []
        jenis_pekerjaan_list = []
        
        # Akses field yang benar: 'detail_pekerjaan' bukan 'laporan_pekerjaan'
        try:
            detail_pekerjaan = laporan.detail_pekerjaan.all()
            if detail_pekerjaan.exists():
                for detail in detail_pekerjaan:
                    pekerjaan_list.append(detail.deskripsi)  # Field 'deskripsi' di DetailPekerjaan
                    jenis_pekerjaan_list.append(detail.jenis_pekerjaan)
        except Exception as e:
            print(f"Error accessing detail_pekerjaan: {e}")
        
        # Gabungin masalah utama dengan detail pekerjaan
        masalah_lengkap = laporan.masalah
        if pekerjaan_list:
            masalah_lengkap += " | Detail: " + " • ".join(pekerjaan_list)
        
        combined_data.append({
            'id': laporan.id,
            'type': 'laporan',  # buat identifier
            'tanggal': laporan.tanggal,
            'jam': None,  # laporan ga punya jam spesifik
            'shift': laporan.shift,
            'nomor_wo': None,  # laporan ga punya nomor WO
            'waktu_pengerjaan': laporan.lama_pekerjaan if hasattr(laporan, 'lama_pekerjaan') else None,
            'line': None,
            'mesin': None,
            'nomer': None,
            'masalah': masalah_lengkap,  # masalah + detail pekerjaan
            'penyebab': None,
            'pekerjaan': " • ".join(pekerjaan_list) if pekerjaan_list else None,
            'status_pekerjaan': None,
            'tindakan_perbaikan': None,
            'tindakan_pencegahan': None,
            'tindakan': None,  # laporan ga punya field tindakan
            'status': None,  # laporan ga punya status
            'image': laporan.image,
            'pic': laporan.pic.all(),
            'pic_lembur': laporan.piclembur.all() if hasattr(laporan, 'piclembur') else [],
            'user': laporan.user,
            'catatan': laporan.catatan if hasattr(laporan, 'catatan') else None,
            'lama_pekerjaan': laporan.lama_pekerjaan if hasattr(laporan, 'lama_pekerjaan') else None,
            'pic_masalah': laporan.pic_pekerjaan if hasattr(laporan, 'pic_pekerjaan') else None,  # Field 'pic_pekerjaan' di LaporanData
            'jenis_pekerjaan': " • ".join(jenis_pekerjaan_list) if jenis_pekerjaan_list else None,
            'location': None,  # laporan ga punya location
            'machine': None,  # laporan ga punya machine
            'category': None,  # laporan ga punya category
        })
    
    # Sort combined data berdasarkan tanggal dan jam (simple sorting)
    def safe_sort_key2(item):
        """Helper function untuk sorting yang aman"""
        tanggal = item['tanggal']
        jam = item['jam']
        
        # Convert semua ke string buat sorting, atau set default value
        if jam is None:
            jam_sort = "00:00:00"  # Default time string
        elif isinstance(jam, datetime):
            jam_sort = jam.strftime("%H:%M:%S")
        elif hasattr(jam, 'strftime'):
            jam_sort = jam.strftime("%H:%M:%S")  
        else:
            jam_sort = str(jam) if jam else "00:00:00"
        
        return (tanggal, jam_sort, -item['id'])
    
    try:
        combined_data.sort(key=safe_sort_key2, reverse=True)
    except Exception as sort_error:
        # Fallback: sort cuma berdasarkan tanggal dan ID aja
        print(f"Sorting error: {sort_error}, using fallback sort")
        combined_data.sort(key=lambda x: (x['tanggal'], -x['id']), reverse=True)
    
    context = {
        'combined_data': combined_data,
        'utility_data': utility_data,  # tetep kirim yang original buat compatibility
        'laporan_data': laporan_data,
        'selected_date': tanggal_parsed,
        'total_utility': utility_data.count(),
        'total_laporan': laporan_data.count(),
        'total_combined': len(combined_data),
        'data_info': {
            'utility_model': 'UtilityData',  # info model yang dipake
            'laporan_model': 'LaporanData',
            'has_utility': utility_data.exists(),
            'has_laporan': laporan_data.exists(),
        }
    }
    
    return render(request, 'dailyactivity_app/data_utility.html', context)


def shift_index(request):
    shifts = Shift.objects.all()  # Mengambil semua data shift

    if request.method == "POST":
        shift_name = request.POST.get('shift_name')
        if shift_name:
            Shift.objects.create(name=shift_name)
            messages.success(request, "Shift added successfully!")
            return redirect('dailyactivity_app:shift_index')

    return render(request, 'dailyactivity_app/shift_index.html', {'shifts': shifts})

def edit_shift(request, pk):
    shift = get_object_or_404(Shift, pk=pk)

    if request.method == "POST":
        shift_name = request.POST.get('shift_name')
        if shift_name:
            shift.name = shift_name
            shift.save()
            messages.success(request, "Shift updated successfully!")
            return redirect('dailyactivity_app:shift_index')

    return render(request, 'dailyactivity_app/edit_shift.html', {'shift': shift})

def delete_shift(request, pk):
    shift = get_object_or_404(Shift, pk=pk)
    shift.delete()
    messages.success(request, "Shift deleted successfully!")
    return redirect('dailyactivity_app:shift_index')

def location_index(request):
    locations = Location.objects.all()  # Mengambil semua data shift

    if request.method == "POST":
        location_name = request.POST.get('location_name')
        if location_name:
            Location.objects.create(name=location_name)
            messages.success(request, "Location added successfully!")
            return redirect('dailyactivity_app:location_index')

    return render(request, 'dailyactivity_app/location_index.html', {'locations': locations})

def edit_location(request, pk):
    location = get_object_or_404(Location, pk=pk)

    if request.method == "POST":
        location_name = request.POST.get('location_name')
        if location_name:
            location.name = location_name
            location.save()
            messages.success(request, "Locationupdated successfully!")
            return redirect('dailyactivity_app:location_index')

    return render(request, 'dailyactivity_app/edit_location.html', {'location': location})

def delete_location(request, pk):
    location = get_object_or_404(Location, pk=pk)
    location.delete()
    messages.success(request, "Location deleted successfully!")
    return redirect('dailyactivity_app:location_index')

def category_index(request):
    categorys = Category.objects.all()  # Mengambil semua data shift

    if request.method == "POST":
        category_name = request.POST.get('category_name')
        if category_name:
            Category.objects.create(name=category_name)
            messages.success(request, "Category added successfully!")
            return redirect('dailyactivity_app:category_index')

    return render(request, 'dailyactivity_app/category_index.html', {'categorys': categorys})

def edit_category(request, pk):
    category = get_object_or_404(Category, pk=pk)

    if request.method == "POST":
        category_name = request.POST.get('category_name')
        if category_name:
            category.name = category_name
            category.save()
            messages.success(request, "Category updated successfully!")
            return redirect('dailyactivity_app:category_index')

    return render(request, 'dailyactivity_app/edit_category.html', {'category': category})

def delete_category(request, pk):
    location = get_object_or_404(Location, pk=pk)
    location.delete()
    messages.success(request, "Category deleted successfully!")
    return redirect('dailyactivity_app:category_index')

def status_index(request):
    status = Status.objects.all()  # Mengambil semua data shift

    if request.method == "POST":
        status_name = request.POST.get('status_name')
        if status_name:
            Status.objects.create(name=status_name)
            messages.success(request, "Status added successfully!")
            return redirect('dailyactivity_app:status_index')

    return render(request, 'dailyactivity_app/status_index.html', {'status': status})

def edit_status(request, pk):
    status = get_object_or_404(Status, pk=pk)

    if request.method == "POST":
        status_name = request.POST.get('status_name')
        if status_name:
            status.name = status_name
            status.save()
            messages.success(request, "Status updated successfully!")
            return redirect('dailyactivity_app:status_index')

    return render(request, 'dailyactivity_app/edit_status.html', {'status': status})

def delete_status(request, pk):
    status = get_object_or_404(Status, pk=pk)
    status.delete()
    messages.success(request, "Status deleted successfully!")
    return redirect('dailyactivity_app:status_index')


def machinemechanical_index(request):
    machinemechanicals = Machinemechanical.objects.all()  # Mengambil semua data dari model MachineMechanical
    locations = Location.objects.all()  # Mengambil semua data Location untuk dipilih

    if request.method == "POST":
        machinemechanical_name = request.POST.get('machinemechanical_name')
        location_id = request.POST.get('location')  # Mendapatkan location dari form
        nomor = request.POST.get('nomor')  # Mendapatkan nomor dari form

        if machinemechanical_name and location_id:
            location = Location.objects.get(id=location_id)  # Mengambil Location berdasarkan ID
            Machinemechanical.objects.create(
                name=machinemechanical_name,
                location=location,
                nomor=nomor  # Menyimpan nomor
            )
            messages.success(request, "Machine Mechanical added successfully!")
            return redirect('dailyactivity_app:machinemechanical_index')

    return render(request, 'dailyactivity_app/machinemechanical_index.html', {
        'machinemechanicals': machinemechanicals,
        'locations': locations
    })

def edit_machinemechanical(request, pk):
    machinemechanical = get_object_or_404(Machinemechanical, pk=pk)  # Model yang tepat
    locations = Location.objects.all()  # Mengambil semua data Location untuk dipilih

    if request.method == "POST":
        machinemechanical_name = request.POST.get('machinemechanical_name')
        location_id = request.POST.get('location')  # Mendapatkan location dari form
        nomor = request.POST.get('nomor')  # Mendapatkan nomor dari form

        if machinemechanical_name and location_id:
            location = Location.objects.get(id=location_id)  # Mengambil Location berdasarkan ID
            machinemechanical.name = machinemechanical_name
            machinemechanical.location = location
            machinemechanical.nomor = nomor  # Mengupdate nomor
            machinemechanical.save()
            messages.success(request, "Machine mechanical updated successfully!")
            return redirect('dailyactivity_app:machinemechanical_index')

    return render(request, 'dailyactivity_app/edit_machinemechanical.html', {
        'machinemechanical': machinemechanical,
        'locations': locations
    })


def delete_machinemechanical(request, pk):
    machinemechanical = get_object_or_404(Machinemechanical, pk=pk)  # Model yang benar digunakan
    machinemechanical.delete()
    messages.success(request, "Machine mechanical deleted successfully!")
    return redirect('dailyactivity_app:machinemechanical_index')

def machineelectrical_index(request):
    machineelectricals = Machineelectrical.objects.all()  # Mengambil semua data dari model Machineelectrical
    locations = Location.objects.all()  # Mengambil semua data Location untuk dipilih

    # Inisialisasi variabel yang akan digunakan
    machineelectrical_name = None
    location_id = None
    nomor = None

    if request.method == "POST":
        machineelectrical_name = request.POST.get('machineelectrical_name')  # Nama sesuai form
        location_id = request.POST.get('location')  # Mendapatkan location dari form
        nomor = request.POST.get('nomor')  # Mendapatkan nomor dari form

        if machineelectrical_name and location_id:
            try:
                location = Location.objects.get(id=location_id)  # Mengambil Location berdasarkan ID
                Machineelectrical.objects.create(
                    name=machineelectrical_name,
                    location=location,
                    nomor=nomor  # Menyimpan nomor
                )
                messages.success(request, "Machine Electrical added successfully!")
                return redirect('dailyactivity_app:machineelectrical_index')
            except Location.DoesNotExist:
                messages.error(request, "Location tidak ditemukan.")
    
    return render(request, 'dailyactivity_app/machineelectrical_index.html', {
        'machineelectricals': machineelectricals,
        'locations': locations
    })


def edit_machineelectrical(request, pk):
    machineelectrical = get_object_or_404(Machineelectrical, pk=pk)  # Model yang tepat
    locations = Location.objects.all()  # Mengambil semua data Location untuk dipilih

    if request.method == "POST":
        machineelectrical_name = request.POST.get('machineelectrical_name')
        location_id = request.POST.get('location')  # Mendapatkan location dari form
        nomor = request.POST.get('nomor')  # Mendapatkan nomor dari form

        if machineelectrical_name and location_id:
            location = Location.objects.get(id=location_id)  # Mengambil Location berdasarkan ID
            machineelectrical.name = machineelectrical_name
            machineelectrical.location = location
            machineelectrical.nomor = nomor  # Mengupdate nomor
            machineelectrical.save()
            messages.success(request, "Machine mechanical updated successfully!")
            return redirect('dailyactivity_app:machineelectrical_index')

    return render(request, 'dailyactivity_app/edit_machineelectrical.html', {
        'machineelectrical': machineelectrical,
        'locations': locations
    })


def delete_machineelectrical(request, pk):
    machineelectrical = get_object_or_404(Machineelectrical, pk=pk)  # Model yang benar digunakan
    machineelectrical.delete()
    messages.success(request, "Machine machineelectrical deleted successfully!")
    return redirect('dailyactivity_app:machineelectrical_index')


def machineutility_index(request):
    machineutilitys = Machineutility.objects.all()  # Mengambil semua data dari model Machineelectrical
    locations = Location.objects.all()  # Mengambil semua data Location untuk dipilih

    # Inisialisasi variabel yang akan digunakan
    machineutility_name = None
    location_id = None
    nomor = None

    if request.method == "POST":
        machineutility_name = request.POST.get('machineutility_name')  # Nama sesuai form
        location_id = request.POST.get('location')  # Mendapatkan location dari form
        nomor = request.POST.get('nomor')  # Mendapatkan nomor dari form

        if machineutility_name and location_id:
            try:
                location = Location.objects.get(id=location_id)  # Mengambil Location berdasarkan ID
                Machineutility.objects.create(
                    name=machineutility_name,
                    location=location,
                    nomor=nomor  # Menyimpan nomor
                )
                messages.success(request, "Machine Electrical added successfully!")
                return redirect('dailyactivity_app:machineutility_index')
            except Location.DoesNotExist:
                messages.error(request, "Location tidak ditemukan.")
    
    return render(request, 'dailyactivity_app/machineutility_index.html', {
        'machineutilitys': machineutilitys,
        'locations': locations
    })


def edit_machineutility(request, pk):
    machineutility = get_object_or_404(Machineutility, pk=pk)  # Model yang tepat
    locations = Location.objects.all()  # Mengambil semua data Location untuk dipilih

    if request.method == "POST":
        machineutility_name = request.POST.get('machineutility_name')
        location_id = request.POST.get('location')  # Mendapatkan location dari form
        nomor = request.POST.get('nomor')  # Mendapatkan nomor dari form

        if machineutility_name and location_id:
            location = Location.objects.get(id=location_id)  # Mengambil Location berdasarkan ID
            machineutility.name = machineutility_name
            machineutility.location = location
            machineutility.nomor = nomor  # Mengupdate nomor
            machineutility.save()
            messages.success(request, "Machine mechanical updated successfully!")
            return redirect('dailyactivity_app:machineutility_index')

    return render(request, 'dailyactivity_app/edit_machineutility.html', {
        'machineutility': machineutility,
        'locations': locations
    })


def delete_machineutility(request, pk):
    machineutility = get_object_or_404(Machineutility, pk=pk)  # Model yang benar digunakan
    machineutility.delete()
    messages.success(request, "Machine Utility deleted successfully!")
    return redirect('dailyactivity_app:machineutility_index')

# Fungsi untuk menampilkan semua data PIC Mechanical
def picmechanical_index(request):
    picmechanicals = PICMechanical2.objects.all()  # Mengambil semua data dari model PicMechanical

    if request.method == "POST":
        picmechanical_name = request.POST.get('name')  # Nama PIC Mechanical
        picmechanical_nokar = request.POST.get('nokar')  # No. Karyawan PIC Mechanical
        
        if picmechanical_name and picmechanical_nokar:
            PICMechanical2.objects.create(name=picmechanical_name, nokar=picmechanical_nokar)
            messages.success(request, "PIC Mechanical added successfully!")
            return redirect('dailyactivity_app:picmechanical_index')

    return render(request, 'dailyactivity_app/picmechanical_index.html', {'picmechanicals': picmechanicals})

# Fungsi untuk mengedit data PIC Mechanical
def edit_picmechanical(request, pk):
    # Mendapatkan objek PICMechanical yang akan diedit
    picmechanical = get_object_or_404(PICMechanical2, pk=pk)
    
    # Memeriksa apakah request adalah POST (artinya data form dikirim)
    if request.method == 'POST':
        form = PICMechanical2Form(request.POST, instance=picmechanical)
        if form.is_valid():
            form.save()  # Menyimpan perubahan pada objek yang ada
            return redirect('dailyactivity_app:picmechanical_index')  # Redirect ke halaman yang sesuai
    else:
        form = PICMechanical2Form(instance=picmechanical)  # Mengisi form dengan data yang ada
    
    return render(request, 'dailyactivity_app/edit_picmechanical.html', {'form': form})


# Fungsi untuk menghapus data PIC Mechanical
def delete_picmechanical(request, pk):
    picmechanical = get_object_or_404(PICMechanical2, pk=pk)  # Mengambil data PIC Mechanical berdasarkan pk
    picmechanical.delete()
    messages.success(request, "PIC Mechanical deleted successfully!")
    return redirect('dailyactivity_app:picmechanical_index')

# Fungsi untuk menampilkan semua data PIC Electrical
def picelectrical_index(request):
    picelectricals = PICElectrical.objects.all()  # Mengambil semua data dari model PicMechanical

    if request.method == "POST":
        picelectrical_name = request.POST.get('name')  # Nama PIC Mechanical
        picelectrical_nokar = request.POST.get('nokar')  # No. Karyawan PIC Mechanical
        
        if picelectrical_name and picelectrical_nokar:
            PICElectrical.objects.create(name=picelectrical_name, nokar=picelectrical_nokar)
            messages.success(request, "PIC Electrical added successfully!")
            return redirect('dailyactivity_app:picelectrical_index')

    return render(request, 'dailyactivity_app/picelectrical_index.html', {'picelectricals': picelectricals})

# Fungsi untuk mengedit data PIC Mechanical
def edit_picelectrical(request, pk):
    # Mendapatkan objek PICMechanical yang akan diedit
    picelectrical = get_object_or_404(PICElectrical, pk=pk)
    
    # Memeriksa apakah request adalah POST (artinya data form dikirim)
    if request.method == 'POST':
        form = PICElectricalForm(request.POST, instance=picelectrical)
        if form.is_valid():
            form.save()  # Menyimpan perubahan pada objek yang ada
            return redirect('dailyactivity_app:picmechanical_index')  # Redirect ke halaman yang sesuai
    else:
        form = PICElectricalForm(instance=picelectrical)  # Mengisi form dengan data yang ada
    
    return render(request, 'dailyactivity_app/edit_picelectrical.html', {'form': form})


# Fungsi untuk menghapus data PIC Mechanical
def delete_picelectrical(request, pk):
    picelectrical = get_object_or_404(PICElectrical, pk=pk)  # Mengambil data PIC Mechanical berdasarkan pk
    picelectrical.delete()
    messages.success(request, "PIC Electrical deleted successfully!")
    return redirect('dailyactivity_app:picelectrical_index')

# Fungsi untuk menampilkan semua data PIC Electrical
def picutility_index(request):
    picutilitys = PICUtility2.objects.all()  # Mengambil semua data dari model PicMechanical

    if request.method == "POST":
        picutility_name = request.POST.get('name')  # Nama PIC Mechanical
        picutility_nokar = request.POST.get('nokar')  # No. Karyawan PIC Mechanical
        
        if picutility_name and picutility_nokar:
            PICUtility2.objects.create(name=picutility_name, nokar=picutility_nokar)
            messages.success(request, "PIC Electrical added successfully!")
            return redirect('dailyactivity_app:picutility_index')

    return render(request, 'dailyactivity_app/picutility_index.html', {'picutilitys': picutilitys})

# Fungsi untuk mengedit data PIC Mechanical
def edit_picutility(request, pk):
    # Mendapatkan objek PICMechanical yang akan diedit
    picutility = get_object_or_404(PICUtility2, pk=pk)
    
    # Memeriksa apakah request adalah POST (artinya data form dikirim)
    if request.method == 'POST':
        form = PICUtility2Form(request.POST, instance=picutility)
        if form.is_valid():
            form.save()  # Menyimpan perubahan pada objek yang ada
            return redirect('dailyactivity_app:picutility_index')  # Redirect ke halaman yang sesuai
    else:
        form = PICUtility2Form(instance=picutility)  # Mengisi form dengan data yang ada
    
    return render(request, 'dailyactivity_app/edit_picutility.html', {'form': form})


# Fungsi untuk menghapus data PIC Mechanical
def delete_picutility(request, pk):
    picutility = get_object_or_404(PICUtility2, pk=pk)  # Mengambil data PIC Mechanical berdasarkan pk
    picutility.delete()
    messages.success(request, "PIC Utility deleted successfully!")
    return redirect('dailyactivity_app:picutility_index')

# Fungsi untuk menampilkan semua data PIC Electrical
def picit_index(request):
    picits = PICIt.objects.all()  # Mengambil semua data dari model PicMechanical

    if request.method == "POST":
        picit_name = request.POST.get('name')  # Nama PIC Mechanical
        picit_nokar = request.POST.get('nokar')  # No. Karyawan PIC Mechanical
        
        if picit_name and picit_nokar:
            PICIt.objects.create(name=picit_name, nokar=picit_nokar)
            messages.success(request, "PIC Electrical added successfully!")
            return redirect('dailyactivity_app:picit_index')

    return render(request, 'dailyactivity_app/picit_index.html', {'picits': picits})

# Fungsi untuk mengedit data PIC Mechanical
def edit_picit(request, pk):
    # Mendapatkan objek PICMechanical yang akan diedit
    picutility = get_object_or_404(PICIt, pk=pk)
    
    # Memeriksa apakah request adalah POST (artinya data form dikirim)
    if request.method == 'POST':
        form = PICItForm(request.POST, instance=picit)
        if form.is_valid():
            form.save()  # Menyimpan perubahan pada objek yang ada
            return redirect('dailyactivity_app:picit_index')  # Redirect ke halaman yang sesuai
    else:
        form = PICItForm(instance=picit)  # Mengisi form dengan data yang ada
    
    return render(request, 'dailyactivity_app/edit_picit.html', {'form': form})


# Fungsi untuk menghapus data PIC Mechanical
def delete_picit(request, pk):
    picutility = get_object_or_404(PICIt, pk=pk)  # Mengambil data PIC Mechanical berdasarkan pk
    picutility.delete()
    messages.success(request, "PIC Utility deleted successfully!")
    return redirect('dailyactivity_app:picutility_index')

@login_required
def it_index(request):
    shifts = Shift.objects.all()
    status = Status.objects.all()
    pic_it = PICIt.objects.all()

    # Ambil daftar nomor WO
    nomor_wo_list = []
    with connections['DB_Maintenance'].cursor() as cursor:
        cursor.execute("""
            SELECT number_wo, status_pekerjaan
            FROM dbo.view_main
            WHERE id_section = 8
        AND YEAR(tgl_his) BETWEEN 2024 AND 2024
        ORDER BY history_id DESC
        """)
        nomor_wo_list = [(row[0], row[1]) for row in cursor.fetchall()]  # Simpan hanya kolom number_wo dan status_pekerjaan

    # Default untuk deskripsi_perbaikan, tgl_his, dan penyebab
    deskripsi_perbaikan = None
    tgl_his = None
    penyebab = None
    line = None
    mesin = None
    nomer = None
    pekerjaan = None
    status_pekerjaan = None
    tindakan_perbaikan = None
    tindakan_pencegahan = None

    # Tangani Form Submission
    if request.method == 'POST':
        form = ItDataForm(request.POST, request.FILES)
        if form.is_valid():
            # Proses data form
            machine_number = form.cleaned_data.get('machine_number')
            machine_instance = form.cleaned_data.get('machine')

            # Cek apakah mesin sudah memiliki nomor
            if machine_number and machine_instance and not machine_instance.nomor:
                machine_instance.nomor = machine_number
                machine_instance.save()

            # Simpan data MechanicalData
            it_data = form.save(commit=False)
            it_data.user = request.user
            it_data.machine = machine_instance
            it_data.save()

            # Hubungkan PIC dengan MechanicalData
            pic_ids = form.cleaned_data.get('pic')
            it_data.pic.set(pic_ids)

            # Simpan nomor WO dan waktu pengerjaan
            it_data.nomor_wo = form.cleaned_data.get('nomor_wo')
            it_data.waktu_pengerjaan = form.cleaned_data.get('waktu_pengerjaan')
            it_data.save()
            
            return redirect('success_page')
    else:
        form = ItDataForm()

    # Tangani Permintaan AJAX
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.method == 'GET':
        nomor_wo_selected = request.GET.get('nomor_wo')
        with connections['DB_Maintenance'].cursor() as cursor:
            cursor.execute("""
                SELECT deskripsi_perbaikan, tgl_his, penyebab, line, mesin, nomer, pekerjaan, status_pekerjaan, tindakan_perbaikan, tindakan_pencegahan
                FROM dbo.view_main
                WHERE number_wo = %s
            """, [nomor_wo_selected])
            row = cursor.fetchone()

            if row:
                deskripsi_perbaikan, tgl_his, penyebab, line, mesin, nomer, pekerjaan, status_pekerjaan, tindakan_perbaikan, tindakan_pencegahan = row
            else:
                # Jika nomor WO tidak ditemukan, ambil data line dan mesin dari tabel mesin
                with connections['DB_Maintenance'].cursor() as machine_cursor:
                    machine_cursor.execute("""
                        SELECT line, mesin
                        FROM dbo.mesin
                        ORDER BY id ASC
                        LIMIT 1
                    """)
                    machine_row = machine_cursor.fetchone()
                    if machine_row:
                        line, mesin = machine_row

        return JsonResponse({
            'deskripsi_perbaikan': deskripsi_perbaikan,
            'tgl_his': tgl_his,
            'penyebab': penyebab,
            'line': line,
            'mesin': mesin,
            'nomer': nomer,
            'pekerjaan': pekerjaan,
            'status_pekerjaan': status_pekerjaan,
            'tindakan_perbaikan': tindakan_perbaikan,
            'tindakan_pencegahan': tindakan_pencegahan
        })

    # Context untuk template
    context = {
        'shifts': shifts,
        'status': status,
        'pic_it': pic_it,
        'nomor_wo_list': nomor_wo_list,
        'deskripsi_perbaikan': deskripsi_perbaikan,
        'tgl_his': tgl_his,
        'penyebab': penyebab,
        'line': line,
        'mesin': mesin,
        'nomer': nomer,
        'pekerjaan': pekerjaan,
        'status_pekerjaan': status_pekerjaan,
        'tindakan_perbaikan': tindakan_perbaikan,
        'tindakan_pencegahan': tindakan_pencegahan,
        'form': form,
    }
    return render(request, 'dailyactivity_app/it_index.html', context)

@login_required  # Pastikan pengguna harus login untuk mengakses fungsi ini
def it_submit(request):
    if request.method == 'POST':
        # Ambil data dari formulir
        tanggal = request.POST.get('tanggal')
        jam = request.POST.get('jam')
        tgl_his = request.POST.get('tgl_his')
        shift_id = request.POST.get('shift')
        # machine_id = request.POST.get('machine')
        status_id = request.POST.get('status')
        masalah = request.POST.get('masalah')
        penyebab = request.POST.get('penyebab')  # Default None jika tidak diisi
        line = request.POST.get('line')   # Default None jika tidak diisi
        mesin = request.POST.get('mesin')  # Default None jika tidak diisi
        nomer = request.POST.get('nomer') # Default None jika tidak diisi
        pekerjaan = request.POST.get('pekerjaan')   # Default None jika tidak diisi
        status_pekerjaan = request.POST.get('status_pekerjaan')   # Default None jika tidak diisi
        tindakan_perbaikan = request.POST.get('tindakan_perbaikan')   # Default None jika tidak diisi
        tindakan_pencegahan = request.POST.get('tindakan_pencegahan')   # Default None jika tidak diisi
        # tindakan = request.POST.get('tindakan')
        image = request.FILES.get('image')
        pic_ids = request.POST.getlist('pic')  # Ambil daftar PIC yang dipilih
        nomor_wo = request.POST.get('nomor_wo')  # Ambil nomor WO
        waktu_pengerjaan = request.POST.get('waktu_pengerjaan')  # Ambil waktu pengerjaan

        # Ambil instance Shift
        try:
            shift_instance = Shift.objects.get(id=shift_id)
        except Shift.DoesNotExist:
            messages.error(request, 'Shift tidak ditemukan.')
            return redirect('dailyactivity_app:it_index')
        try:
            status_instance = Status.objects.get(id=status_id)
        except Status.DoesNotExist:
            messages.error(request, 'Status tidak ditemukan.')
            return redirect('dailyactivity_app:it_index')

        # Ambil user_id dari pengguna yang sedang login
        user_id = request.user.id

        jam_value = tgl_his if tgl_his else jam

        # Simpan data ke database
        it_data = ItData.objects.create(
            tanggal=tanggal,
            jam=jam_value,
            shift=shift_instance,
            # machine=machine_instance,
            status=status_instance,
            user_id=user_id,
            masalah=masalah,
            penyebab=penyebab,
            line=line,
            mesin=mesin,
            nomer=nomer,
            pekerjaan=pekerjaan,
            status_pekerjaan=status_pekerjaan,
            # tindakan=tindakan,
            tindakan_perbaikan=tindakan_perbaikan,
            tindakan_pencegahan=tindakan_pencegahan,
            image=image,
            nomor_wo=nomor_wo,
            waktu_pengerjaan=waktu_pengerjaan
        )
        # Menyimpan PIC yang dipilih
        for pic_id in pic_ids:
            try:
                pic_instance = PICIt.objects.get(id=pic_id)
                it_data.pic.add(pic_instance)
            except PICIt.DoesNotExist:
                continue
        # Simpan pesan sukses
        messages.success(request, 'Data berhasil disimpan!')

        # Redirect ke mechanical_index
        return redirect('dailyactivity_app:it_index')

    return redirect('dailyactivity_app:it_index')

@login_required
def edit_it_data(request, id):
    it_data = get_object_or_404(ItData, id=id)
    shifts = Shift.objects.all()
    status = Status.objects.all()
    pic_it = PICIt.objects.all()

    if request.method == 'POST':
        # Memproses form data yang telah diisi ulang
        form = ItDataForm(request.POST, request.FILES, instance=it_data)
        if form.is_valid():
            updated_data = form.save(commit=False)
            updated_data.user = request.user
            updated_data.nomor_wo = form.cleaned_data.get('nomor_wo')
            updated_data.waktu_pengerjaan = form.cleaned_data.get('waktu_pengerjaan')
            updated_data.line = form.cleaned_data.get('line')
            updated_data.mesin = form.cleaned_data.get('mesin')
            updated_data.nomer = form.cleaned_data.get('nomer')
            updated_data.pekerjaan = form.cleaned_data.get('pekerjaan')

            updated_data.save()

            # Update PIC terkait
            pic_ids = form.cleaned_data.get('pic')
            updated_data.pic.set(pic_ids)

            messages.success(request, 'Data berhasil diperbarui!')
            return redirect('dailyactivity_app:data_it.html', tanggal=updated_data.tanggal.strftime('%Y-%m-%d'))
        else:
            messages.error(request, 'Terjadi kesalahan saat memperbarui data. Periksa kembali isian Anda.')

    else:
        # Memuat form dengan data yang ada untuk ditampilkan di template
        form = ItDataForm(instance=it_data)

    context = {
        'form': form,
        'shifts': shifts,
        # 'locations': locations,
        # 'machines': machines,
        # 'categories': categories,
        'status': status,
        'pic_it': pic_it,
        'data': it_data,
        'tanggal': it_data.tanggal,
        'jam': it_data.jam,
        'nomor_wo': it_data.nomor_wo,
        'waktu_pengerjaan': it_data.waktu_pengerjaan,
        'line': it_data.line,
        'mesin': it_data.mesin,
        'nomer': it_data.nomer,
        'pekerjaan': it_data.pekerjaan,
        'pic': it_data.pic.all(),
    }
    return render(request, 'dailyactivity_app/edit_it_data.html', context)

@login_required
def delete_it_data(request, id):
    # Mengambil data berdasarkan ID
    it_data = get_object_or_404(ItData, id=id)

    if request.method == 'POST':
        # Hapus data dari database
        it_data.delete()
        messages.success(request, 'Data berhasil dihapus!')
        return redirect('dailyactivity_app:data_it.html', tanggal=it_data.tanggal.strftime('%Y-%m-%d'))
    
    # Jika bukan POST, redirect ke halaman data electrical
    return redirect('dailyactivity_app:data_it.html', tanggal=it_data.tanggal.strftime('%Y-%m-%d'))

@login_required
def tanggal_it(request):
    # Mengambil tanggal-tanggal unik dari data mechanical
    dates = ItData.objects.annotate(date=Trunc('tanggal', 'day', output_field=DateField())).values('date').distinct().order_by('-date')
    context = {
        'dates': dates,
    }
    return render(request, 'dailyactivity_app/tanggal_it.html', context)

@login_required
def data_it(request, tanggal):
    # Parsing tanggal dari URL
    tanggal_parsed = parse_date(tanggal)
    # Menyaring data berdasarkan tanggal yang dipilih
    it_data = ItData.objects.filter(tanggal=tanggal_parsed)
    context = {
        'it_data': it_data,
        'selected_date': tanggal_parsed,
    }
    return render(request, 'dailyactivity_app/data_it.html', context)


# @login_required
# def laporan_index(request):
#     shifts = Shift.objects.all()
#     pic_laporan = PICLaporan.objects.all()  # Fetch all PICLaporan instances
#     pic_lembur = PICLembur.objects.all()    # Fetch all PICLembur instances
#     laporan_data = LaporanData.objects.all()

#     # Handle form submission
#     if request.method == 'POST':
#         form = LaporanDataForm(request.POST, request.FILES)
#         if form.is_valid():
#             # Save the LaporanData instance without committing to DB
#             laporan_data = form.save(commit=False)
#             laporan_data.user = request.user
#             laporan_data.save()  # Save the main LaporanData record

#             # Save selected PICs (PICLaporan)
#             pic_ids = form.cleaned_data.get('pic')
#             laporan_data.pic.set(pic_ids)

#             # Save selected PICLembur if available
#             piclembur_ids = request.POST.getlist('piclembur')  # Fetch piclembur data from POST
#             laporan_data.piclembur.set(piclembur_ids)  # Link piclembur instances

#             laporan_data.save()  # Save changes to LaporanData
#             return redirect('success_page')
#     else:
#         form = LaporanDataForm()

#     context = {
#         'shifts': shifts,
#         'pic_laporan': pic_laporan,  # Pass PICLaporan instances to the template
#         'pic_lembur': pic_lembur,    # Pass PICLembur instances to the template
#         'form': form,
#     }
#     return render(request, 'dailyactivity_app/laporan_index.html', context)

@login_required
def laporan_index(request):
    """View utama untuk input laporan utility enhanced dengan WO & Maintenance per row"""
    
    # Data yang udah ada (JANGAN DIHAPUS)
    shifts = Shift.objects.all()
    pic_laporan = PICLaporan.objects.all()
    pic_lembur = PICLembur.objects.all()
    laporan_data = LaporanData.objects.all()
    
    # Handle form submission dengan multiple rows - FIXED COMPLETE
    if request.method == 'POST':
        print("=== DEBUG: Form Submission Started ===")
        
        # Ambil data common (tanggal, shift, pic, dll)
        tanggal = request.POST.get('tanggal')
        shift_id = request.POST.get('shift')
        catatan = request.POST.get('catatan', '')
        pic_ids = request.POST.getlist('pic')
        piclembur_ids = request.POST.getlist('piclembur')
        image = request.FILES.get('image')
        
        print(f"Common data - Tanggal: {tanggal}, Shift: {shift_id}, Catatan: {catatan}")
        print(f"PIC IDs: {pic_ids}, PIC Lembur IDs: {piclembur_ids}")
        
        # Ambil array data pekerjaan (JENIS PEKERJAAN UTILITY DIHILANGKAN DARI FORM)
        deskripsi_list = request.POST.getlist('deskripsi_pekerjaan[]')
        lama_pekerjaan_list = request.POST.getlist('lama_pekerjaan[]')
        pic_pekerjaan_list = request.POST.getlist('pic_pekerjaan[]')
        
        # Ambil array data WO & Maintenance baru (SEMUA FIELD BARU)
        nomor_wo_list = request.POST.getlist('nomor_wo[]')
        status_utility_list = request.POST.getlist('status_utility[]')
        lokasi_list = request.POST.getlist('lokasi[]')
        mesin_list = request.POST.getlist('mesin[]')
        nomor_mesin_list = request.POST.getlist('nomor_mesin[]')
        jenis_pekerjaan_maintenance_list = request.POST.getlist('jenis_pekerjaan_maintenance[]')
        penyebab_list = request.POST.getlist('penyebab[]')
        tindakan_perbaikan_list = request.POST.getlist('tindakan_perbaikan[]')
        
        # Debug print semua arrays
        print(f"Deskripsi: {deskripsi_list}")
        print(f"Lama Pekerjaan: {lama_pekerjaan_list}")
        print(f"PIC Pekerjaan: {pic_pekerjaan_list}")
        print(f"Nomor WO: {nomor_wo_list}")
        print(f"Status Utility: {status_utility_list}")
        print(f"Lokasi: {lokasi_list}")
        print(f"Mesin: {mesin_list}")
        print(f"Nomor Mesin: {nomor_mesin_list}")
        print(f"Jenis Pekerjaan Maintenance: {jenis_pekerjaan_maintenance_list}")
        print(f"Penyebab: {penyebab_list}")
        print(f"Tindakan Perbaikan: {tindakan_perbaikan_list}")
        
        # Validasi basic
        if not tanggal or not shift_id or not deskripsi_list:
            messages.error(request, 'Data tanggal, shift, dan minimal 1 deskripsi pekerjaan harus diisi!')
            return redirect('dailyactivity_app:laporan_index')
        
        try:
            with transaction.atomic():
                # Dapatkan shift object
                shift = Shift.objects.get(id=shift_id)
                created_laporans = []
                
                print(f"=== Processing {len(deskripsi_list)} rows ===")
                
                # Loop untuk setiap baris pekerjaan - HANDLE SEMUA FIELD BARU
                for i in range(len(deskripsi_list)):
                    if not deskripsi_list[i].strip():  # Skip empty descriptions
                        continue
                        
                    print(f"--- Processing row {i+1} ---")
                    
                    # Create LaporanData instance dengan SEMUA FIELD
                    laporan = LaporanData(
                        tanggal=tanggal,
                        shift=shift,
                        user=request.user,
                        
                        # Data utility (field lama)
                        masalah=deskripsi_list[i].strip(),  # deskripsi -> masalah
                        jenis_pekerjaan='',  # Dikosongkan karena input utility dihilangkan
                        lama_pekerjaan=lama_pekerjaan_list[i].strip() if i < len(lama_pekerjaan_list) else '',
                        pic_pekerjaan=pic_pekerjaan_list[i].strip() if i < len(pic_pekerjaan_list) else '',
                        catatan=catatan,
                        
                        # Data WO & Maintenance (field baru) - COMPLETE MAPPING
                        nomor_wo=nomor_wo_list[i].strip() if i < len(nomor_wo_list) and nomor_wo_list[i] else '',
                        status_utility=status_utility_list[i] if i < len(status_utility_list) and status_utility_list[i] else 'proses',
                        lokasi=lokasi_list[i].strip() if i < len(lokasi_list) and lokasi_list[i] else '',
                        mesin=mesin_list[i].strip() if i < len(mesin_list) and mesin_list[i] else '',
                        nomor_mesin=nomor_mesin_list[i].strip() if i < len(nomor_mesin_list) and nomor_mesin_list[i] else '',
                        jenis_pekerjaan_maintenance=jenis_pekerjaan_maintenance_list[i].strip() if i < len(jenis_pekerjaan_maintenance_list) and jenis_pekerjaan_maintenance_list[i] else '',
                        penyebab=penyebab_list[i].strip() if i < len(penyebab_list) and penyebab_list[i] else '',
                        tindakan_perbaikan=tindakan_perbaikan_list[i].strip() if i < len(tindakan_perbaikan_list) and tindakan_perbaikan_list[i] else '',
                        
                        # Image hanya untuk record pertama
                        image=image if i == 0 else None
                    )
                    
                    # Save record
                    laporan.save()
                    created_laporans.append(laporan)
                    
                    print(f"Saved laporan {laporan.id} - WO: {laporan.nomor_wo}, Lokasi: {laporan.lokasi}")
                    
                    # Set PIC untuk setiap laporan (Many-to-Many)
                    if pic_ids:
                        pic_objects = PICLaporan.objects.filter(id__in=pic_ids)
                        laporan.pic.set(pic_objects)
                        print(f"Set PIC: {list(pic_objects.values_list('name', flat=True))}")
                    
                    if piclembur_ids:
                        piclembur_objects = PICLembur.objects.filter(id__in=piclembur_ids)
                        laporan.piclembur.set(piclembur_objects)
                        print(f"Set PIC Lembur: {list(piclembur_objects.values_list('name', flat=True))}")
                
                # Success message
                messages.success(request, f'✅ Berhasil menyimpan {len(created_laporans)} data laporan utility!')
                print(f"=== SUCCESS: Saved {len(created_laporans)} records ===")
                
                # Redirect ke halaman sukses
                return redirect('dailyactivity_app:tanggal_laporan')
            
        except Shift.DoesNotExist:
            messages.error(request, '❌ Shift yang dipilih tidak ditemukan!')
            print("ERROR: Shift not found")
        except Exception as e:
            messages.error(request, f'❌ Error menyimpan data: {str(e)}')
            print(f"ERROR: {e}")
            import traceback
            traceback.print_exc()
        
        # Jika error, tetap tampilkan form
        return redirect('dailyactivity_app:laporan_index')
    
    else:
        # GET request - tampilkan form
        form = LaporanDataForm()
    
    # Data untuk dropdown - pake fallback kalau error
    try:
        nomor_wo_list = get_nomor_wo_list()
    except Exception as e:
        print(f"Error loading nomor_wo_list: {e}")
        nomor_wo_list = []
    
    try:
        lokasi_list = get_lokasi_list()
    except Exception as e:
        print(f"Error loading lokasi_list: {e}")
        lokasi_list = []
    
    try:
        jenis_pekerjaan_list = get_jenis_pekerjaan_list()
    except Exception as e:
        print(f"Error loading jenis_pekerjaan_list: {e}")
        jenis_pekerjaan_list = []
    
    context = {
        'shifts': shifts,
        'pic_laporan': pic_laporan,
        'pic_lembur': pic_lembur,
        'form': form,
        # Data baru dari DB_Maintenance (dengan fallback)
        'nomor_wo_list': nomor_wo_list,
        'lokasi_list': lokasi_list,
        'jenis_pekerjaan_list': jenis_pekerjaan_list,
    }
    return render(request, 'dailyactivity_app/laporan_index.html', context)


def get_nomor_wo_list():
    """Ambil daftar nomor WO dari DB_Maintenance"""
    nomor_wo_list = []
    try:
        with connections['DB_Maintenance'].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT number_wo
                FROM dbo.view_main
                WHERE id_section = 6
                AND YEAR(tgl_his) BETWEEN 2024 AND 2025
                AND number_wo IS NOT NULL
                AND number_wo != ''
                ORDER BY number_wo DESC
            """)
            nomor_wo_list = [{'value': row[0], 'text': row[0]} for row in cursor.fetchall()]
    except Exception as e:
        print(f"Error fetching nomor WO: {e}")
        # Fallback data
        nomor_wo_list = [
            {'value': 'WO001', 'text': 'WO001 - Sample'},
            {'value': 'WO002', 'text': 'WO002 - Sample'}
        ]
    return nomor_wo_list


def get_lokasi_list():
    """Ambil daftar lokasi dari tabel line DB_Maintenance - FIX DUPLICATE"""
    lokasi_list = []
    try:
        with connections['DB_Maintenance'].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT id_line, line
                FROM tabel_line
                WHERE status = 'A'
                AND line IS NOT NULL
                AND line != ''
                ORDER BY line
            """)
            lokasi_list = [{'value': str(row[0]), 'text': row[1]} for row in cursor.fetchall()]
    except Exception as e:
        print(f"Error fetching lokasi: {e}")
        # Fallback data
        lokasi_list = [
            {'value': '1', 'text': 'Line 1'},
            {'value': '2', 'text': 'Line 2'}
        ]
    return lokasi_list


def get_jenis_pekerjaan_list():
    """Ambil daftar jenis pekerjaan dari tabel_pekerjaan DB_Maintenance"""
    jenis_pekerjaan_list = []
    try:
        with connections['DB_Maintenance'].cursor() as cursor:
            cursor.execute("""
                SELECT id_pekerjaan, pekerjaan
                FROM tabel_pekerjaan
                WHERE status = 'A'
                AND pekerjaan IS NOT NULL
                AND pekerjaan != ''
                ORDER BY pekerjaan
            """)
            jenis_pekerjaan_list = [{'value': str(row[0]), 'text': row[1]} for row in cursor.fetchall()]
    except Exception as e:
        print(f"Error fetching jenis pekerjaan: {e}")
        # Fallback data
        jenis_pekerjaan_list = [
            {'value': '1', 'text': 'Maintenance'},
            {'value': '2', 'text': 'Cleaning'},
            {'value': '3', 'text': 'Repair'}
        ]
    return jenis_pekerjaan_list

@login_required
def get_wo_details(request):
    """AJAX endpoint untuk auto-fill data berdasarkan nomor WO"""
    nomor_wo = request.GET.get('nomor_wo')
    
    if not nomor_wo:
        return JsonResponse({'error': 'Nomor WO required'})
    
    try:
        with connections['DB_Maintenance'].cursor() as cursor:
            cursor.execute("""
                SELECT 
                    vm.deskripsi_perbaikan,
                    vm.id_line,
                    vm.line,
                    vm.id_mesin,
                    vm.mesin,
                    vm.nomer,
                    vm.penyebab,
                    vm.tindakan_perbaikan
                FROM dbo.view_main vm
                WHERE vm.number_wo = %s
                ORDER BY vm.history_id DESC
            """, [nomor_wo])
            
            result = cursor.fetchone()
            
            if result:
                return JsonResponse({
                    'success': True,
                    'data': {
                        'deskripsi_perbaikan': result[0] or '',
                        'id_line': str(result[1]) if result[1] else '',
                        'line': result[2] or '',
                        'id_mesin': str(result[3]) if result[3] else '',
                        'mesin': result[4] or '',
                        'nomer': result[5] or '',
                        'penyebab': result[6] or '',
                        'tindakan_perbaikan': result[7] or ''
                    }
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Nomor WO tidak ditemukan'
                })
                
    except Exception as e:
        print(f"Error fetching WO details: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def get_mesin_by_lokasi_utility(request, lokasi_id):
    """AJAX endpoint untuk dapetin daftar mesin berdasarkan lokasi - FIXED untuk handle float/string"""
    
    # Convert lokasi_id to integer (handle float values seperti "28.0")
    try:
        lokasi_id_int = int(float(lokasi_id))
    except (ValueError, TypeError):
        return JsonResponse([], safe=False)
    
    mesin_list = []
    try:
        with connections['DB_Maintenance'].cursor() as cursor:
            cursor.execute("""
                SELECT tm.id_mesin, tm.mesin, tm.nomer
                FROM tabel_mesin tm
                WHERE tm.id_line = %s 
                AND tm.status = 'A'
                AND tm.mesin IS NOT NULL
                AND tm.mesin != ''
                ORDER BY tm.mesin, tm.nomer
            """, [lokasi_id_int])
            
            results = cursor.fetchall()
            
            # Format response sama seperti mechanical
            for row in results:
                mesin_list.append({
                    'id': row[0],
                    'name': row[1],
                    'location_id': str(lokasi_id_int),
                    'nomor': row[2] if row[2] else ''
                })
    
    except Exception as e:
        print(f"Error fetching mesin by lokasi {lokasi_id}: {e}")
        # Return empty list instead of fallback untuk debugging
        return JsonResponse([], safe=False)
    
    # Return format sama persis seperti mechanical
    return JsonResponse(mesin_list, safe=False)


@login_required
def get_machine_number_utility(request, machine_id):
    """AJAX endpoint untuk dapetin nomor mesin berdasarkan machine_id - FIXED untuk handle float/string"""
    
    # Convert machine_id to integer (handle float values)
    try:
        machine_id_int = int(float(machine_id))
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid machine ID'}, status=400)
        
    try:
        with connections['DB_Maintenance'].cursor() as cursor:
            cursor.execute("""
                SELECT nomer
                FROM tabel_mesin
                WHERE id_mesin = %s
            """, [machine_id_int])
            
            result = cursor.fetchone()
            
            if result and result[0]:
                return JsonResponse({'nomor': result[0]})
            else:
                return JsonResponse({'nomor': ''})
                
    except Exception as e:
        print(f"Error fetching nomor mesin {machine_id}: {e}")
        return JsonResponse({'error': 'Machine not found'}, status=404)
    
@login_required
def get_mesin_by_lokasi(request):
    """AJAX endpoint untuk dapetin daftar mesin berdasarkan lokasi yang dipilih - FIXED"""
    lokasi_id = request.GET.get('lokasi_id')
    
    if not lokasi_id:
        return JsonResponse({'mesin_list': []})
    
    mesin_list = []
    try:
        with connections['DB_Maintenance'].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT tm.id_mesin, tm.mesin, tm.nomer
                FROM tabel_mesin tm
                WHERE tm.id_line = %s 
                AND tm.status = 'A'
                AND tm.mesin IS NOT NULL
                AND tm.mesin != ''
                ORDER BY tm.mesin
            """, [lokasi_id])
            
            mesin_list = [
                {
                    'value': str(row[0]), 
                    'text': row[1], 
                    'nomer': row[2] if row[2] else ''
                } 
                for row in cursor.fetchall()
            ]
    except Exception as e:
        print(f"Error fetching mesin by lokasi: {e}")
        # Fallback data
        mesin_list = [
            {'value': '1', 'text': 'Mesin A', 'nomer': 'M001'},
            {'value': '2', 'text': 'Mesin B', 'nomer': 'M002'}
        ]
    
    return JsonResponse({'mesin_list': mesin_list})


@login_required 
def get_nomor_mesin(request):
    """AJAX endpoint untuk dapetin nomor mesin berdasarkan mesin yang dipilih"""
    mesin_id = request.GET.get('mesin_id')
    
    if not mesin_id:
        return JsonResponse({'nomor_mesin': ''})
    
    nomor_mesin = ''
    try:
        with connections['DB_Maintenance'].cursor() as cursor:
            cursor.execute("""
                SELECT nomer
                FROM tabel_mesin
                WHERE id_mesin = %s
            """, [mesin_id])
            
            result = cursor.fetchone()
            if result and result[0]:
                nomor_mesin = result[0]
    except Exception as e:
        print(f"Error fetching nomor mesin: {e}")
        nomor_mesin = f"M{mesin_id:03d}"  # Fallback format
    
    return JsonResponse({'nomor_mesin': nomor_mesin})


# @login_required
# def laporan_submit(request):
#     if request.method == 'POST':
#         tanggal = request.POST.get('tanggal')
#         shift_id = request.POST.get('shift')
#         masalah = request.POST.get('masalah')
#         catatan = request.POST.get('catatan')  # Tambahan untuk catatan
#         image = request.FILES.get('image')
#         pic_ids = request.POST.getlist('pic')  # PIC Laporan yang dipilih
#         pic_lembur_ids = request.POST.getlist('piclembur')  # PIC Lembur yang dipilih

#         # Debugging: Menampilkan data yang diterima dari form
#         print(f"Form Data - Tanggal: {tanggal}, Shift: {shift_id}, Masalah: {masalah}, Catatan: {catatan}, Image: {image}, PIC: {pic_ids}, PIC Lembur: {pic_lembur_ids}")

#         # Ambil instance Shift
#         try:
#             shift_instance = Shift.objects.get(id=shift_id)
#         except Shift.DoesNotExist:
#             messages.error(request, 'Shift tidak ditemukan.')
#             return redirect('dailyactivity_app:laporan_index')

#         # Ambil user_id dari pengguna yang sedang login
#         user_id = request.user.id

#         # Simpan data laporan
#         laporan_data = LaporanData.objects.create(
#             tanggal=tanggal,
#             shift=shift_instance,
#             user_id=user_id,
#             masalah=masalah,
#             catatan=catatan,  # Tambahan untuk catatan
#             image=image,
#         )

#         # Debugging: Menampilkan data laporan yang telah disimpan
#         print(f"Laporan Data saved: {laporan_data}")

#         # Menyimpan PIC Laporan
#         for pic_id in pic_ids:
#             try:
#                 pic_instance = PICLaporan.objects.get(id=pic_id)
#                 laporan_data.pic.add(pic_instance)
#             except PICLaporan.DoesNotExist:
#                 continue

#         # Debugging: Menampilkan data PIC Laporan yang ditambahkan
#         print(f"PIC Laporan added: {laporan_data.pic.all()}")

#         # Menyimpan PIC Lembur
#         for pic_lembur_id in pic_lembur_ids:
#             try:
#                 pic_lembur_instance = PICLembur.objects.get(id=pic_lembur_id)
#                 laporan_data.piclembur.add(pic_lembur_instance)
#             except PICLembur.DoesNotExist:
#                 continue

#         # Debugging: Menampilkan data PIC Lembur yang ditambahkan
#         print(f"PIC Lembur added: {laporan_data.piclembur.all()}")

#         # Simpan pesan sukses
#         messages.success(request, 'Data berhasil disimpan!')
#         return redirect('dailyactivity_app:laporan_index')

#     return redirect('dailyactivity_app:laporan_index')

# Update fungsi laporan_submit
@login_required
def laporan_submit(request):
    if request.method == 'POST':
        tanggal = request.POST.get('tanggal')
        shift_id = request.POST.get('shift')
        catatan = request.POST.get('catatan')  # Catatan khusus untuk keseluruhan
        image = request.FILES.get('image')
        pic_ids = request.POST.getlist('pic')
        pic_lembur_ids = request.POST.getlist('piclembur')
        
        # Ambil data multiple pekerjaan dari form
        deskripsi_list = request.POST.getlist('deskripsi_pekerjaan[]')
        jenis_pekerjaan_list = request.POST.getlist('jenis_pekerjaan[]')
        lama_pekerjaan_list = request.POST.getlist('lama_pekerjaan[]')
        pic_pekerjaan_list = request.POST.getlist('pic_pekerjaan[]')

        # Debugging
        print(f"Form Data - Tanggal: {tanggal}, Shift: {shift_id}")
        print(f"Pekerjaan - Deskripsi: {deskripsi_list}")
        print(f"Jenis: {jenis_pekerjaan_list}, Lama: {lama_pekerjaan_list}, PIC: {pic_pekerjaan_list}")

        # Ambil instance Shift
        try:
            shift_instance = Shift.objects.get(id=shift_id)
        except Shift.DoesNotExist:
            messages.error(request, 'Shift tidak ditemukan.')
            return redirect('dailyactivity_app:laporan_index')

        user_id = request.user.id

        try:
            with transaction.atomic():
                created_laporans = []
                
                # Simpan setiap pekerjaan sebagai record LaporanData terpisah
                for deskripsi, jenis, lama, pic_pekerjaan in zip(
                    deskripsi_list, jenis_pekerjaan_list, lama_pekerjaan_list, pic_pekerjaan_list
                ):
                    if deskripsi.strip():  # Hanya simpan jika ada deskripsi
                        laporan_data = LaporanData.objects.create(
                            tanggal=tanggal,
                            shift=shift_instance,
                            user_id=user_id,
                            masalah=deskripsi.strip(),  # Deskripsi pekerjaan
                            jenis_pekerjaan=jenis.strip(),
                            lama_pekerjaan=lama.strip(),
                            pic_pekerjaan=pic_pekerjaan.strip(),
                            catatan=catatan,  # Catatan sama untuk semua
                            image=image if len(created_laporans) == 0 else None,  # Image hanya di record pertama
                        )
                        created_laporans.append(laporan_data)

                # Set PIC dan PIC Lembur untuk semua record yang dibuat
                for laporan in created_laporans:
                    # Set PIC Laporan
                    for pic_id in pic_ids:
                        try:
                            pic_instance = PICLaporan.objects.get(id=pic_id)
                            laporan.pic.add(pic_instance)
                        except PICLaporan.DoesNotExist:
                            continue

                    # Set PIC Lembur
                    for pic_lembur_id in pic_lembur_ids:
                        try:
                            pic_lembur_instance = PICLembur.objects.get(id=pic_lembur_id)
                            laporan.piclembur.add(pic_lembur_instance)
                        except PICLembur.DoesNotExist:
                            continue

                print(f"Berhasil membuat {len(created_laporans)} record laporan")

        except Exception as e:
            messages.error(request, f'Kesalahan saat menyimpan data: {e}')
            return redirect('dailyactivity_app:laporan_index')

        messages.success(request, f'Data laporan utility berhasil disimpan! ({len(created_laporans)} pekerjaan)')
        return redirect('dailyactivity_app:laporan_index')

    return redirect('dailyactivity_app:laporan_index')


# @login_required
# def tanggal_laporan(request):
#     # Mengambil tanggal-tanggal unik dan mengurutkan berdasarkan tanggal secara descending
#     dates = LaporanData.objects.annotate(
#         date=TruncDate('tanggal', output_field=DateField())  # Truncate to date only
#     ).values('date').distinct().order_by('-date')  # Order by date descending

#     context = {
#         'dates': dates,
#     }
#     return render(request, 'dailyactivity_app/tanggal_laporan.html', context)   


# @login_required
# def data_laporan(request, tanggal):
#     tanggal_parsed = parse_date(tanggal)
#     # print(f"Parsed tanggal: {tanggal_parsed}")
#     laporan_data = LaporanData.objects.filter(tanggal=tanggal_parsed)
#     # print(f"Laporan data count: {laporan_data.count()}")  # Debug jumlah data
#     context = {
#         'laporan_data': laporan_data,
#         'selected_date': tanggal_parsed,
#     }
#     return render(request, 'dailyactivity_app/data_laporan.html', context)

# Update view data_laporan yang sudah ada
@login_required
def data_laporan(request, tanggal):
    tanggal_parsed = parse_date(tanggal)
    
    # Ambil semua laporan untuk tanggal tersebut, dikelompokkan per shift dan user
    laporan_data = LaporanData.objects.filter(tanggal=tanggal_parsed).order_by('shift', 'user', 'created_at')
    
    # Group laporan berdasarkan shift, user, dan waktu yang berdekatan (dalam 5 menit)
    grouped_laporan = {}
    
    for laporan in laporan_data:
        # Buat key berdasarkan shift, user, dan waktu (dibulatkan ke 5 menit)
        time_key = laporan.created_at.replace(second=0, microsecond=0)
        # Bulatkan ke 5 menit terdekat
        minute = (time_key.minute // 5) * 5
        time_key = time_key.replace(minute=minute)
        
        key = f"{laporan.shift.id}_{laporan.user.id}_{time_key.strftime('%Y%m%d_%H%M')}"
        
        if key not in grouped_laporan:
            grouped_laporan[key] = {
                'meta': {
                    'tanggal': laporan.tanggal,
                    'shift': laporan.shift,
                    'user': laporan.user,
                    'catatan': laporan.catatan,
                    'image': laporan.image,
                    'pic': laporan.pic.all(),
                    'piclembur': laporan.piclembur.all(),
                    'created_at': laporan.created_at,
                },
                'pekerjaan_list': []
            }
        
        grouped_laporan[key]['pekerjaan_list'].append(laporan)

    context = {
        'grouped_laporan': grouped_laporan,
        'selected_date': tanggal_parsed,
        'current_timestamp': timezone.now().timestamp(),  # Tambahkan timestamp untuk polling
    }
    return render(request, 'dailyactivity_app/data_laporan.html', context)

@login_required
def tanggal_laporan(request):
    # Ambil parameter bulan dan tahun dari URL jika ada
    selected_month = request.GET.get('month')
    selected_year = request.GET.get('year')
    
    print(f"DEBUG - Raw parameters: month='{selected_month}', year='{selected_year}'")
    
    if selected_month and selected_year:
        try:
            # Fungsi untuk membersihkan dan konversi nilai
            def clean_and_convert(value):
                if value is None:
                    return None
                
                # Convert to string terlebih dahulu
                str_value = str(value).strip()
                
                # Jika kosong, return None
                if not str_value:
                    return None
                
                # Jika ada decimal (seperti 2.025), ambil bagian integer saja
                if '.' in str_value:
                    # Split dan ambil bagian sebelum decimal
                    integer_part = str_value.split('.')[0]
                    return int(integer_part) if integer_part.isdigit() else None
                
                # Jika sudah integer/string digit biasa
                if str_value.isdigit():
                    return int(str_value)
                
                # Coba konversi float dulu baru ke int (untuk handle kasus lain)
                try:
                    return int(float(str_value))
                except (ValueError, TypeError):
                    return None
            
            # Konversi parameter
            selected_month = clean_and_convert(selected_month)
            selected_year = clean_and_convert(selected_year)
            
            print(f"DEBUG - Cleaned: month={selected_month}, year={selected_year}")
            
            # Validasi hasil konversi
            if selected_month is None or selected_year is None:
                messages.error(request, 'Parameter bulan atau tahun tidak dapat dikonversi')
                return redirect('dailyactivity_app:tanggal_laporan')
            
            # Validasi range
            if not (1 <= selected_month <= 12):
                messages.error(request, f'Bulan harus antara 1-12, diterima: {selected_month}')
                return redirect('dailyactivity_app:tanggal_laporan')
                
            # Validasi tahun dengan range yang realistis
            current_year = datetime.now().year
            if not (2020 <= selected_year <= current_year + 2):
                messages.error(request, f'Tahun harus antara 2020-{current_year + 2}, diterima: {selected_year}')
                return redirect('dailyactivity_app:tanggal_laporan')
            
            # Query data dengan error handling
            try:
                dates = LaporanData.objects.filter(
                    tanggal__month=selected_month,
                    tanggal__year=selected_year
                ).annotate(
                    date=TruncDate('tanggal', output_field=DateField())
                ).values('date').distinct().order_by('-date')
                
                dates_count = dates.count()
                print(f"DEBUG - Query successful, found {dates_count} dates")
                
            except Exception as query_error:
                print(f"DEBUG - Query error: {query_error}")
                messages.error(request, f'Error saat mengambil data: {query_error}')
                return redirect('dailyactivity_app:tanggal_laporan')
            
            # Nama bulan
            month_names = {
                1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
                5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
                9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
            }
            
            context = {
                'dates': dates,
                'selected_month': selected_month,
                'selected_year': selected_year,
                'selected_month_name': month_names.get(selected_month, f'Bulan {selected_month}'),
                'show_dates': True,
            }
            
        except Exception as e:
            print(f"DEBUG - General exception: {e}")
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
            return redirect('dailyactivity_app:tanggal_laporan')
    else:
        print("DEBUG - Showing month/year list")
        
        try:
            # Cek apakah ada data sama sekali
            total_count = LaporanData.objects.count()
            print(f"DEBUG - Total LaporanData records: {total_count}")
            
            if total_count == 0:
                messages.info(request, 'Belum ada data mechanical dalam sistem')
                context = {
                    'month_year_data': [],
                    'show_dates': False,
                }
                return render(request, 'dailyactivity_app/tanggal_laporan.html', context)
            
            # Ambil sample data untuk debugging
            sample_data = LaporanData.objects.first()
            if sample_data and sample_data.tanggal:
                print(f"DEBUG - Sample date: {sample_data.tanggal}, type: {type(sample_data.tanggal)}")
            
            # Ambil semua data dengan tanggal yang valid
            all_laporan_data = LaporanData.objects.filter(
                tanggal__isnull=False
            ).values('tanggal')
            
            print(f"DEBUG - Records with valid dates: {all_laporan_data.count()}")
            
            # Dictionary untuk menyimpan count
            month_year_dict = {}
            
            for data in all_laporan_data:
                tanggal = data['tanggal']
                if tanggal:
                    try:
                        month = tanggal.month
                        year = tanggal.year
                        key = f"{year}-{month:02d}"
                        
                        if key in month_year_dict:
                            month_year_dict[key]['count'] += 1
                        else:
                            month_year_dict[key] = {
                                'month': month,
                                'year': year,
                                'count': 1
                            }
                    except Exception as date_error:
                        print(f"DEBUG - Error processing date {tanggal}: {date_error}")
                        continue
            
            print(f"DEBUG - Month/Year dict keys: {list(month_year_dict.keys())}")
            
            # Konversi ke list
            month_names = {
                1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
                5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
                9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
            }
            
            data_list = []
            for key, data in month_year_dict.items():
                month = data['month']
                year = data['year']
                count = data['count']
                
                data_list.append({
                    'month': month,
                    'year': year,
                    'count': count,
                    'month_name': month_names.get(month, f'Bulan {month}')
                })
            
            # Sorting
            current_date = datetime.now()
            current_month = current_date.month
            current_year = current_date.year
            
            def get_sort_priority(item):
                year = item['year']
                month = item['month']
                
                if year == current_year:
                    if month <= current_month:
                        return (0, current_month - month)
                    else:
                        return (1, month)
                elif year < current_year:
                    return (2, -year, -month)
                else:
                    return (3, year, month)
            
            data_list.sort(key=get_sort_priority)
            
            print(f"DEBUG - Final data list count: {len(data_list)}")
            
            context = {
                'month_year_data': data_list,
                'show_dates': False,
            }
            
        except Exception as list_error:
            print(f"DEBUG - Error creating month/year list: {list_error}")
            messages.error(request, f'Error saat mengambil daftar bulan/tahun: {list_error}')
            context = {
                'month_year_data': [],
                'show_dates': False,
            }
    
    return render(request, 'dailyactivity_app/tanggal_laporan.html', context)


# @login_required
# def edit_laporan(request, id):
#     laporan_data = get_object_or_404(LaporanData, id=id)
#     shifts = Shift.objects.all()
#     pic_laporan = PICLaporan.objects.all()
#     pic_lembur_list = PICLembur.objects.all()  # Semua PIC Lembur

#     if request.method == 'POST':
#         form = LaporanDataForm(request.POST, request.FILES, instance=laporan_data)
#         if form.is_valid():
#             # Ambil data catatan dari POST
#             catatan = request.POST.get('catatan')
            
#             # Simpan form
#             updated_laporan = form.save(commit=False)
#             updated_laporan.catatan = catatan  # Set catatan
#             updated_laporan.save()

#             # Update PIC laporan
#             pic_ids = form.cleaned_data.get('pic')
#             laporan_data.pic.set(pic_ids)

#             # Update PIC lembur (gunakan `through` model)
#             pic_lembur_ids = request.POST.getlist('pic_lembur')
#             # Hapus data lama
#             LemburDataPIC.objects.filter(laporan_data=laporan_data).delete()
#             # Tambah data baru
#             for pic_lembur_id in pic_lembur_ids:
#                 pic_lembur = PICLembur.objects.get(id=pic_lembur_id)
#                 LemburDataPIC.objects.create(laporan_data=laporan_data, pic_lembur=pic_lembur)

#             messages.success(request, 'Data berhasil diperbarui!')
#             return redirect('dailyactivity_app:data_laporan', tanggal=laporan_data.tanggal)
#     else:
#         form = LaporanDataForm(instance=laporan_data)

#     context = {
#         'form': form,
#         'laporan': laporan_data,
#         'shifts': shifts,
#         'pic_laporan': pic_laporan,
#         'pic_lembur_list': pic_lembur_list,
#     }
#     return render(request, 'dailyactivity_app/edit_laporan.html', context)

# Update fungsi edit_laporan untuk memastikan updated_at ter-update
@login_required
def edit_laporan(request, id):
    # Ambil laporan yang akan diedit
    main_laporan = get_object_or_404(LaporanData, id=id)
    
    # Ambil semua laporan yang dibuat bersamaan (dalam rentang 5 menit)
    time_range_start = main_laporan.created_at - timedelta(minutes=5)
    time_range_end = main_laporan.created_at + timedelta(minutes=5)
    
    related_laporan = LaporanData.objects.filter(
        tanggal=main_laporan.tanggal,
        shift=main_laporan.shift,
        user=main_laporan.user,
        created_at__range=(time_range_start, time_range_end)
    ).order_by('created_at')
    
    shifts = Shift.objects.all()
    pic_laporan = PICLaporan.objects.all()
    pic_lembur_list = PICLembur.objects.all()

    if request.method == 'POST':
        tanggal = request.POST.get('tanggal')
        shift_id = request.POST.get('shift')
        catatan = request.POST.get('catatan')
        image = request.FILES.get('image')
        pic_ids = request.POST.getlist('pic')
        pic_lembur_ids = request.POST.getlist('piclembur')
        
        # Ambil data multiple pekerjaan dari form
        deskripsi_list = request.POST.getlist('deskripsi_pekerjaan[]')
        jenis_pekerjaan_list = request.POST.getlist('jenis_pekerjaan[]')
        lama_pekerjaan_list = request.POST.getlist('lama_pekerjaan[]')
        pic_pekerjaan_list = request.POST.getlist('pic_pekerjaan[]')

        try:
            shift_instance = Shift.objects.get(id=shift_id)
        except Shift.DoesNotExist:
            messages.error(request, 'Shift tidak ditemukan.')
            return redirect('dailyactivity_app:edit_laporan', id=id)

        try:
            with transaction.atomic():
                # Hapus semua laporan terkait
                related_laporan.delete()
                
                # Buat ulang dengan data baru
                created_laporans = []
                current_time = timezone.now()
                
                for deskripsi, jenis, lama, pic_pekerjaan in zip(
                    deskripsi_list, jenis_pekerjaan_list, lama_pekerjaan_list, pic_pekerjaan_list
                ):
                    if deskripsi.strip():
                        laporan_data = LaporanData.objects.create(
                            tanggal=tanggal,
                            shift=shift_instance,
                            user=main_laporan.user,
                            masalah=deskripsi.strip(),
                            jenis_pekerjaan=jenis.strip(),
                            lama_pekerjaan=lama.strip(),
                            pic_pekerjaan=pic_pekerjaan.strip(),
                            catatan=catatan,
                            image=image if len(created_laporans) == 0 else None,
                            created_at=current_time,  # Set sama agar grouping tetap sama
                        )
                        
                        # Update updated_at manually jika perlu
                        if hasattr(laporan_data, 'updated_at'):
                            laporan_data.updated_at = current_time
                            laporan_data.save()
                        
                        created_laporans.append(laporan_data)

                # Set PIC untuk semua record baru
                for laporan in created_laporans:
                    laporan.pic.set(pic_ids)
                    laporan.piclembur.set(pic_lembur_ids)

            messages.success(request, f'Data berhasil diperbarui! ({len(created_laporans)} pekerjaan)')
            return redirect('dailyactivity_app:data_laporan', tanggal=tanggal)
            
        except Exception as e:
            messages.error(request, f'Kesalahan saat memperbarui data: {e}')
            return redirect('dailyactivity_app:edit_laporan', id=id)

    context = {
        'main_laporan': main_laporan,
        'related_laporan': related_laporan,
        'shifts': shifts,
        'pic_laporan': pic_laporan,
        'pic_lembur_list': pic_lembur_list,
    }
    return render(request, 'dailyactivity_app/edit_laporan.html', context)

# @login_required
# def delete_laporan(request, id):
#     laporan_utility = get_object_or_404(LaporanData, id=id)
#     tanggal_laporan = laporan_utility.tanggal  # Ambil tanggal laporan yang akan dihapus
#     laporan_utility.delete()  # Hapus data
#     messages.success(request, 'Data berhasil dihapus!')
#     return redirect('dailyactivity_app:data_laporan', tanggal=tanggal_laporan)

@login_required
def delete_laporan(request, id):
    # Ambil laporan utama
    main_laporan = get_object_or_404(LaporanData, id=id)
    tanggal_laporan = main_laporan.tanggal
    
    # Ambil semua laporan terkait (dibuat bersamaan)
    time_range_start = main_laporan.created_at - timedelta(minutes=5)
    time_range_end = main_laporan.created_at + timedelta(minutes=5)
    
    related_laporan = LaporanData.objects.filter(
        tanggal=main_laporan.tanggal,
        shift=main_laporan.shift,
        user=main_laporan.user,
        created_at__range=(time_range_start, time_range_end)
    )
    
    # Hapus semua laporan terkait
    count = related_laporan.count()
    related_laporan.delete()
    
    messages.success(request, f'Data berhasil dihapus! ({count} pekerjaan)')
    return redirect('dailyactivity_app:data_laporan', tanggal=tanggal_laporan)

# Fungsi untuk menampilkan semua data PIC Electrical
@login_required
def piclaporan_index(request):
    piclaporans = PICLaporan.objects.all()  # Mengambil semua data dari model PicMechanical

    if request.method == "POST":
        piclaporan_name = request.POST.get('name')  # Nama PIC Mechanical
        piclaporan_nokar = request.POST.get('nokar')  # No. Karyawan PIC Mechanical
        
        if piclaporan_name and piclaporan_nokar:
            PICLaporan.objects.create(name=piclaporan_name, nokar=piclaporan_nokar)
            messages.success(request, "PIC Pelaporan added successfully!")
            return redirect('dailyactivity_app:piclaporan_index')

    return render(request, 'dailyactivity_app/piclaporan_index.html', {'piclaporans': piclaporans})

# Fungsi untuk mengedit data PIC Mechanical
def edit_piclaporan(request, pk):
    # Mendapatkan objek PICMechanical yang akan diedit
    piclaporan = get_object_or_404(PICLaporan, pk=pk)   
    # Memeriksa apakah request adalah POST (artinya data form dikirim)
    if request.method == 'POST':
        form = PICLaporanForm(request.POST, instance=piclaporan)
        if form.is_valid():
            form.save()  # Menyimpan perubahan pada objek yang ada
            return redirect('dailyactivity_app:piclaporan_index')  # Redirect ke halaman yang sesuai
    else:
        form = PICLaporanForm(instance=piclaporan)  # Mengisi form dengan data yang ada   
    return render(request, 'dailyactivity_app/edit_piclaporan.html', {'form': form})
# Fungsi untuk menghapus data PIC Mechanical

def delete_piclaporan(request, pk):
    piclaporan = get_object_or_404(PICLaporan, pk=pk)  # Mengambil data PIC Mechanical berdasarkan pk
    piclaporan.delete()
    messages.success(request, "PIC Electrical deleted successfully!")
    return redirect('dailyactivity_app:piclaporan_index')

@login_required
def piclembur_index(request):
    piclemburs = PICLembur.objects.all()  # Mengambil semua data dari model PicMechanical

    if request.method == "POST":
        piclembur_name = request.POST.get('name')  # Nama PIC Mechanical
        piclembur_nokar = request.POST.get('nokar')  # No. Karyawan PIC Mechanical
        
        if piclembur_name and piclembur_nokar:
            PICLaporan.objects.create(name=piclembur_name, nokar=piclembur_nokar)
            messages.success(request, "PIC Pelaporan added successfully!")
            return redirect('dailyactivity_app:piclembur_index')
    return render(request, 'dailyactivity_app/piclembur_index.html', {'piclemburs': piclemburs})

# Fungsi untuk mengedit data PIC Mechanical
def edit_piclembur(request, pk):
    # Mendapatkan objek PICMechanical yang akan diedit
    piclembur = get_object_or_404(PICLembur, pk=pk)   
    # Memeriksa apakah request adalah POST (artinya data form dikirim)
    if request.method == 'POST':
        form = PICLemburForm(request.POST, instance=piclembur)
        if form.is_valid():
            form.save()  # Menyimpan perubahan pada objek yang ada
            return redirect('dailyactivity_app:piclembur_index')  # Redirect ke halaman yang sesuai
    else:
        form = PICLemburForm(instance=piclembur)  # Mengisi form dengan data yang ada 
    return render(request, 'dailyactivity_app/edit_piclembur.html', {'form': form})

# Fungsi untuk menghapus data PIC Mechanical
def delete_piclembur(request, pk):
    piclembur = get_object_or_404(PICLembur, pk=pk)  # Mengambil data PIC Mechanical berdasarkan pk
    piclembur.delete()
    messages.success(request, "PIC Electrical deleted successfully!")
    return redirect('dailyactivity_app:piclembur_index')


@login_required
def laporanmechanical_index(request):
    shifts = Shift.objects.all()
    locations = Location.objects.all()
    machines = Machinemechanical.objects.all()
    categories = Category.objects.all()
    status = Status.objects.all()
    pic_mechanical = PICMechanical2.objects.all()

    # Ambil daftar nomor WO dari section mechanical (id_section = 4)
    nomor_wo_list = []
    with connections['DB_Maintenance'].cursor() as cursor:
        cursor.execute("""
            SELECT number_wo
            FROM dbo.view_main
            WHERE id_section = 4
            AND YEAR(tgl_his) BETWEEN 2023 AND 2024
            ORDER BY history_id DESC
        """)
        nomor_wo_list = [row[0] for row in cursor.fetchall()]

    # Default values
    deskripsi_perbaikan = None
    tgl_his = None
    penyebab = None
    line = None
    mesin = None
    nomer = None
    pekerjaan = None
    status_pekerjaan = None
    tindakan_perbaikan = None
    tindakan_pencegahan = None

    # Tangani Form Submission
    if request.method == 'POST':
        form = MechanicalData2Form(request.POST, request.FILES)
        if form.is_valid():
            # Proses data form
            machine_number = form.cleaned_data.get('machine_number')
            machine_instance = form.cleaned_data.get('machine')

            # Cek apakah mesin sudah memiliki nomor
            if machine_number and machine_instance and not machine_instance.nomor:
                machine_instance.nomor = machine_number
                machine_instance.save()

            # Simpan data MechanicalData
            mechanical_data = form.save(commit=False)
            mechanical_data.user = request.user
            mechanical_data.machine = machine_instance
            mechanical_data.save()

            # Hubungkan PIC dengan MechanicalData
            pic_ids = form.cleaned_data.get('pic')
            mechanical_data.pic.set(pic_ids)

            # Simpan nomor WO dan waktu pengerjaan
            mechanical_data.nomor_wo = form.cleaned_data.get('nomor_wo')
            mechanical_data.waktu_pengerjaan = form.cleaned_data.get('waktu_pengerjaan')
            mechanical_data.save()
            
            return redirect('success_page')
    else:
        form = MechanicalData2Form()

    # Tangani Permintaan AJAX untuk nomor WO
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.method == 'GET':
        nomor_wo_selected = request.GET.get('nomor_wo')
        with connections['DB_Maintenance'].cursor() as cursor:
            cursor.execute("""
                SELECT deskripsi_perbaikan, tgl_his, penyebab, line, mesin, nomer, pekerjaan, status_pekerjaan, tindakan_perbaikan, tindakan_pencegahan
                FROM dbo.view_main
                WHERE number_wo = %s
            """, [nomor_wo_selected])
            row = cursor.fetchone()
            if row:
                deskripsi_perbaikan, tgl_his, penyebab, line, mesin, nomer, pekerjaan, status_pekerjaan, tindakan_perbaikan, tindakan_pencegahan = row

        return JsonResponse({
            'deskripsi_perbaikan': deskripsi_perbaikan,
            'tgl_his': tgl_his,
            'penyebab': penyebab,
            'line': line,
            'mesin': mesin,
            'nomer': nomer,
            'pekerjaan': pekerjaan,
            'status_pekerjaan': status_pekerjaan,
            'tindakan_perbaikan': tindakan_perbaikan,
            'tindakan_pencegahan': tindakan_pencegahan
        })

    # Context untuk template
    context = {
        'shifts': shifts,
        'locations': locations,
        'machines': machines,
        'categories': categories,
        'status': status,
        'pic_mechanical': pic_mechanical,
        'nomor_wo_list': nomor_wo_list,
        'deskripsi_perbaikan': deskripsi_perbaikan,
        'tgl_his': tgl_his,
        'penyebab': penyebab,
        'line': line,
        'mesin': mesin,
        'nomer': nomer,
        'pekerjaan': pekerjaan,
        'status_pekerjaan': status_pekerjaan,
        'tindakan_perbaikan': tindakan_perbaikan,
        'tindakan_pencegahan': tindakan_pencegahan,
        'form': form,
    }
    return render(request, 'dailyactivity_app/laporanmechanical_index.html', context)


@login_required
def laporanmechanical_submit(request):
    if request.method == 'POST':
        tanggal = request.POST.get('tanggal')
        shift_id = request.POST.get('shift')
        masalah_list = request.POST.getlist('masalah[]')
        jenis_pekerjaan_list = request.POST.getlist('jenis_pekerjaan[]')
        lama_pekerjaan_list = request.POST.getlist('lama_pekerjaan[]')
        pic_masalah_list = request.POST.getlist('pic_masalah[]')
        pic_ids = request.POST.getlist('pic')
        pic_lembur_ids = request.POST.getlist('piclembur')
        image = request.FILES.get('image')

        try:
            shift_instance = Shift.objects.get(id=shift_id)
        except Shift.DoesNotExist:
            messages.error(request, 'Shift tidak ditemukan.')
            return redirect('dailyactivity_app:laporanmechanical_index')

        user_id = request.user.id

        try:
            with transaction.atomic():  # Gunakan transaksi untuk memastikan konsistensi data
                # Simpan data utama ke tabel LaporanMechanicalData
                laporan_data = LaporanMechanicalData.objects.create(
                    tanggal=tanggal,
                    shift=shift_instance,
                    user_id=user_id,
                    image=image,
                )

                # Simpan setiap baris pekerjaan ke tabel dailyactivity_app_laporanpekerjaan
                for masalah, jenis_pekerjaan, lama_pekerjaan, pic_masalah in zip(masalah_list, jenis_pekerjaan_list, lama_pekerjaan_list, pic_masalah_list):
                    LaporanPekerjaan.objects.create(
                        laporan_data_id=laporan_data.id,
                        masalah=masalah,
                        jenis_pekerjaan=jenis_pekerjaan,
                        lama_pekerjaan=lama_pekerjaan,
                        pic_masalah=pic_masalah,
                    )

                # Simpan PIC dan PIC Lembur (ManyToMany relationships)
                for pic_id in pic_ids:
                    try:
                        pic_instance = PICLaporanMechanical.objects.get(id=pic_id)
                        laporan_data.pic.add(pic_instance)
                    except PICLaporanMechanical.DoesNotExist:
                        continue

                for pic_lembur_id in pic_lembur_ids:
                    try:
                        pic_lembur_instance = PICLemburMechanical.objects.get(id=pic_lembur_id)
                        laporan_data.piclembur.add(pic_lembur_instance)
                    except PICLemburMechanical.DoesNotExist:
                        continue

            messages.success(request, 'Data berhasil disimpan!')
        except Exception as e:
            messages.error(request, f'Kesalahan saat menyimpan data: {e}')
        
        return redirect('dailyactivity_app:laporanmechanical_index')

    return redirect('dailyactivity_app:laporanmechanical_index')

@login_required
def tanggal_laporanmechanical(request):
    # Ambil parameter bulan dan tahun dari URL jika ada
    selected_month = request.GET.get('month')
    selected_year = request.GET.get('year')
    
    print(f"DEBUG - Raw parameters: month='{selected_month}', year='{selected_year}'")
    
    if selected_month and selected_year:
        try:
            # Fungsi untuk membersihkan dan konversi nilai
            def clean_and_convert(value):
                if value is None:
                    return None
                
                # Convert to string terlebih dahulu
                str_value = str(value).strip()
                
                # Jika kosong, return None
                if not str_value:
                    return None
                
                # Jika ada decimal (seperti 2.025), ambil bagian integer saja
                if '.' in str_value:
                    # Split dan ambil bagian sebelum decimal
                    integer_part = str_value.split('.')[0]
                    return int(integer_part) if integer_part.isdigit() else None
                
                # Jika sudah integer/string digit biasa
                if str_value.isdigit():
                    return int(str_value)
                
                # Coba konversi float dulu baru ke int (untuk handle kasus lain)
                try:
                    return int(float(str_value))
                except (ValueError, TypeError):
                    return None
            
            # Konversi parameter
            selected_month = clean_and_convert(selected_month)
            selected_year = clean_and_convert(selected_year)
            
            print(f"DEBUG - Cleaned: month={selected_month}, year={selected_year}")
            
            # Validasi hasil konversi
            if selected_month is None or selected_year is None:
                messages.error(request, 'Parameter bulan atau tahun tidak dapat dikonversi')
                return redirect('dailyactivity_app:tanggal_laporanmechanical')
            
            # Validasi range
            if not (1 <= selected_month <= 12):
                messages.error(request, f'Bulan harus antara 1-12, diterima: {selected_month}')
                return redirect('dailyactivity_app:tanggal_laporanmechanical')
                
            # Validasi tahun dengan range yang realistis
            current_year = datetime.now().year
            if not (2020 <= selected_year <= current_year + 2):
                messages.error(request, f'Tahun harus antara 2020-{current_year + 2}, diterima: {selected_year}')
                return redirect('dailyactivity_app:tanggal_laporanmechanical')
            
            # Query data dengan error handling
            try:
                dates = LaporanMechanicalData.objects.filter(
                    tanggal__month=selected_month,
                    tanggal__year=selected_year
                ).annotate(
                    date=TruncDate('tanggal', output_field=DateField())
                ).values('date').distinct().order_by('-date')
                
                dates_count = dates.count()
                print(f"DEBUG - Query successful, found {dates_count} dates")
                
            except Exception as query_error:
                print(f"DEBUG - Query error: {query_error}")
                messages.error(request, f'Error saat mengambil data: {query_error}')
                return redirect('dailyactivity_app:tanggal_laporanmechanical')
            
            # Nama bulan
            month_names = {
                1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
                5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
                9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
            }
            
            context = {
                'dates': dates,
                'selected_month': selected_month,
                'selected_year': selected_year,
                'selected_month_name': month_names.get(selected_month, f'Bulan {selected_month}'),
                'show_dates': True,
            }
            
        except Exception as e:
            print(f"DEBUG - General exception: {e}")
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
            return redirect('dailyactivity_app:tanggal_laporanmechanical')
    else:
        print("DEBUG - Showing month/year list")
        
        try:
            # Cek apakah ada data sama sekali
            total_count = LaporanMechanicalData.objects.count()
            print(f"DEBUG - Total LaporanMechanicalData records: {total_count}")
            
            if total_count == 0:
                messages.info(request, 'Belum ada data mechanical dalam sistem')
                context = {
                    'month_year_data': [],
                    'show_dates': False,
                }
                return render(request, 'dailyactivity_app/tanggal_laporanmechanical.html', context)
            
            # Ambil sample data untuk debugging
            sample_data = LaporanMechanicalData.objects.first()
            if sample_data and sample_data.tanggal:
                print(f"DEBUG - Sample date: {sample_data.tanggal}, type: {type(sample_data.tanggal)}")
            
            # Ambil semua data dengan tanggal yang valid
            all_laporanmechanical_data = LaporanMechanicalData.objects.filter(
                tanggal__isnull=False
            ).values('tanggal')
            
            print(f"DEBUG - Records with valid dates: {all_laporanmechanical_data.count()}")
            
            # Dictionary untuk menyimpan count
            month_year_dict = {}
            
            for data in all_laporanmechanical_data:
                tanggal = data['tanggal']
                if tanggal:
                    try:
                        month = tanggal.month
                        year = tanggal.year
                        key = f"{year}-{month:02d}"
                        
                        if key in month_year_dict:
                            month_year_dict[key]['count'] += 1
                        else:
                            month_year_dict[key] = {
                                'month': month,
                                'year': year,
                                'count': 1
                            }
                    except Exception as date_error:
                        print(f"DEBUG - Error processing date {tanggal}: {date_error}")
                        continue
            
            print(f"DEBUG - Month/Year dict keys: {list(month_year_dict.keys())}")
            
            # Konversi ke list
            month_names = {
                1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
                5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
                9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
            }
            
            data_list = []
            for key, data in month_year_dict.items():
                month = data['month']
                year = data['year']
                count = data['count']
                
                data_list.append({
                    'month': month,
                    'year': year,
                    'count': count,
                    'month_name': month_names.get(month, f'Bulan {month}')
                })
            
            # Sorting
            current_date = datetime.now()
            current_month = current_date.month
            current_year = current_date.year
            
            def get_sort_priority(item):
                year = item['year']
                month = item['month']
                
                if year == current_year:
                    if month <= current_month:
                        return (0, current_month - month)
                    else:
                        return (1, month)
                elif year < current_year:
                    return (2, -year, -month)
                else:
                    return (3, year, month)
            
            data_list.sort(key=get_sort_priority)
            
            print(f"DEBUG - Final data list count: {len(data_list)}")
            
            context = {
                'month_year_data': data_list,
                'show_dates': False,
            }
            
        except Exception as list_error:
            print(f"DEBUG - Error creating month/year list: {list_error}")
            messages.error(request, f'Error saat mengambil daftar bulan/tahun: {list_error}')
            context = {
                'month_year_data': [],
                'show_dates': False,
            }
    
    return render(request, 'dailyactivity_app/tanggal_laporanmechanical.html', context)


@login_required
def data_laporanmechanical(request, tanggal, laporan_id=None):
    tanggal_parsed = parse_date(tanggal)
    laporan_data = LaporanMechanicalData.objects.filter(tanggal=tanggal_parsed)

    # Tangani jika ada komentar yang dikirimkan
    if request.method == 'POST':
        if laporan_id:
            komentar = request.POST.get('comment')
            laporan = get_object_or_404(LaporanMechanicalData, id=laporan_id)
            laporan.catatan = komentar
            laporan.save()
            return redirect('dailyactivity_app:data_laporanmechanical', tanggal=tanggal)

    context = {
        'laporan_data': laporan_data,
        'selected_date': tanggal_parsed,
    }
    return render(request, 'dailyactivity_app/data_laporanmechanical.html', context)


@login_required
def add_comment(request, laporan_id):
    # Mencari data laporan berdasarkan ID terlebih dahulu
    laporan = get_object_or_404(LaporanMechanicalData, id=laporan_id)
    tanggal = laporan.tanggal  # Ambil tanggal dari laporan

    if request.method == 'POST':
        komentar = request.POST.get('comment')

        # Menyimpan komentar ke kolom 'catatan'
        laporan.catatan = komentar
        laporan.save()

        # Redirect ke halaman laporan dengan tanggal yang benar
        return redirect('dailyactivity_app:data_laporanmechanical', tanggal=tanggal)

    # Jika bukan POST, tetap redirect ke halaman yang sama (tanpa perubahan)
    return redirect('dailyactivity_app:data_laporanmechanical', tanggal=tanggal)

@login_required
def edit_laporanmechanical(request, id):
    # Ambil data laporan yang akan diedit
    laporan_data = get_object_or_404(LaporanMechanicalData, id=id)

    # Ambil data shift, PIC, dan PIC Lembur yang ada
    shifts = Shift.objects.all()
    pic_laporan = PICLaporanMechanical.objects.all()
    pic_lembur = PICLemburMechanical.objects.all()

    # Ambil pekerjaan terkait
    pekerjaan_list = laporan_data.laporan_pekerjaan.all()

    if request.method == 'POST':
        # Ambil data POST
        tanggal = request.POST.get('tanggal')
        shift_id = request.POST.get('shift')
        masalah_list = request.POST.getlist('masalah[]')
        jenis_pekerjaan_list = request.POST.getlist('jenis_pekerjaan[]')
        lama_pekerjaan_list = request.POST.getlist('lama_pekerjaan[]')
        pic_masalah_list = request.POST.getlist('pic_masalah[]')
        pic_ids = request.POST.getlist('pic')
        pic_lembur_ids = request.POST.getlist('piclembur')
        image = request.FILES.get('image')

        try:
            shift_instance = Shift.objects.get(id=shift_id)
        except Shift.DoesNotExist:
            messages.error(request, 'Shift tidak ditemukan.')
            return redirect('dailyactivity_app:edit_laporanmechanical', id=id)

        try:
            with transaction.atomic():  # Gunakan transaksi untuk memastikan konsistensi data
                # Perbarui data laporan utama
                laporan_data.tanggal = tanggal
                laporan_data.shift = shift_instance
                if image:
                    laporan_data.image = image  # Hanya update jika ada file baru
                laporan_data.save()

                # Hapus pekerjaan lama terkait laporan dan tambahkan yang baru
                laporan_data.laporan_pekerjaan.all().delete()
                for masalah, jenis_pekerjaan, lama_pekerjaan, pic_masalah in zip(
                    masalah_list, jenis_pekerjaan_list, lama_pekerjaan_list, pic_masalah_list
                ):
                    LaporanPekerjaan.objects.create(
                        laporan_data_id=laporan_data.id,
                        masalah=masalah,
                        jenis_pekerjaan=jenis_pekerjaan,
                        lama_pekerjaan=lama_pekerjaan,
                        pic_masalah=pic_masalah,
                    )

                # Perbarui ManyToMany untuk PIC dan PIC Lembur
                laporan_data.pic.set(pic_ids)
                laporan_data.piclembur.set(pic_lembur_ids)

            messages.success(request, 'Data berhasil diperbarui!')
            return redirect('dailyactivity_app:data_laporanmechanical', tanggal=laporan_data.tanggal)

        except Exception as e:
            messages.error(request, f'Kesalahan saat memperbarui data: {e}')
            return redirect('dailyactivity_app:edit_laporanmechanical', id=id)

    # Jika request GET, tampilkan form edit
    context = {
        'laporan': laporan_data,
        'shifts': shifts,
        'pic_laporan': pic_laporan,
        'pic_lembur': pic_lembur,
        'pekerjaan_list': pekerjaan_list,
        'selected_pic': laporan_data.pic.values_list('id', flat=True),
        'selected_piclembur': laporan_data.piclembur.values_list('id', flat=True),
    }
    return render(request, 'dailyactivity_app/edit_laporanmechanical.html', context)

@login_required
def delete_laporanmechanical(request, id):
    laporan_data = get_object_or_404(LaporanMechanicalData, id=id)
    tanggal_laporan = laporan_data.tanggal  # Ambil tanggal laporan yang akan dihapus
    laporan_data.delete()  # Hapus data
    messages.success(request, 'Data berhasil dihapus!')
    return redirect('dailyactivity_app:data_laporanmechanical', tanggal=tanggal_laporan)

# Fungsi untuk menampilkan semua data PIC Electrical
@login_required
def piclaporanmechanical_index(request):
    piclaporans = PICLaporanMechanical.objects.all()  # Mengambil semua data dari model PicMechanical

    if request.method == "POST":
        piclaporan_name = request.POST.get('name')  # Nama PIC Mechanical
        piclaporan_nokar = request.POST.get('nokar')  # No. Karyawan PIC Mechanical
        
        if piclaporan_name and piclaporan_nokar:
            PICLaporanMechanical.objects.create(name=piclaporan_name, nokar=piclaporan_nokar)
            messages.success(request, "PIC Pelaporan added successfully!")
            return redirect('dailyactivity_app:piclaporanmechanical_index')

    return render(request, 'dailyactivity_app/piclaporanmechanical_index.html', {'piclaporans': piclaporans})

# Fungsi untuk mengedit data PIC Mechanical
def edit_piclaporanmechanical(request, pk):
    # Mendapatkan objek PICMechanical yang akan diedit
    piclaporan = get_object_or_404(PICLaporanMechanical, pk=pk)   
    # Memeriksa apakah request adalah POST (artinya data form dikirim)
    if request.method == 'POST':
        form = PICLaporanMechanicalForm(request.POST, instance=piclaporan)
        if form.is_valid():
            form.save()  # Menyimpan perubahan pada objek yang ada
            return redirect('dailyactivity_app:piclaporanmechanical_index')  # Redirect ke halaman yang sesuai
    else:
        form = PICLaporanMechanicalForm(instance=piclaporan)  # Mengisi form dengan data yang ada   
    return render(request, 'dailyactivity_app/edit_piclaporanmechanical.html', {'form': form})
# Fungsi untuk menghapus data PIC Mechanical

def delete_piclaporanmechanical(request, pk):
    piclaporan = get_object_or_404(PICLaporanMechanical, pk=pk)  # Mengambil data PIC Mechanical berdasarkan pk
    piclaporan.delete()
    messages.success(request, "PIC Electrical deleted successfully!")
    return redirect('dailyactivity_app:piclaporanmechanical_index')

@login_required
def piclemburmechanical_index(request):
    piclemburs = PICLemburMechanical.objects.all()  # Mengambil semua data dari model PicMechanical

    if request.method == "POST":
        piclembur_name = request.POST.get('name')  # Nama PIC Mechanical
        piclembur_nokar = request.POST.get('nokar')  # No. Karyawan PIC Mechanical
        
        if piclembur_name and piclembur_nokar:
            PICLaporanMechanical.objects.create(name=piclembur_name, nokar=piclembur_nokar)
            messages.success(request, "PIC Pelaporan added successfully!")
            return redirect('dailyactivity_app:piclemburmechanical_index')
    return render(request, 'dailyactivity_app/piclemburmechanical_index.html', {'piclemburs': piclemburs})

# Fungsi untuk mengedit data PIC Mechanical
def edit_piclemburmechanical(request, pk):
    # Mendapatkan objek PICMechanical yang akan diedit
    piclembur = get_object_or_404(PICLemburMechanical, pk=pk)   
    # Memeriksa apakah request adalah POST (artinya data form dikirim)
    if request.method == 'POST':
        form = PICLemburMechanicalForm(request.POST, instance=piclembur)
        if form.is_valid():
            form.save()  # Menyimpan perubahan pada objek yang ada
            return redirect('dailyactivity_app:piclemburmechanical_index')  # Redirect ke halaman yang sesuai
    else:
        form = PICLemburMechanicalForm(instance=piclembur)  # Mengisi form dengan data yang ada 
    return render(request, 'dailyactivity_app/edit_piclemburmechanical.html', {'form': form})

# Fungsi untuk menghapus data PIC Mechanical
def delete_piclemburmechanical(request, pk):
    piclembur = get_object_or_404(PICLemburMechanical, pk=pk)
    piclembur.delete()
    messages.success(request, "PIC Electrical deleted successfully!")
    return redirect('dailyactivity_app:piclemburmechanical_index')

@login_required
def schedule_index(request):
    # Mendapatkan data Shift dan PIC untuk pilihan di form
    shifts = Shift.objects.all()
    pic_lembur = PICLemburMechanical.objects.all()

    # Menangani form jika dikirim dengan POST
    if request.method == 'POST':
        # Mengambil data POST dan file (gambar)
        tanggal = request.POST.get('tanggal')
        shift_id = request.POST.get('shift')
        pic_ids = request.POST.getlist('pic')  # multiple PIC
        pekerjaan = request.POST.get('pekerjaan')
        image = request.FILES.get('image')
        
        # Menangani jam kerja untuk setiap PIC
        pic_jam_data = {}
        for pic_id in pic_ids:
            jam_key = f'input_{pic_id}'  # Nama input untuk jam berdasarkan PIC
            jam_value = request.POST.get(jam_key)  # Mendapatkan nilai jam untuk PIC tersebut
            if jam_value:
                pic_jam_data[pic_id] = jam_value  # Menyimpan jam yang dimasukkan

        # Mendapatkan objek shift berdasarkan ID
        try:
            shift = Shift.objects.get(id=shift_id)
        except Shift.DoesNotExist:
            messages.error(request, "Shift yang dipilih tidak valid.")
            return redirect('dailyactivity_app:schedule_index')

        # Membuat dan menyimpan objek ScheduleMechanicalData
        schedule = ScheduleMechanicalData(
            tanggal=tanggal,
            shift=shift,
            pekerjaan=pekerjaan,
            image=image,
            user=request.user, 
        )
        schedule.save()

        # Menambahkan PIC yang dipilih ke ManyToMany field
        for pic_id in pic_ids:
            try:
                pic = PICLemburMechanical.objects.get(id=pic_id)
                schedule.pic_lemburmechanical.add(pic)

                # Menambahkan jam untuk PIC yang bersangkutan
                if pic_id in pic_jam_data:
                    schedule.jam_set.create(pic=pic, jam=pic_jam_data[pic_id])
            except PICLemburMechanical.DoesNotExist:
                messages.error(request, f"PIC dengan ID {pic_id} tidak ditemukan.")
                continue

        # Menampilkan pesan sukses dan redirect
        messages.success(request, "Data Jadwal Pekerjaan berhasil disimpan!")
        return redirect('dailyactivity_app:schedule_index')  # Redirect kembali ke form input   
    
    # Menampilkan form kosong ketika pertama kali membuka halaman
    return render(request, 'dailyactivity_app/schedule_index.html', {
        'shifts': shifts,
        'pic_lembur': pic_lembur
    })

@login_required
def schedule_submit(request):
    if request.method == 'POST':
        tanggal = request.POST.get('tanggal') 
        shift_id = request.POST.get('shift')
        pic_ids = request.POST.getlist('pic')  # multiple PIC
        pekerjaan = request.POST.get('pekerjaan')
        image = request.FILES.get('image')

        # Mendapatkan objek Shift
        try:
            shift = Shift.objects.get(id=shift_id)
        except Shift.DoesNotExist:
            messages.error(request, "Shift tidak valid.")
            return redirect('dailyactivity_app:schedule_index')

        # Simpan jadwal
        schedule = ScheduleMechanicalData(
            tanggal=tanggal,
            shift=shift,
            pekerjaan=pekerjaan,
            image=image,
            user=request.user,
        )
        schedule.save()

        # Simpan PIC dan jam terkait
        for pic_id in pic_ids:
            try:
                pic = PICLemburMechanical.objects.get(id=pic_id)
                jam_key = f'jam_{pic_id}'  # Nama input untuk jam
                jam_value = request.POST.get(jam_key)

                if jam_value:
                    # Tambahkan PIC dan jam ke jadwal
                    schedule.pic_lemburmechanical.add(pic)
                    schedule.jam_set.create(pic=pic, jam=jam_value)  # Contoh relasi Many-to-Many dengan tambahan
            except PICLemburMechanical.DoesNotExist:
                continue

        messages.success(request, "Jadwal berhasil disimpan!")
        return redirect('dailyactivity_app:schedule_index')
    
    return redirect('dailyactivity_app:schedule_index')

@login_required
def tanggal_schedule(request):
    # Mengambil tanggal-tanggal unik dan mengurutkan berdasarkan tanggal secara descending
    dates = ScheduleMechanicalData.objects.annotate(
        date=Trunc('tanggal', 'day', output_field=DateField())
    ).values('date').distinct().order_by('-date')
    
    # Menambahkan nomor urut manual
    dates_with_number = []
    for index, date_obj in enumerate(dates, 1):
        dates_with_number.append({
            'number': index,
            'date': date_obj['date']
        })

    context = {
        'dates': dates_with_number,
    }
    return render(request, 'dailyactivity_app/tanggal_schedule.html', context)

@login_required
def data_schedule(request, tanggal):
    tanggal_parsed = parse_date(tanggal)
    print(f"Parsed tanggal: {tanggal_parsed}")
    laporan_data = ScheduleMechanicalData.objects.filter(tanggal=tanggal_parsed)
    print(f"Laporan data count: {laporan_data.count()}")  # Debug jumlah data

    context = {
        'laporan_data': laporan_data,
        'selected_date': tanggal_parsed,
    }
    return render(request, 'dailyactivity_app/data_schedule.html', context)

@login_required
def edit_schedule(request, id):
    # Fetch the schedule data by ID
    schedule_data = get_object_or_404(ScheduleMechanicalData, id=id)

    if request.method == 'POST':
        # Bind the form to the POST data and the current instance
        form = ScheduleMechanicalDataForm(request.POST, request.FILES, instance=schedule_data)
        if form.is_valid():
            form.save()
            messages.success(request, 'Schedule updated successfully!')
            # Redirect to the schedule detail or listing page
            return redirect('dailyactivity_app:data_schedule', tanggal=schedule_data.tanggal)
        else:
            # Add an error message if form validation fails
            messages.error(request, 'Failed to update the schedule. Please check the form.')
    else:
        # Initialize the form with the existing schedule data
        form = ScheduleMechanicalDataForm(instance=schedule_data)

    # Context for the template
    context = {
        'form': form,
        'schedule_data': schedule_data,  # Use a consistent variable name for schedule data
        'selected_date': schedule_data.tanggal,  # Keep selected_date for date context
    }

    # Render the edit schedule template
    return render(request, 'dailyactivity_app/edit_schedule.html', context)


@login_required
def delete_schedule(request, id):
    schedule = get_object_or_404(ScheduleMechanicalData, id=id)
    schedule.delete()
    return redirect('dailyactivity_app:data_schedule', tanggal=schedule.tanggal)  # Redirect kembali ke halaman yang sesuai

@login_required
def it_project(request):
    if request.method == 'POST':
        # Ambil data dari form
        project_name = request.POST.get('project_name')
        pic_project = request.POST.get('pic_project')
        department = request.POST.get('department')
        start_date = request.POST.get('start_date')
        finish_date = request.POST.get('finish_date')

        if not all([project_name, pic_project, department, start_date, finish_date]):
            messages.error(request, "Semua field wajib diisi.")
            return render(request, 'dailyactivity_app/it_project.html')

        # Simpan data proyek
        try:
            project = Project.objects.create(
                project_name=project_name,
                pic_project=pic_project,
                department=department,
                start_date=start_date,
                finish_date=finish_date
            )
        except Exception as e:
            messages.error(request, f"Terjadi kesalahan saat menyimpan proyek: {str(e)}")
            return render(request, 'dailyactivity_app/it_project.html')

        # Menyimpan setiap issue terkait proyek
        issues_count = len([key for key in request.POST.keys() if key.startswith('issue_')])
        for i in range(1, issues_count + 1):
            issue = request.POST.get(f'issue_{i}', '').strip()
            if not issue:
                continue  # Lewati issue jika tidak diisi

            pic = request.POST.get(f'issue_pic_{i}', '').strip() or None
            due_date = request.POST.get(f'issue_due_{i}', '').strip() or None
            status = request.POST.get(f'issue_status_{i}', 'Pending').strip()
            remark = request.POST.get(f'issue_remark_{i}', '').strip() or None

            try:
                ProjectIssue.objects.create(
                    project=project,
                    issue=issue,
                    pic=pic,
                    due_date=due_date,
                    status=status,
                    remark=remark
                )
            except Exception as e:
                messages.error(request, f"Terjadi kesalahan saat menyimpan issue {i}: {str(e)}")

        # Berikan feedback sukses
        messages.success(request, "Proyek dan issue berhasil disimpan.")
        return render(request, 'dailyactivity_app/it_project.html', {'project': project})

    # Jika GET, tampilkan halaman kosong
    return render(request, 'dailyactivity_app/it_project.html')

@login_required
def it_tanggal_project(request):
    search_query = request.GET.get('search', '')
    
    # Base query
    project_issues = ProjectIssue.objects.all()
    
    # Apply search filter jika ada query
    if search_query:
        project_issues = project_issues.filter(
            project__project_name__icontains=search_query
        )
    
    # Ambil tanggal unik dari hasil query
    dates = project_issues.annotate(
        date=TruncDate('due_date', output_field=DateField())
    ).values('date').distinct().order_by('-date')
    
    context = {
        'dates': dates,
        'search_query': search_query,
    }
    return render(request, 'dailyactivity_app/it_tanggal_project.html', context)

@login_required
def it_data_project(request, date=None):
    if date:
        # Konversi tanggal dari string ke objek date
        if isinstance(date, str):
            try:
                # Parse tanggal dari format YYYY-MM-DD
                date_obj = datetime.strptime(date, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, f'Format tanggal tidak valid: {date}')
                return redirect('dailyactivity_app:it_tanggal_project')
        else:
            date_obj = date
        
        # Filter project berdasarkan ProjectIssue dengan due_date yang sesuai
        project_issues = ProjectIssue.objects.filter(due_date=date_obj).select_related('project')
    else:
        # Jika tidak ada tanggal, tampilkan semua project issue
        project_issues = ProjectIssue.objects.all().order_by('-due_date').select_related('project')
    
    # Menangani duplikasi project_name
    unique_projects = {}
    filtered_issues = []
    
    for issue in project_issues:
        project_name = issue.project.project_name
        if project_name not in unique_projects:
            unique_projects[project_name] = issue
            filtered_issues.append(issue)
    
    context = {
        'project_issues': filtered_issues,
        'selected_date': date_obj if date else None
    }
    return render(request, 'dailyactivity_app/it_data_project.html', context)

@login_required
def it_project_detail(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    issues = project.issues.all()

    if request.method == 'POST':
        issue_keys = [
            key for key in request.POST
            if key.startswith('issue_') and not key.endswith(('pic', 'due', 'status', 'remark'))
        ]

        for key in issue_keys:
            issue_index = key.split('_')[1]
            issue_text = request.POST.get(key, "").strip()
            if not issue_text:
                continue  # Lewati jika issue kosong

            # Ambil input terkait berdasarkan index
            pic_text = request.POST.get(f'issue_pic_{issue_index}', "").strip() or None
            due_date = request.POST.get(f'issue_due_{issue_index}', "").strip()
            # status = request.POST.get(f'issue_status_{issue_index}', "Pending").strip()
            status = request.POST.get(f'issue_status_{issue_index}', "0") + "%"
            remark = request.POST.get(f'issue_remark_{issue_index}', "").strip() or None
            # Format tanggal jika kosong
            if not due_date:
                due_date = None
            try:
                ProjectIssue.objects.create(
                    project=project,
                    issue=issue_text,
                    pic=pic_text,
                    due_date=due_date,
                    status=status,
                    remark=remark,
                )
            except Exception as e:
                # Tambahkan logging atau flash message untuk menangani error
                print(f"Error saat menyimpan issue: {e}")
                # Anda dapat menggunakan messages framework untuk menampilkan error di UI
        # Redirect untuk mencegah form resubmission
        return redirect('dailyactivity_app:it_project_detail', project_id=project_id)

    context = {
        'project': project,
        'issues': issues,
    }
    return render(request, 'dailyactivity_app/it_project_detail.html', context)

@login_required
def submit_issues(request):
    if request.method == 'POST':
        print("Received POST data:", request.POST)  # Cetak semua data POST
        project_id = request.POST.get('project_id')
        print("Project ID:", project_id)

        if not project_id:
            messages.error(request, "Project ID is missing!")
            return redirect('dailyactivity_app:it_project')

        try:
            project = Project.objects.get(id=project_id)
            print("Project found:", project)
        except Project.DoesNotExist:
            messages.error(request, "Project not found!")
            return redirect('dailyactivity_app:it_project')

        issues = []
        for key in request.POST:
            if key.startswith('issue_'):
                index = key.split('_')[1]
                issue = request.POST.get(f'issue_{index}', '').strip()
                pic = request.POST.get(f'issue_pic_{index}', '').strip()
                due_date = request.POST.get(f'issue_due_{index}', '').strip()
                status = request.POST.get(f'issue_status_{index}', "0") + "%"
                remark = request.POST.get(f'issue_remark_{index}', '').strip()

                if issue:
                    issues.append({
                        'issue': issue,
                        'pic': pic,
                        'due_date': due_date,
                        'status': status,
                        'remark': remark,
                    })
                    print(f"Issue {index}: {issue}, PIC: {pic}, Due Date: {due_date}, Status: {status}, Remark: {remark}")

        for issue_data in issues:
            ProjectIssue.objects.create(
                project=project,
                issue=issue_data['issue'],
                pic=issue_data['pic'],
                due_date=issue_data['due_date'],
                status=issue_data['status'],
                remark=issue_data['remark'],
            )
            print(f"Issue saved: {issue_data['issue']}")

        messages.success(request, "Issues submitted successfully!")
        return redirect('dailyactivity_app:it_project_detail', project_id=project.id)

    return render(request, 'dailyactivity_app/it_project.html')

@login_required
def export_excel(request, project_id):
    # Ambil project berdasarkan ID
    project = get_object_or_404(Project, id=project_id)

    # Buat workbook dan sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Monitoring IT Project'

    # Set font dan alignment untuk judul
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    alignment_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    # Tambahkan judul
    sheet.merge_cells("A1:F1")
    sheet["A1"] = f"Monitoring IT Project - {project.project_name}"
    sheet["A1"].font = title_font
    sheet["A1"].alignment = alignment_center

    # Detail proyek
    details = [
        ("Name of project", project.project_name),
        ("PIC Project", project.pic_project),
        ("Dept", project.department),
        ("Start project", project.start_date.strftime('%b-%y') if project.start_date else "N/A"),
        ("Finish Project", project.finish_date.strftime('%b-%y') if project.finish_date else "N/A"),
    ]
    # Tambahkan detail proyek ke sheet
    row = 3
    for detail in details:
        sheet[f"A{row}"] = detail[0]
        sheet[f"B{row}"] = f": {detail[1]}" 
        sheet[f"A{row}"].font = bold_font
        row += 1

    # Tambahkan header tabel
    headers = ["NO", "Issue", "PIC", "Due Date", "Status", "Remark"]
    header_row = row + 1
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col_num, value=header)
        cell.font = bold_font
        cell.alignment = alignment_center
        cell.border = thin_border

    # Tambahkan data tabel
    data_start_row = header_row + 1
    current_row = data_start_row
    row_num = 1

    # Ambil issues untuk project spesifik
    issues = ProjectIssue.objects.filter(project=project)
    for issue in issues:
        sheet.cell(row=current_row, column=1, value=row_num).alignment = alignment_center
        sheet.cell(row=current_row, column=1).border = thin_border
        sheet.cell(row=current_row, column=2, value=issue.issue).border = thin_border
        sheet.cell(row=current_row, column=3, value=issue.pic).border = thin_border
        sheet.cell(row=current_row, column=4, value=issue.due_date.strftime('%Y-%m-%d') if issue.due_date else "").border = thin_border
        sheet.cell(row=current_row, column=5, value=issue.status).border = thin_border
        sheet.cell(row=current_row, column=6, value=issue.remark).border = thin_border

        current_row += 1
        row_num += 1

    # Atur lebar kolom
    column_widths = [5, 50, 15, 15, 15, 20]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Konfigurasi response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Monitoring_IT_Project_{project.project_name}.xlsx"'

    # Simpan workbook ke response
    workbook.save(response)
    return response

def report_data_it(request):
    # Inisialisasi projects sebagai QuerySet kosong
    projects = Project.objects.none()
    filtered = False
    
    # Ambil tanggal dari form filter jika ada
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Filter data hanya jika kedua tanggal disediakan
    if start_date and end_date:
        filtered = True
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            # Query untuk mendapatkan projects dalam rentang tanggal
            projects = Project.objects.filter(
                start_date__gte=start_date_obj, 
                start_date__lte=end_date_obj
            ).order_by('start_date')
            
        except ValueError as e:
            messages.error(request, f'Format tanggal tidak valid: {e}')
            projects = Project.objects.none()
        
        # Export ke Excel jika tombol export ditekan
        if 'export' in request.GET:
            return export_to_excel(projects, start_date, end_date)
    
    # Context untuk template
    context = {
        'projects': projects,
        'start_date': start_date,
        'end_date': end_date,
        'filtered': filtered
    }
    
    return render(request, 'dailyactivity_app/report_data_it.html', context)

def export_to_excel(projects, start_date, end_date):
    # Buat response dengan content type Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="IT_Report_{start_date}_to_{end_date}.xls"'
    
    # Buat workbook Excel
    wb = xlwt.Workbook(encoding='utf-8')
    
    # Style untuk header
    header_style = xlwt.easyxf('font: bold on; pattern: pattern solid, fore_color gray25;')
    project_style = xlwt.easyxf('font: bold on; pattern: pattern solid, fore_color light_green;')
    issue_style = xlwt.easyxf('pattern: pattern solid, fore_color light_yellow;')
    
    # Sheet untuk report gabungan
    ws = wb.add_sheet('IT Report')
    
    # Tulis header untuk report
    row_num = 0
    headers = ['Project ID', 'Project Name', 'PIC Project', 'Department', 'Start Date', 'Finish Date', 
               'Issue ID', 'Issue Description', 'PIC Issue', 'Due Date', 'Status', 'Remark']
    
    for col_num, header in enumerate(headers):
        ws.write(row_num, col_num, header, header_style)
    
    # Tulis data projects dan issues
    row_num = 1
    for project in projects:
        # Ambil issues untuk project ini
        issues = project.issues.all()
        
        if issues.exists():
            # Jika project memiliki issues, tulis project dan issue bersama
            for issue in issues:
                # Data project
                ws.write(row_num, 0, f'P{project.id}', project_style)
                ws.write(row_num, 1, project.project_name, project_style)
                ws.write(row_num, 2, project.pic_project, project_style)
                ws.write(row_num, 3, project.department, project_style)
                ws.write(row_num, 4, project.start_date.strftime('%Y-%m-%d'), project_style)
                ws.write(row_num, 5, project.finish_date.strftime('%Y-%m-%d'), project_style)
                
                # Data issue
                ws.write(row_num, 6, f'I{issue.id}', issue_style)
                ws.write(row_num, 7, issue.issue, issue_style)
                ws.write(row_num, 8, issue.pic, issue_style)
                ws.write(row_num, 9, issue.due_date.strftime('%Y-%m-%d'), issue_style)
                ws.write(row_num, 10, issue.status + '%' if issue.status else '0%', issue_style)
                ws.write(row_num, 11, issue.remark or '', issue_style)
                
                row_num += 1
        else:
            # Jika project tidak memiliki issues, tulis hanya project
            ws.write(row_num, 0, f'P{project.id}', project_style)
            ws.write(row_num, 1, project.project_name, project_style)
            ws.write(row_num, 2, project.pic_project, project_style)
            ws.write(row_num, 3, project.department, project_style)
            ws.write(row_num, 4, project.start_date.strftime('%Y-%m-%d'), project_style)
            ws.write(row_num, 5, project.finish_date.strftime('%Y-%m-%d'), project_style)
            
            # Isi kolom issue dengan kosong
            for i in range(6, 12):
                ws.write(row_num, i, '', issue_style)
                
            row_num += 1
    
    # Atur lebar kolom
    for i in range(12):
        if i in [1, 7, 11]:  # Kolom dengan teks panjang
            ws.col(i).width = 256 * 40  # 40 karakter
        else:
            ws.col(i).width = 256 * 15  # 15 karakter
    
    wb.save(response)
    return response


@login_required
def delete_issue(request, issue_id):
    issue = get_object_or_404(ProjectIssue, id=issue_id)
    issue.delete()  # Menghapus issue dari database
    return redirect('dailyactivity_app:it_project_detail', project_id=issue.project.id)  

def delete_project(request, project_id):
    if request.method == 'POST':
        project = get_object_or_404(Project, id=project_id)
        project.delete()
        return JsonResponse({'success': True, 'message': 'Project deleted successfully'})
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)